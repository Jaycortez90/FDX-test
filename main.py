import json
import asyncio
import os
import time
import math
import urllib.parse
import urllib.request
import re
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

from fastapi import FastAPI, HTTPException, Query, Request, Body
from fastapi.responses import HTMLResponse, Response, FileResponse
from fastapi.staticfiles import StaticFiles

try:
    from pywebpush import webpush
    _PUSH_OK = True
except Exception:
    _PUSH_OK = False

try:
    from dateutil import parser as dtparser
    _DATEUTIL_OK = True
except Exception:
    _DATEUTIL_OK = False

try:
    import openpyxl  # type: ignore
    _OPENPYXL_OK = True
except Exception:
    openpyxl = None  # type: ignore
    _OPENPYXL_OK = False


app = FastAPI(title="Driver Status")

# =============================
# Paths (data + static)
# =============================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
STATIC_DIR = os.path.join(BASE_DIR, "static")

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(STATIC_DIR, exist_ok=True)

# Background image path used by the website:
# Put your uploaded image into: static/bg.png
# (same folder as this main.py, inside "static")
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

# Excel lookup files (server-side destination calculation)
# Optional env overrides:
#   - LOCATIONS_XLSX: path to FedEx_locations.xlsx / .xlsm
#   - DEST_LAND_XLSX: path to dest-land.xlsx / .xlsm
LOCATIONS_XLSX_ENV = os.environ.get("LOCATIONS_XLSX", "").strip()
DEST_LAND_XLSX_ENV = os.environ.get("DEST_LAND_XLSX", "").strip()


def _pick_existing_path(candidates: List[str]) -> str:
    first_non_empty = ""
    for p in candidates:
        p = (p or "").strip()
        if p and not first_non_empty:
            first_non_empty = p
        if p and os.path.exists(p):
            return p
    return first_non_empty

def _locations_path() -> str:
    # Prefer env var, otherwise try common locations
    candidates = [
        LOCATIONS_XLSX_ENV,
        os.path.join(DATA_DIR, "FedEx_locations.xlsx"),
        os.path.join(DATA_DIR, "FedEx_locations.xlsm"),
        os.path.join(BASE_DIR, "FedEx_locations.xlsx"),
        os.path.join(os.getcwd(), "data", "FedEx_locations.xlsx"),
        os.path.join(os.getcwd(), "FedEx_locations.xlsx"),
    ]
    return _pick_existing_path(candidates)


def _dest_land_path() -> str:
    candidates = [
        DEST_LAND_XLSX_ENV,
        os.path.join(DATA_DIR, "dest-land.xlsx"),
        os.path.join(DATA_DIR, "dest-land.xlsm"),
        os.path.join(DATA_DIR, "dest_land.xlsx"),
        os.path.join(DATA_DIR, "dest_land.xlsm"),
        os.path.join(BASE_DIR, "dest-land.xlsx"),
        os.path.join(BASE_DIR, "dest-land.xlsm"),
        os.path.join(os.getcwd(), "data", "dest-land.xlsx"),
        os.path.join(os.getcwd(), "dest-land.xlsx"),
    ]
    return _pick_existing_path(candidates)


LOCATIONS_XLSX = _locations_path()
DEST_LAND_XLSX = _dest_land_path()

# Loaded at startup
LOCATION_BY_CODE: Dict[str, Dict[str, Any]] = {}
DESTLAND_BY_CODE: Dict[str, Dict[str, Any]] = {}

# =============================
# Geofence (QAR Duiven) - still enforced, but NOT displayed on website
# =============================
HUB_NAME = "QAR Duiven"
HUB_LAT = 51.9672245
HUB_LON = 6.0205411
GEOFENCE_RADIUS_KM = 30.0
MAX_LOCATION_AGE_SECONDS = 120

# =============================
# Upload secret (required for desktop uploads)
# =============================
ADMIN_UPLOAD_SECRET = os.environ.get("ADMIN_UPLOAD_SECRET", "").strip()

# =============================
# Routing (optional) - OpenRouteService (truck route geometry)
# =============================
ORS_API_KEY = os.environ.get("ORS_API_KEY", "").strip()
ORS_DIRECTIONS_URL = "https://api.openrouteservice.org/v2/directions/driving-hgv/geojson"

# =============================
# Live Traffic (optional) - HERE Routing v8 (server-side)
# =============================
HERE_API_KEY = os.environ.get("HERE_API_KEY", "").strip()
HERE_ROUTING_URL = "https://router.hereapi.com/v8/routes"

# In-memory cache for traffic delay (Render restarts will clear these)
_TRAFFIC_CACHE: Dict[Tuple[float, float, float, float, str], Tuple[float, Dict[str, Any]]] = {}
_TRAFFIC_TTL_SEC = 90

# =============================
# Web Push (optional)
# =============================
VAPID_PUBLIC_KEY = os.environ.get("VAPID_PUBLIC_KEY", "").strip()
VAPID_PRIVATE_KEY = os.environ.get("VAPID_PRIVATE_KEY", "").strip()
VAPID_SUBJECT = os.environ.get("VAPID_SUBJECT", "mailto:admin@example.com").strip()
PUSH_ENABLED = bool(_PUSH_OK and VAPID_PUBLIC_KEY and VAPID_PRIVATE_KEY)

# =============================
# In-memory stores (Render restarts will clear these)
# =============================
SNAPSHOT: Optional[Dict[str, Any]] = None
LAST_STATUS_KEY_BY_PLATE: Dict[str, str] = {}
SUBSCRIPTIONS_BY_PLATE: Dict[str, List[Dict[str, Any]]] = {}
MANUAL_STATUS_BY_PLATE: Dict[str, str] = {}
VIEWED_BY_PLATE: Dict[str, Dict[str, Any]] = {}  # plate -> {count:int, last_view:str}
HOUSE_RULES_ACCEPTED_BY_PLATE: Dict[str, str] = {}  # plate -> ISO timestamp (in-memory, resets on restart)
STATUS_POLL_INTERVAL_SECONDS = 30

# =============================
# Developer monitor (in-memory)
# =============================
DEV_PLATE = "KLETH743"  # developer window "password" (normalized license plate)
CHECK_LOG_WINDOW_HOURS = 12
CHECK_LOG_WINDOW_SECONDS = CHECK_LOG_WINDOW_HOURS * 3600

# List of plate check events in the last ~12 hours (pruned continuously)
# Each item: {"plate": "AB123CD", "ts": 1700000000, "iso": "2026-02-28T12:34:56Z"}
CHECK_LOG: List[Dict[str, Any]] = []

# Admin monitor toggle (requires DEV_PLATE in UI; not cryptographically secure)
ADMIN_NOTIFY_ENABLED: bool = False
ADMIN_NOTIFY_CHANGED_AT: str = ""

# Simple de-duplication to avoid spamming admin on frequent refreshes
LAST_ADMIN_CHECK_PUSH_TS_BY_PLATE: Dict[str, float] = {}
ADMIN_CHECK_PUSH_DEDUP_SECONDS = 60



# -----------------------------
# Startup
# -----------------------------
@app.on_event("startup")
async def _startup():
    _load_destination_lookups()

    # Periodically re-evaluate statuses so time-based changes (45 min threshold)
    # can trigger push even without new uploads.
    if not PUSH_ENABLED:
        return

    async def _loop():
        global SNAPSHOT
        while True:
            try:
                if SNAPSHOT:
                    for m in _snapshot_movements():
                        plate = normalize_plate(m.get("license_plate", ""))
                        if not plate:
                            continue
                        st = compute_driver_status(m)
                        new_key = st["status_key"]
                        old_key = LAST_STATUS_KEY_BY_PLATE.get(plate)
                        if old_key is None:
                            LAST_STATUS_KEY_BY_PLATE[plate] = new_key
                            continue
                        if new_key != old_key:
                            LAST_STATUS_KEY_BY_PLATE[plate] = new_key
                            _push_status_change_to_plate(plate, m)
                            _maybe_admin_push_status_change(plate, m)
            except Exception:
                pass
            await asyncio.sleep(STATUS_POLL_INTERVAL_SECONDS)

    asyncio.create_task(_loop())


# -----------------------------
# Helpers
# -----------------------------
def normalize_plate(value: str) -> str:
    v = (value or "").upper().strip()
    v = v.replace(" ", "").replace("-", "")
    return v


def _norm_code(value: Any) -> str:
    s = str(value or "").strip().upper()
    s = s.replace(" ", "").replace("-", "")
    return s


def haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    r = 6371.0
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)

    a = math.sin(dphi / 2.0) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2.0) ** 2
    c = 2.0 * math.atan2(math.sqrt(a), math.sqrt(1.0 - a))
    return r * c


def geofence_check(lat: float, lon: float, ts: int) -> None:
    now = int(time.time())
    if abs(now - int(ts)) > MAX_LOCATION_AGE_SECONDS:
        raise HTTPException(status_code=401, detail="Location timestamp too old. Refresh and try again.")

    dist = haversine_km(float(lat), float(lon), HUB_LAT, HUB_LON)
    if dist > float(GEOFENCE_RADIUS_KM):
        raise HTTPException(status_code=403, detail=f"Access denied (outside {GEOFENCE_RADIUS_KM:.0f} km of {HUB_NAME}).")


def _parse_dt(val: Any) -> Optional[datetime]:
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in {"nan", "none", "nat"}:
        return None

    try:
        if s.endswith("Z"):
            return datetime.fromisoformat(s[:-1])
        return datetime.fromisoformat(s)
    except Exception:
        pass

    if _DATEUTIL_OK:
        try:
            return dtparser.parse(s, dayfirst=True, fuzzy=True)
        except Exception:
            return None

    return None

def _format_dt_like(dt: datetime, sample: Any) -> str:
    """Format dt to match the date/time style of sample (scheduled_departure string)."""
    try:
        s = str(sample or "").strip()
    except Exception:
        s = ""

    # Default (ISO-like)
    date_fmt = "%Y-%m-%d"
    sep = " "

    # Preserve separator if sample uses 'T'
    if "T" in s and " " not in s:
        sep = "T"

    # Pick date format based on sample
    if re.search(r"\b\d{2}\.\d{2}\.\d{4}\b", s):
        date_fmt = "%d.%m.%Y"
    elif re.search(r"\b\d{2}/\d{2}/\d{4}\b", s):
        date_fmt = "%d/%m/%Y"
    elif re.search(r"\b\d{2}-\d{2}-\d{4}\b", s):
        date_fmt = "%d-%m-%Y"
    elif re.search(r"\b\d{4}/\d{2}/\d{2}\b", s):
        date_fmt = "%Y/%m/%d"
    elif re.search(r"\b\d{4}\.\d{2}\.\d{2}\b", s):
        date_fmt = "%Y.%m.%d"
    elif re.search(r"\b\d{4}-\d{2}-\d{2}\b", s):
        date_fmt = "%Y-%m-%d"

    has_seconds = bool(re.search(r":\d{2}:\d{2}(?!\d)", s))
    time_fmt = "%H:%M:%S" if has_seconds else "%H:%M"

    try:
        return dt.strftime(f"{date_fmt}{sep}{time_fmt}")
    except Exception:
        return dt.strftime("%Y-%m-%d %H:%M")


def _format_scheduled_departure(sched_raw: Any) -> str:
    """Format scheduled_departure so it matches report_in_office_at style."""
    dt = _parse_dt(sched_raw)
    if dt:
        return _format_dt_like(dt, sched_raw)
    try:
        return str(sched_raw or "").strip()
    except Exception:
        return ""


def _has(v: Any) -> bool:
    s = str(v or "").strip()
    return bool(s) and s.lower() not in {"nan", "none", "nat"}


def _clean_location_value(v: Any) -> str:
    s = str(v or "").strip()
    if not s:
        return ""
    if s.lower() == "wait":
        return ""
    return s


def _traffic_cache_get(key: Tuple[float, float, float, float, str]) -> Optional[Dict[str, Any]]:
    item = _TRAFFIC_CACHE.get(key)
    if not item:
        return None
    ts, payload = item
    if (time.time() - float(ts)) > float(_TRAFFIC_TTL_SEC):
        _TRAFFIC_CACHE.pop(key, None)
        return None
    return payload


def _traffic_cache_set(key: Tuple[float, float, float, float, str], payload: Dict[str, Any]) -> None:
    _TRAFFIC_CACHE[key] = (time.time(), payload)


def _here_fetch_delay_minutes(
    origin_lat: float,
    origin_lon: float,
    dest_lat: float,
    dest_lon: float,
    departure_iso: str = "",
) -> Tuple[Optional[int], Optional[str]]:
    """Return (delay_minutes, error). delay_minutes is >=0 on success."""
    key = (HERE_API_KEY or "").strip()
    if not key:
        return None, "HERE_API_KEY missing on server"

    params = {
        "transportMode": "truck",
        "origin": f"{origin_lat:.6f},{origin_lon:.6f}",
        "destination": f"{dest_lat:.6f},{dest_lon:.6f}",
        "return": "summary",
        "apiKey": key,
    }
    if departure_iso:
        params["departureTime"] = departure_iso

    try:
        qs = urllib.parse.urlencode(params)
        url = f"{HERE_ROUTING_URL}?{qs}"
        req = urllib.request.Request(
            url,
            headers={"Accept": "application/json", "User-Agent": "DriverStatus/TrafficDelay"},
            method="GET",
        )
        with urllib.request.urlopen(req, timeout=8) as resp:
            raw = resp.read()

        data = json.loads(raw.decode("utf-8", errors="ignore") or "{}")
        routes = data.get("routes") or []
        if not routes:
            return None, "HERE: no routes"
        sections = routes[0].get("sections") or []
        if not sections:
            return None, "HERE: no sections"
        summary = sections[0].get("summary") or {}

        base_sec = summary.get("baseDuration")
        traffic_sec = summary.get("duration")
        if base_sec is None or traffic_sec is None:
            return None, "HERE: missing duration fields"

        delay_sec = max(0, int(traffic_sec) - int(base_sec))
        delay_min = int(round(delay_sec / 60.0))
        if delay_min < 0:
            delay_min = 0
        return delay_min, None
    except Exception as e:
        # If the request is rejected, HERE often returns JSON with 'title'/'message',
        # but urllib raises HTTPError; keep it simple.
        return None, f"HERE request failed ({type(e).__name__})"

SUPPORTED_LANGS = {"en", "de", "nl", "fr", "tr", "sv", "es", "it", "ro", "ru", "lt", "kk", "hi", "pl", "hu", "uz", "tg", "ky", "be"}

def normalize_lang(value: Any) -> str:
    """Return one of: en, de, nl, fr, tr, sv, es, it, ro, ru, lt, kk, hi, pl, hu, uz, tg, ky, be."""
    s = str(value or "").strip().lower()
    if not s:
        return "en"
    # normalize common forms: en-US, de_DE, etc.
    s = s.replace("_", "-")
    base = s.split("-", 1)[0]
    if base in SUPPORTED_LANGS:
        return base
    # allow common aliases
    if base in {"eng"}:
        return "en"
    if base in {"ger", "deu"}:
        return "de"
    if base in {"dut", "nld"}:
        return "nl"
    if base in {"fre", "fra"}:
        return "fr"
    if base in {"tur"}:
        return "tr"
    if base in {"swe"}:
        return "sv"
    if base in {"rus"}:
        return "ru"
    if base in {"lit"}:
        return "lt"
    if base in {"kaz", "kz"}:
        return "kk"
    if base in {"hin"}:
        return "hi"
    if base in {"pol"}:
        return "pl"
    if base in {"hun"}:
        return "hu"
    if base in {"uzb"}:
        return "uz"
    if base in {"tgk", "taj", "tj"}:
        return "tg"
    if base in {"kir", "kg"}:
        return "ky"
    if base in {"bel", "by"}:
        return "be"

    if base in {"spa", "esp"}:
        return "es"
    if base in {"ita"}:
        return "it"
    if base in {"rom", "ron", "rum"}:
        return "ro"

    return "en"


_I18N_STATUS: Dict[str, Dict[str, str]] = {
    "en": {
        "DEPARTED": "Drive safe, we wait you back!",
        "LOCATION_WITH_TRAILER": "Please connect the {trailer} trailer on location: {location} and pick up the CMR documents in the office!",
        "LOCATION_NO_TRAILER": "Please connect the trailer on location: {location} and pick up the CMR documents in the office!",
        "CLOSEDOOR_NO_LOCATION": "Your trailer is ready, please report in the office for further information!",
        "LOADING_WAIT": "Your trailer is being loaded, please wait!",
        "REPORT_OFFICE": "Please report in the office!",
    },
    "de": {
        "DEPARTED": "Fahren Sie vorsichtig – wir erwarten Sie zurück!",
        "LOCATION_WITH_TRAILER": "Bitte koppeln Sie den Anhänger {trailer} am Standort: {location} an und holen Sie die CMR-Dokumente im Büro ab!",
        "LOCATION_NO_TRAILER": "Bitte koppeln Sie den Anhänger am Standort: {location} an und holen Sie die CMR-Dokumente im Büro ab!",
        "CLOSEDOOR_NO_LOCATION": "Ihr Anhänger ist fertig. Bitte melden Sie sich im Büro für weitere Informationen!",
        "LOADING_WAIT": "Ihr Anhänger wird beladen – bitte warten!",
        "REPORT_OFFICE": "Bitte melden Sie sich im Büro!",
    },
    "nl": {
        "DEPARTED": "Rij veilig – we wachten op je terugkeer!",
        "LOCATION_WITH_TRAILER": "Koppel de trailer {trailer} op locatie: {location} en haal de CMR-documenten op in het kantoor!",
        "LOCATION_NO_TRAILER": "Koppel de trailer op locatie: {location} en haal de CMR-documenten op in het kantoor!",
        "CLOSEDOOR_NO_LOCATION": "Je trailer is gereed. Meld je in het kantoor voor verdere informatie!",
        "LOADING_WAIT": "Je trailer wordt geladen – even wachten!",
        "REPORT_OFFICE": "Meld je in het kantoor!",
    },
    "fr": {
        "DEPARTED": "Bonne route, on vous attend de retour !",
        "LOCATION_WITH_TRAILER": "Veuillez atteler la remorque {trailer} à l’emplacement : {location} et récupérer les documents CMR au bureau !",
        "LOCATION_NO_TRAILER": "Veuillez atteler la remorque à l’emplacement : {location} et récupérer les documents CMR au bureau !",
        "CLOSEDOOR_NO_LOCATION": "Votre remorque est prête. Veuillez vous présenter au bureau pour plus d’informations !",
        "LOADING_WAIT": "Votre remorque est en cours de chargement — veuillez patienter !",
        "REPORT_OFFICE": "Veuillez vous présenter au bureau !",
    },
    "tr": {
        "DEPARTED": "İyi yolculuklar, geri dönmenizi bekliyoruz!",
        "LOCATION_WITH_TRAILER": "Lütfen {location} konumunda {trailer} dorsesini bağlayın ve CMR belgelerini ofisten alın!",
        "LOCATION_NO_TRAILER": "Lütfen {location} konumunda dorsenizi bağlayın ve CMR belgelerini ofisten alın!",
        "CLOSEDOOR_NO_LOCATION": "Dorseniz hazır. Daha fazla bilgi için lütfen ofise başvurun!",
        "LOADING_WAIT": "Dorseniz yükleniyor — lütfen bekleyin!",
        "REPORT_OFFICE": "Lütfen ofise başvurun!",
    },
    "sv": {
        "DEPARTED": "Kör försiktigt – vi väntar på att du kommer tillbaka!",
        "LOCATION_WITH_TRAILER": "Koppla släpet {trailer} på plats: {location} och hämta CMR-dokumenten på kontoret!",
        "LOCATION_NO_TRAILER": "Koppla släpet på plats: {location} och hämta CMR-dokumenten på kontoret!",
        "CLOSEDOOR_NO_LOCATION": "Ditt släp är klart. Anmäl dig på kontoret för mer information!",
        "LOADING_WAIT": "Ditt släp lastas — vänligen vänta!",
        "REPORT_OFFICE": "Anmäl dig på kontoret!",
    },


    "es": {
        "DEPARTED": "Buen viaje — ¡te esperamos de vuelta!",
        "LOCATION_WITH_TRAILER": "Por favor conecta el remolque {trailer} en la ubicación: {location} y recoge los documentos CMR en la oficina!",
        "LOCATION_NO_TRAILER": "Por favor conecta el remolque en la ubicación: {location} y recoge los documentos CMR en la oficina!",
        "CLOSEDOOR_NO_LOCATION": "Tu remolque está listo. Preséntate en la oficina para más información!",
        "LOADING_WAIT": "Tu remolque se está cargando — espera, por favor!",
        "REPORT_OFFICE": "¡Preséntate en la oficina!",
    },
    "it": {
        "DEPARTED": "Buon viaggio — ti aspettiamo di nuovo!",
        "LOCATION_WITH_TRAILER": "Collega il rimorchio {trailer} alla posizione: {location} e ritira i documenti CMR in ufficio!",
        "LOCATION_NO_TRAILER": "Collega il rimorchio alla posizione: {location} e ritira i documenti CMR in ufficio!",
        "CLOSEDOOR_NO_LOCATION": "Il tuo rimorchio è pronto. Presentati in ufficio per ulteriori informazioni!",
        "LOADING_WAIT": "Il tuo rimorchio è in carico — attendi!",
        "REPORT_OFFICE": "Presentati in ufficio!",
    },
    "ro": {
        "DEPARTED": "Drum bun — te așteptăm înapoi!",
        "LOCATION_WITH_TRAILER": "Vă rugăm să conectați remorca {trailer} la locația: {location} și să ridicați documentele CMR din birou!",
        "LOCATION_NO_TRAILER": "Vă rugăm să conectați remorca la locația: {location} și să ridicați documentele CMR din birou!",
        "CLOSEDOOR_NO_LOCATION": "Remorca ta este gata. Te rugăm să te prezinți la birou pentru informații suplimentare!",
        "LOADING_WAIT": "Remorca ta este în curs de încărcare — te rugăm să aștepți!",
        "REPORT_OFFICE": "Te rugăm să te prezinți la birou!",
    },
    "ru": {
        "DEPARTED": "Счастливого пути — ждём вас обратно!",
        "LOCATION_WITH_TRAILER": "Пожалуйста, прицепите прицеп {trailer} на месте: {location} и заберите документы CMR в офисе!",
        "LOCATION_NO_TRAILER": "Пожалуйста, прицепите прицеп на месте: {location} и заберите документы CMR в офисе!",
        "CLOSEDOOR_NO_LOCATION": "Ваш прицеп готов. Пожалуйста, подойдите в офис за дальнейшей информацией!",
        "LOADING_WAIT": "Ваш прицеп загружается — пожалуйста, подождите!",
        "REPORT_OFFICE": "Пожалуйста, подойдите в офис!",
    },
    "lt": {
        "DEPARTED": "Saugios kelionės – laukiame jūsų sugrįžtant!",
        "LOCATION_WITH_TRAILER": "Prašome prijungti priekabą {trailer} vietoje: {location} ir pasiimti CMR dokumentus biure!",
        "LOCATION_NO_TRAILER": "Prašome prijungti priekabą vietoje: {location} ir pasiimti CMR dokumentus biure!",
        "CLOSEDOOR_NO_LOCATION": "Jūsų priekaba paruošta. Prašome užsukti į biurą dėl tolimesnės informacijos!",
        "LOADING_WAIT": "Jūsų priekaba kraunama – prašome palaukti!",
        "REPORT_OFFICE": "Prašome užsukti į biurą!",
    },
    "kk": {
        "DEPARTED": "Сәтті жол — сізді қайта күтеміз!",
        "LOCATION_WITH_TRAILER": "{location} орнында {trailer} тіркемесін қосып, CMR құжаттарын кеңседен алыңыз!",
        "LOCATION_NO_TRAILER": "{location} орнында тіркемені қосып, CMR құжаттарын кеңседен алыңыз!",
        "CLOSEDOOR_NO_LOCATION": "Тіркеме дайын. Қосымша ақпарат үшін кеңсеге келіңіз!",
        "LOADING_WAIT": "Тіркеме тиелуде — күтіңіз!",
        "REPORT_OFFICE": "Кеңсеге келіңіз!",
    },
    "hi": {
        "DEPARTED": "सुरक्षित यात्रा करें — हम आपका वापस इंतज़ार करेंगे!",
        "LOCATION_WITH_TRAILER": "कृपया {location} स्थान पर ट्रेलर {trailer} जोड़ें और CMR दस्तावेज़ कार्यालय से लें!",
        "LOCATION_NO_TRAILER": "कृपया {location} स्थान पर ट्रेलर जोड़ें और CMR दस्तावेज़ कार्यालय से लें!",
        "CLOSEDOOR_NO_LOCATION": "आपका ट्रेलर तैयार है। आगे की जानकारी के लिए कृपया कार्यालय में रिपोर्ट करें!",
        "LOADING_WAIT": "आपका ट्रेलर लोड हो रहा है — कृपया प्रतीक्षा करें!",
        "REPORT_OFFICE": "कृपया कार्यालय में रिपोर्ट करें!",
    },
    "pl": {
        "DEPARTED": "Szerokiej drogi — czekamy na Twój powrót!",
        "LOCATION_WITH_TRAILER": "Proszę podpiąć naczepę {trailer} na lokalizacji: {location} i odebrać dokumenty CMR w biurze!",
        "LOCATION_NO_TRAILER": "Proszę podpiąć naczepę na lokalizacji: {location} i odebrać dokumenty CMR w biurze!",
        "CLOSEDOOR_NO_LOCATION": "Twoja naczepa jest gotowa. Proszę zgłosić się do biura po dalsze informacje!",
        "LOADING_WAIT": "Twoja naczepa jest ładowana — proszę czekać!",
        "REPORT_OFFICE": "Proszę zgłosić się do biura!",
    },
    "hu": {
        "DEPARTED": "Vezess óvatosan – várunk vissza!",
        "LOCATION_WITH_TRAILER": "Kérjük, csatlakoztasd a(z) {trailer} pótkocsit a következő helyen: {location}, és vedd fel a CMR dokumentumokat az irodában!",
        "LOCATION_NO_TRAILER": "Kérjük, csatlakoztasd a pótkocsit a következő helyen: {location}, és vedd fel a CMR dokumentumokat az irodában!",
        "CLOSEDOOR_NO_LOCATION": "A pótkocsid kész. További információért jelentkezz az irodában!",
        "LOADING_WAIT": "A pótkocsid rakodás alatt – kérjük, várj!",
        "REPORT_OFFICE": "Kérjük, jelentkezz az irodában!",
    },
    "uz": {
        "DEPARTED": "Xavfsiz haydang — sizni qaytib kelishingizni kutamiz!",
        "LOCATION_WITH_TRAILER": "Iltimos, {location} joyida {trailer} treylerini ulang va CMR hujjatlarini ofisdan oling!",
        "LOCATION_NO_TRAILER": "Iltimos, {location} joyida treylerini ulang va CMR hujjatlarini ofisdan oling!",
        "CLOSEDOOR_NO_LOCATION": "Treyleringiz tayyor. Qo‘shimcha ma’lumot uchun ofisga murojaat qiling!",
        "LOADING_WAIT": "Treyleringiz yuklanmoqda — iltimos, kuting!",
        "REPORT_OFFICE": "Iltimos, ofisga murojaat qiling!",
    },
    "tg": {
        "DEPARTED": "Сафар ба хайр — мо шуморо боз интизорем!",
        "LOCATION_WITH_TRAILER": "Лутфан прицепи {trailer}-ро дар ҷойи {location} васл кунед ва ҳуҷҷатҳои CMR-ро аз офис гиред!",
        "LOCATION_NO_TRAILER": "Лутфан прицепро дар ҷойи {location} васл кунед ва ҳуҷҷатҳои CMR-ро аз офис гиред!",
        "CLOSEDOOR_NO_LOCATION": "Прицепи шумо тайёр аст. Барои маълумоти бештар ба офис ҳозир шавед!",
        "LOADING_WAIT": "Прицепи шумо бор карда мешавад — лутфан интизор шавед!",
        "REPORT_OFFICE": "Лутфан ба офис ҳозир шавед!",
    },
    "ky": {
        "DEPARTED": "Жолуңуз болсун — кайра келишиңизди күтөбүз!",
        "LOCATION_WITH_TRAILER": "Сураныч, {location} жерде {trailer} чиркегичин туташтырып, CMR документтерин кеңседен алыңыз!",
        "LOCATION_NO_TRAILER": "Сураныч, {location} жерде чиркегичти туташтырып, CMR документтерин кеңседен алыңыз!",
        "CLOSEDOOR_NO_LOCATION": "Чиркегичиңиз даяр. Кошумча маалымат үчүн кеңсеге келиңиз!",
        "LOADING_WAIT": "Чиркегичиңиз жүктөлүүдө — сураныч, күтө туруңуз!",
        "REPORT_OFFICE": "Сураныч, кеңсеге келиңиз!",
    },
    "be": {
        "DEPARTED": "Шчаслівай дарогі — чакаем вас назад!",
        "LOCATION_WITH_TRAILER": "Калі ласка, прычапіце прычэп {trailer} у месцы: {location} і забярыце дакументы CMR у офісе!",
        "LOCATION_NO_TRAILER": "Калі ласка, прычапіце прычэп у месцы: {location} і забярыце дакументы CMR у офісе!",
        "CLOSEDOOR_NO_LOCATION": "Ваш прычэп гатовы. Калі ласка, зайдзіце ў офіс для далейшай інфармацыі!",
        "LOADING_WAIT": "Ваш прычэп загружаецца — калі ласка, пачакайце!",
        "REPORT_OFFICE": "Калі ласка, зайдзіце ў офіс!",
    },
}

_I18N_PUSH_TITLES: Dict[str, Dict[str, str]] = {
    "en": {"STATUS_UPDATE": "Status update", "MESSAGE_FROM_DISPATCH": "Message from dispatch", "ADMIN_MONITOR": "Admin monitor"},
    "de": {"STATUS_UPDATE": "Status-Update", "MESSAGE_FROM_DISPATCH": "Nachricht von der Disposition", "ADMIN_MONITOR": "Admin Monitor"},
    "nl": {"STATUS_UPDATE": "Statusupdate", "MESSAGE_FROM_DISPATCH": "Bericht van de planning", "ADMIN_MONITOR": "Admin monitor"},
    "fr": {"STATUS_UPDATE": "Mise à jour du statut", "MESSAGE_FROM_DISPATCH": "Message de la planification", "ADMIN_MONITOR": "Moniteur admin"},
    "tr": {"STATUS_UPDATE": "Durum güncellemesi", "MESSAGE_FROM_DISPATCH": "Operasyondan mesaj", "ADMIN_MONITOR": "Yönetici izleme"},
    "sv": {"STATUS_UPDATE": "Statusuppdatering", "MESSAGE_FROM_DISPATCH": "Meddelande från dispatch", "ADMIN_MONITOR": "Admin-övervakning"},


    "es": {"STATUS_UPDATE": "Actualización de estado", "MESSAGE_FROM_DISPATCH": "Mensaje del despacho"},
    "it": {"STATUS_UPDATE": "Aggiornamento stato", "MESSAGE_FROM_DISPATCH": "Messaggio dal dispatch"},
    "ro": {"STATUS_UPDATE": "Actualizare status", "MESSAGE_FROM_DISPATCH": "Mesaj de la dispecerat"},
    "ru": {"STATUS_UPDATE": "Обновление статуса", "MESSAGE_FROM_DISPATCH": "Сообщение от диспетчера"},
    "lt": {"STATUS_UPDATE": "Būsenos atnaujinimas", "MESSAGE_FROM_DISPATCH": "Žinutė iš dispečerio"},
    "kk": {"STATUS_UPDATE": "Күй жаңартуы", "MESSAGE_FROM_DISPATCH": "Диспетчерден хабарлама"},
    "hi": {"STATUS_UPDATE": "स्थिति अपडेट", "MESSAGE_FROM_DISPATCH": "डिस्पैच से संदेश"},
    "pl": {"STATUS_UPDATE": "Aktualizacja statusu", "MESSAGE_FROM_DISPATCH": "Wiadomość od dyspozytora"},
    "hu": {"STATUS_UPDATE": "Státusz frissítés", "MESSAGE_FROM_DISPATCH": "Üzenet a diszpécsertől"},
    "uz": {"STATUS_UPDATE": "Holat yangilanishi", "MESSAGE_FROM_DISPATCH": "Dispetcherdan xabar"},
    "tg": {"STATUS_UPDATE": "Навсозии ҳолат", "MESSAGE_FROM_DISPATCH": "Паём аз диспетчер"},
    "ky": {"STATUS_UPDATE": "Абалды жаңыртуу", "MESSAGE_FROM_DISPATCH": "Диспетчерден билдирүү"},
    "be": {"STATUS_UPDATE": "Абнаўленне статусу", "MESSAGE_FROM_DISPATCH": "Паведамленне ад дыспетчара"},
}

_I18N_ROUTE_NOTE: Dict[str, Dict[str, str]] = {
    "en": {"ORS": "Route source: OpenRouteService", "OSRM": "Route source: OSRM", "DIRECT": "Route source: direct line"},
    "de": {"ORS": "Routenquelle: OpenRouteService", "OSRM": "Routenquelle: OSRM", "DIRECT": "Routenquelle: direkte Linie"},
    "nl": {"ORS": "Routebron: OpenRouteService", "OSRM": "Routebron: OSRM", "DIRECT": "Routebron: rechte lijn"},
    "fr": {"ORS": "Source d’itinéraire : OpenRouteService", "OSRM": "Source d’itinéraire : OSRM", "DIRECT": "Source d’itinéraire : ligne directe"},
    "tr": {"ORS": "Rota kaynağı: OpenRouteService", "OSRM": "Rota kaynağı: OSRM", "DIRECT": "Rota kaynağı: doğrudan çizgi"},
    "sv": {"ORS": "Ruttkälla: OpenRouteService", "OSRM": "Ruttkälla: OSRM", "DIRECT": "Ruttkälla: rak linje"},


    "es": {"ORS": "Fuente de ruta: OpenRouteService", "OSRM": "Fuente de ruta: OSRM", "DIRECT": "Fuente de ruta: línea directa"},
    "it": {"ORS": "Fonte percorso: OpenRouteService", "OSRM": "Fonte percorso: OSRM", "DIRECT": "Fonte percorso: linea diretta"},
    "ro": {"ORS": "Sursa rutei: OpenRouteService", "OSRM": "Sursa rutei: OSRM", "DIRECT": "Sursa rutei: linie directă"},
    "ru": {"ORS": "Источник маршрута: OpenRouteService", "OSRM": "Источник маршрута: OSRM", "DIRECT": "Источник маршрута: прямая линия"},
    "lt": {"ORS": "Maršruto šaltinis: OpenRouteService", "OSRM": "Maršruto šaltinis: OSRM", "DIRECT": "Maršruto šaltinis: tiesi linija"},
    "kk": {"ORS": "Маршрут көзі: OpenRouteService", "OSRM": "Маршрут көзі: OSRM", "DIRECT": "Маршрут көзі: түзу сызық"},
    "hi": {"ORS": "मार्ग स्रोत: OpenRouteService", "OSRM": "मार्ग स्रोत: OSRM", "DIRECT": "मार्ग स्रोत: सीधी रेखा"},
    "pl": {"ORS": "Źródło trasy: OpenRouteService", "OSRM": "Źródło trasy: OSRM", "DIRECT": "Źródło trasy: linia prosta"},
    "hu": {"ORS": "Útvonal forrása: OpenRouteService", "OSRM": "Útvonal forrása: OSRM", "DIRECT": "Útvonal forrása: egyenes vonal"},
    "uz": {"ORS": "Marshrut manbai: OpenRouteService", "OSRM": "Marshrut manbai: OSRM", "DIRECT": "Marshrut manbai: to‘g‘ri chiziq"},
    "tg": {"ORS": "Манбаи масир: OpenRouteService", "OSRM": "Манбаи масир: OSRM", "DIRECT": "Манбаи масир: хатти рост"},
    "ky": {"ORS": "Маршрут булагы: OpenRouteService", "OSRM": "Маршрут булагы: OSRM", "DIRECT": "Маршрут булагы: түз сызык"},
    "be": {"ORS": "Крыніца маршруту: OpenRouteService", "OSRM": "Крыніца маршруту: OSRM", "DIRECT": "Крыніца маршруту: прамая лінія"},
}

def route_note_text(route_key: str, lang: str = "en") -> str:
    l = normalize_lang(lang)
    rk = str(route_key or "").strip().upper()
    table = _I18N_ROUTE_NOTE.get(l) or _I18N_ROUTE_NOTE["en"]
    return table.get(rk, _I18N_ROUTE_NOTE["en"].get(rk, ""))


def push_title_text(title_key: str, lang: str = "en") -> str:
    l = normalize_lang(lang)
    tk = str(title_key or "").strip().upper()
    table = _I18N_PUSH_TITLES.get(l) or _I18N_PUSH_TITLES["en"]
    return table.get(tk, _I18N_PUSH_TITLES["en"].get(tk, ""))


def compute_driver_status(m: Dict[str, Any], lang: str = "en") -> Dict[str, Any]:
    """Compute driver-facing status with localization."""
    lang_n = normalize_lang(lang)
    tmpl = _I18N_STATUS.get(lang_n) or _I18N_STATUS["en"]

    # Dispatcher manual status (Driver message) overrides computed status
    plate_n = ""
    try:
        plate_n = normalize_plate(m.get("license_plate", ""))
    except Exception:
        plate_n = ""

    msg = (MANUAL_STATUS_BY_PLATE.get(plate_n) if plate_n else "") or ""
    msg = str(msg).strip()
    if msg:
        try:
            import hashlib
            key = "driver_message:" + hashlib.sha1(msg.encode("utf-8", "ignore")).hexdigest()[:12]
        except Exception:
            key = "driver_message"
        # Manual message is NOT translated (dispatcher text)
        return {"status_key": key, "status_text": msg, "report_in_office_at": ""}

    # Departed override (after manual status)
    departed = m.get("departed", False)
    if isinstance(departed, str):
        departed = departed.strip().lower() in {"1", "true", "yes", "y"}
    if not departed:
        departed = _has(m.get("departed_at", ""))
    if departed:
        return {"status_key": "DEPARTED", "status_text": tmpl["DEPARTED"], "report_in_office_at": ""}

    close_door = m.get("close_door", "")
    location = _clean_location_value(m.get("location", ""))
    trailer = str(m.get("trailer", "") or "").strip()
    sched_raw = m.get("scheduled_departure", "")

    sched_dt = _parse_dt(sched_raw)

    if _has(location):
        if trailer:
            msg2 = tmpl["LOCATION_WITH_TRAILER"].format(trailer=trailer, location=location)
        else:
            msg2 = tmpl["LOCATION_NO_TRAILER"].format(location=location)
        key2 = "LOCATION"
    elif _has(close_door):
        msg2 = tmpl["CLOSEDOOR_NO_LOCATION"]
        key2 = "CLOSEDOOR_NO_LOCATION"
    else:
        minutes_left = None
        if sched_dt:
            minutes_left = (sched_dt - datetime.now()).total_seconds() / 60.0

        if minutes_left is not None and minutes_left > 45:
            msg2 = tmpl["LOADING_WAIT"]
            key2 = "LOADING_WAIT"
        else:
            msg2 = tmpl["REPORT_OFFICE"]
            key2 = "REPORT_OFFICE"

    report_at = ""
    if sched_dt:
        ra = sched_dt - timedelta(minutes=45)
        report_at = _format_dt_like(ra, sched_raw)

    return {
        "status_key": key2,
        "status_text": msg2,
        "report_in_office_at": report_at,
    }


def destination_nav_url(lat: Optional[float], lon: Optional[float], fallback_text: str = "") -> Optional[str]:
    """Return a Google Maps navigation URL.
    - Prefer coordinates (lat/lon) if available.
    - Fallback to destination text search if coordinates are missing.
    """
    try:
        if lat is not None and lon is not None:
            latf = float(lat)
            lonf = float(lon)
            return f"https://www.google.com/maps/dir/?api=1&destination={latf},{lonf}&travelmode=driving"

        fb = (fallback_text or "").strip()
        if fb:
            return f"https://www.google.com/maps/dir/?api=1&destination={urllib.parse.quote(fb)}&travelmode=driving"
        return None
    except Exception:
        return None


def _fetch_ors_route_coords(
    origin_lat: float,
    origin_lon: float,
    dest_lat: float,
    dest_lon: float,
) -> Optional[List[Tuple[float, float]]]:
    """
    Return route coordinates as (lat, lon) pairs using OpenRouteService.
    Returns None if ORS is not configured or on any failure.
    """
    key = (ORS_API_KEY or "").strip()
    if not key:
        return None

    try:
        qs = urllib.parse.urlencode({
            "start": f"{origin_lon:.6f},{origin_lat:.6f}",
            "end": f"{dest_lon:.6f},{dest_lat:.6f}",
        })
        url = f"{ORS_DIRECTIONS_URL}?{qs}"

        req = urllib.request.Request(
            url,
            headers={
                "Authorization": key,
                "Accept": "application/json",
            },
            method="GET",
        )

        with urllib.request.urlopen(req, timeout=7) as resp:
            raw = resp.read()

        data = json.loads(raw.decode("utf-8", errors="ignore") or "{}")
        feats = data.get("features") or []
        if not feats:
            return None

        coords = (feats[0].get("geometry") or {}).get("coordinates") or []
        if not coords:
            return None

        # coords are [lon, lat]
        pts = [(float(lat), float(lon)) for lon, lat in coords]

        # Downsample if extremely dense (keep max ~1200 points)
        if len(pts) > 1200:
            step = int(math.ceil(len(pts) / 1200.0))
            pts = pts[::step]
            if pts and pts[-1] != (float(dest_lat), float(dest_lon)):
                pts.append((float(dest_lat), float(dest_lon)))

        return pts
    except Exception:
        return None

def _fetch_osrm_route_coords(
    origin_lat: float,
    origin_lon: float,
    dest_lat: float,
    dest_lon: float,
) -> Optional[List[Tuple[float, float]]]:
    """
    Return route coordinates as (lat, lon) pairs using OSRM (public demo).
    This does NOT require an API key.
    Returns None on any failure.
    """
    try:
        base = (os.environ.get("OSRM_BASE_URL", "https://router.project-osrm.org") or "").strip().rstrip("/")
        if not base:
            base = "https://router.project-osrm.org"

        url = (
            f"{base}/route/v1/driving/"
            f"{origin_lon:.6f},{origin_lat:.6f};{dest_lon:.6f},{dest_lat:.6f}"
            f"?overview=full&geometries=geojson"
        )

        req = urllib.request.Request(
            url,
            headers={
                "Accept": "application/json",
                "User-Agent": "DriverStatus/1.0",
            },
            method="GET",
        )

        with urllib.request.urlopen(req, timeout=7) as resp:
            raw = resp.read()

        data = json.loads(raw.decode("utf-8", errors="ignore") or "{}")
        routes = data.get("routes") or []
        if not routes:
            return None

        geom = routes[0].get("geometry") or {}
        coords = geom.get("coordinates") or []
        if not coords:
            return None

        # coords are [lon, lat]
        pts = [(float(lat), float(lon)) for lon, lat in coords]

        # Downsample if extremely dense (keep max ~1200 points)
        if len(pts) > 1200:
            step = int(math.ceil(len(pts) / 1200.0))
            pts = pts[::step]
            if pts and pts[-1] != (float(dest_lat), float(dest_lon)):
                pts.append((float(dest_lat), float(dest_lon)))

        return pts
    except Exception:
        return None

def build_route_points(
    origin_lat: float,
    origin_lon: float,
    dest_lat: float,
    dest_lon: float,
) -> Tuple[List[List[float]], str]:
    """Return polyline as [[lat, lon], ...] and a short route-source key."""
    pts = _fetch_ors_route_coords(origin_lat, origin_lon, dest_lat, dest_lon)
    if pts:
        return [[lat, lon] for (lat, lon) in pts], "ORS"

    pts2 = _fetch_osrm_route_coords(origin_lat, origin_lon, dest_lat, dest_lon)
    if pts2:
        return [[lat, lon] for (lat, lon) in pts2], "OSRM"

    return [
        [float(origin_lat), float(origin_lon)],
        [float(dest_lat), float(dest_lon)],
    ], "DIRECT"


    pts2 = _fetch_osrm_route_coords(origin_lat, origin_lon, dest_lat, dest_lon)
    if pts2:
        return [[lat, lon] for (lat, lon) in pts2], "Route source: OSRM"

    return [
        [float(origin_lat), float(origin_lon)],
        [float(dest_lat), float(dest_lon)],
    ], "Route source: direct line"


def _snapshot_movements() -> List[Dict[str, Any]]:
    """Return a sanitized list of movement dicts from SNAPSHOT['movements'].

    Accepts either:
      - list[dict]
      - dict[Any, dict] (values will be used)
    Any non-dict items are ignored.
    """
    if not SNAPSHOT or not isinstance(SNAPSHOT, dict):
        return []

    moves = SNAPSHOT.get("movements")
    if isinstance(moves, dict):
        moves = list(moves.values())

    if not isinstance(moves, list):
        return []

    out: List[Dict[str, Any]] = []
    for m in moves:
        if isinstance(m, dict):
            out.append(m)
    return out


def _get_plate_record(plate: str) -> Optional[Dict[str, Any]]:
    """Return the best record for a given plate.

    Rule:
      - A movement becomes INACTIVE if it is departed OR (scheduled_departure + 30 min) has passed.
      - If multiple movements exist for the same plate, we return the next ACTIVE one (earliest scheduled departure).
      - If none are active, return the most recent (latest scheduled departure).
    """
    moves = _snapshot_movements()
    if not moves:
        return None

    plate_n = normalize_plate(plate)
    matches = [m for m in moves if normalize_plate(m.get("license_plate", "")) == plate_n]
    if not matches:
        return None
    if len(matches) == 1:
        return matches[0]

    # Use a stable local timezone for time-based comparisons (defaults to Europe/Amsterdam).
    try:
        from zoneinfo import ZoneInfo  # py3.9+
        _tz = ZoneInfo((os.environ.get("PORTAL_LOCAL_TZ", "") or "Europe/Amsterdam").strip())
    except Exception:
        _tz = None

    def _now_dt() -> datetime:
        try:
            return datetime.now(_tz) if _tz else datetime.now()
        except Exception:
            return datetime.now()

    def _sched_dt(mv: Dict[str, Any]) -> Optional[datetime]:
        dt0 = _parse_dt(mv.get("scheduled_departure") or "")
        if dt0 and _tz and getattr(dt0, "tzinfo", None) is None:
            try:
                dt0 = dt0.replace(tzinfo=_tz)
            except Exception:
                pass
        return dt0

    def _is_departed_mv(mv: Dict[str, Any]) -> bool:
        departed = mv.get("departed", False)
        if isinstance(departed, str):
            departed = departed.strip().lower() in {"1", "true", "yes", "y"}
        if departed:
            return True
        if _has(mv.get("departed_at", "")):
            return True
        return False

    def _is_inactive_mv(mv: Dict[str, Any]) -> bool:
        if _is_departed_mv(mv):
            return True
        dt0 = _sched_dt(mv)
        if dt0:
            try:
                if _now_dt() > (dt0 + timedelta(minutes=30)):
                    return True
            except Exception:
                pass
        return False

    def _score_mv(mv: Dict[str, Any]) -> int:
        s = 0
        try:
            if _clean_location_value(mv.get("location", "")):
                s += 40
        except Exception:
            pass
        if _has(mv.get("close_door", "")):
            s += 30
        if _has(mv.get("trailer", "")):
            s += 10
        if _has(mv.get("scheduled_departure", "")):
            s += 5
        return s

    active = [mv for mv in matches if not _is_inactive_mv(mv)]
    if active:
        # Next active: earliest scheduled departure; tie-breaker: more complete row.
        def _key_active(mv: Dict[str, Any]):
            dt0 = _sched_dt(mv)
            ts = dt0.timestamp() if dt0 else float("inf")
            return (ts, -_score_mv(mv))
        active.sort(key=_key_active)
        return active[0]

    # All inactive: return most recent by scheduled departure; tie-breaker: more complete row.
    def _key_inactive(mv: Dict[str, Any]):
        dt0 = _sched_dt(mv)
        ts = dt0.timestamp() if dt0 else float("-inf")
        return (ts, _score_mv(mv))

    best = None
    best_key = None
    for mv in matches:
        k = _key_inactive(mv)
        if best_key is None or k > best_key:
            best_key = k
            best = mv
    return best


def _utc_iso_now() -> str:
    try:
        return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    except Exception:
        return datetime.utcnow().isoformat() + "Z"


def _prune_check_log(now_ts: Optional[int] = None) -> None:
    """Keep only the last CHECK_LOG_WINDOW_HOURS hours in CHECK_LOG."""
    global CHECK_LOG
    try:
        now = int(now_ts if now_ts is not None else time.time())
    except Exception:
        now = int(time.time())

    cutoff = now - int(CHECK_LOG_WINDOW_SECONDS)
    try:
        CHECK_LOG = [r for r in CHECK_LOG if int((r or {}).get("ts", 0)) >= cutoff]
    except Exception:
        CHECK_LOG = []


def _log_plate_check_event(plate_n: str) -> None:
    """Record a license plate check event (used by developer monitor)."""
    if not plate_n:
        return
    now_ts = int(time.time())
    _prune_check_log(now_ts)

    CHECK_LOG.append({
        "plate": plate_n,
        "ts": now_ts,
        "iso": _utc_iso_now(),
    })

    # Bound memory (should be plenty for 12h window)
    if len(CHECK_LOG) > 5000:
        CHECK_LOG[:] = CHECK_LOG[-5000:]


def _recent_plate_stats() -> Dict[str, Dict[str, Any]]:
    """Return per-plate stats for the last 12 hours: count + last check time."""
    _prune_check_log()
    out: Dict[str, Dict[str, Any]] = {}
    for r in CHECK_LOG:
        try:
            p = normalize_plate((r or {}).get("plate", ""))
            if not p:
                continue
            ts = int((r or {}).get("ts", 0))
            iso = str((r or {}).get("iso", "") or "")
        except Exception:
            continue

        prev = out.get(p)
        if not prev:
            out[p] = {"count": 1, "last_ts": ts, "last_iso": iso}
            continue

        prev["count"] = int(prev.get("count", 0)) + 1
        if ts >= int(prev.get("last_ts", 0)):
            prev["last_ts"] = ts
            prev["last_iso"] = iso

    return out


def _is_plate_recently_checked(plate_n: str) -> bool:
    """True if plate_n was checked in the last 12 hours."""
    pn = normalize_plate(plate_n)
    if not pn:
        return False
    _prune_check_log()

    cutoff = int(time.time()) - int(CHECK_LOG_WINDOW_SECONDS)
    # Walk from newest to oldest for early exit
    for r in reversed(CHECK_LOG):
        try:
            ts = int((r or {}).get("ts", 0))
            if ts < cutoff:
                break
            if normalize_plate((r or {}).get("plate", "")) == pn:
                return True
        except Exception:
            continue
    return False


def _admin_can_push() -> bool:
    """Admin monitor push is allowed only if push is configured + toggle ON + admin has a subscription."""
    if not PUSH_ENABLED:
        return False
    if not ADMIN_NOTIFY_ENABLED:
        return False
    if not SUBSCRIPTIONS_BY_PLATE.get(DEV_PLATE):
        return False
    return True



def _push_to_plate_localized(plate: str, title_key: str, body_by_lang: Dict[str, str]) -> None:
    """Send localized push to each subscription (best-effort)."""
    if not PUSH_ENABLED:
        return
    subs = SUBSCRIPTIONS_BY_PLATE.get(plate, []) or []
    if not subs:
        return

    vapid_claims = {"sub": VAPID_SUBJECT}
    alive = []

    for sub in subs:
        try:
            lang = normalize_lang((sub or {}).get("lang", "en"))
            title = push_title_text(title_key, lang)
            body = body_by_lang.get(lang) or body_by_lang.get("en") or ""

            payload = json.dumps({
                "title": title,
                "body": body,
                "url": f"/?plate={urllib.parse.quote(plate)}&lang={urllib.parse.quote(lang)}",
            })

            webpush(
                subscription_info=sub,
                data=payload,
                vapid_private_key=VAPID_PRIVATE_KEY,
                vapid_claims=vapid_claims,
            )
            alive.append(sub)
        except Exception:
            # Drop dead subscriptions
            pass

    SUBSCRIPTIONS_BY_PLATE[plate] = alive

def _push_admin_event(title_key: str, body_by_lang: Dict[str, str], target_plate: str = "") -> None:
    """Send a push notification to the admin (DEV_PLATE subscription bucket)."""
    if not PUSH_ENABLED:
        return
    subs = SUBSCRIPTIONS_BY_PLATE.get(DEV_PLATE, []) or []
    if not subs:
        return

    vapid_claims = {"sub": VAPID_SUBJECT}
    alive = []

    for sub in subs:
        try:
            lang = normalize_lang((sub or {}).get("lang", "en"))
            title = push_title_text(title_key, lang)
            body = body_by_lang.get(lang) or body_by_lang.get("en") or ""

            tp = normalize_plate(target_plate) if target_plate else DEV_PLATE
            url = f"/?plate={urllib.parse.quote(tp)}&lang={urllib.parse.quote(lang)}"

            payload = json.dumps({
                "title": title or "Admin",
                "body": body,
                "url": url,
            })

            webpush(
                subscription_info=sub,
                data=payload,
                vapid_private_key=VAPID_PRIVATE_KEY,
                vapid_claims=vapid_claims,
            )
            alive.append(sub)
        except Exception:
            pass

    SUBSCRIPTIONS_BY_PLATE[DEV_PLATE] = alive


def _maybe_admin_push_plate_checked(plate: str, movement: Optional[Dict[str, Any]] = None) -> None:
    """If admin monitor is enabled, push when a plate is checked on the website."""
    if not _admin_can_push():
        return

    pn = normalize_plate(plate)
    if not pn or pn == DEV_PLATE:
        return

    now = float(time.time())
    last = float(LAST_ADMIN_CHECK_PUSH_TS_BY_PLATE.get(pn, 0.0))
    if (now - last) < float(ADMIN_CHECK_PUSH_DEDUP_SECONDS):
        return
    LAST_ADMIN_CHECK_PUSH_TS_BY_PLATE[pn] = now

    status_text = ""
    dest_text = "-"
    sched_disp = "-"

    try:
        if movement:
            st = compute_driver_status(movement, lang="en")
            status_text = str(st.get("status_text", "") or "")
            dest_text, _, _ = resolve_destination(movement)
            sched_disp = _format_scheduled_departure(movement.get("scheduled_departure") or "")
        else:
            status_text = ""
    except Exception:
        pass

    msg = f"Plate checked: {pn}\nStatus: {status_text or '-'}\nDep: {sched_disp or '-'}\nDest: {dest_text or '-'}"
    bodies = {l: msg for l in SUPPORTED_LANGS}
    _push_admin_event("ADMIN_MONITOR", bodies, target_plate=pn)


def _maybe_admin_push_status_change(plate: str, movement: Dict[str, Any]) -> None:
    """If admin monitor is enabled, push when a status changes for a recently checked plate."""
    if not _admin_can_push():
        return
    pn = normalize_plate(plate)
    if not pn or pn == DEV_PLATE:
        return
    if not _is_plate_recently_checked(pn):
        return

    try:
        st = compute_driver_status(movement, lang="en")
        status_text = str(st.get("status_text", "") or "")
    except Exception:
        status_text = ""

    try:
        dest_text, _, _ = resolve_destination(movement)
    except Exception:
        dest_text = "-"

    try:
        sched_disp = _format_scheduled_departure(movement.get("scheduled_departure") or "")
    except Exception:
        sched_disp = "-"

    msg = f"Status changed: {pn}\nNew: {status_text or '-'}\nDep: {sched_disp or '-'}\nDest: {dest_text or '-'}"
    bodies = {l: msg for l in SUPPORTED_LANGS}
    _push_admin_event("ADMIN_MONITOR", bodies, target_plate=pn)



def _push_status_change_to_plate(plate: str, movement: Dict[str, Any]) -> None:
    """Push a status update to a plate, in each subscriber's language."""
    try:
        bodies: Dict[str, str] = {}
        for l in SUPPORTED_LANGS:
            bodies[l] = compute_driver_status(movement, lang=l).get("status_text", "")
        _push_to_plate_localized(plate, "STATUS_UPDATE", bodies)
    except Exception:
        return


def _push_driver_message_to_plate(plate: str, message: str) -> None:
    """Push dispatcher message to a plate (message text is not translated)."""
    try:
        bodies = {l: str(message or "") for l in SUPPORTED_LANGS}
        _push_to_plate_localized(plate, "MESSAGE_FROM_DISPATCH", bodies)
    except Exception:
        return


# -----------------------------
# Excel lookup loading (server-side destination calc)
# -----------------------------
def _clean_header(v: Any) -> str:
    s = str(v or "").strip().lower()
    for ch in [" ", "-", "_", "/", "\\", "(", ")", "[", "]", "{", "}", ".", ",", ":"]:
        s = s.replace(ch, "")
    return s


def _find_col(headers: List[str], candidates: List[str]) -> Optional[int]:
    # exact
    for c in candidates:
        if c in headers:
            return headers.index(c)
    # contains
    for i, h in enumerate(headers):
        for c in candidates:
            if c in h:
                return i
    return None


def _safe_float(v: Any) -> Optional[float]:
    try:
        if v is None:
            return None
        s = str(v).strip()
        if not s or s.lower() in {"nan", "none", "nat"}:
            return None
        return float(s)
    except Exception:
        return None


def _load_xlsx_map_locations(path: str) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    if not _OPENPYXL_OK or not os.path.exists(path):
        return out

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)  # type: ignore
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        return out

    headers = [_clean_header(h) for h in header_row]

    code_i = _find_col(headers, ["dest", "code", "locationcode", "loccode", "stationcode", "facilitycode", "destcode"])
    city_i = _find_col(headers, ["city", "town", "name", "locationname"])
    country_i = _find_col(headers, ["country", "land"])
    lat_i = _find_col(headers, ["lat", "latitude"])
    lon_i = _find_col(headers, ["lon", "lng", "long", "longitude"])

    if code_i is None:
        return out

    for r in rows:
        try:
            code = _norm_code(r[code_i] if code_i < len(r) else "")
            if not code:
                continue

            city = str(r[city_i]).strip() if (city_i is not None and city_i < len(r) and r[city_i] is not None) else ""
            country = str(r[country_i]).strip() if (country_i is not None and country_i < len(r) and r[country_i] is not None) else ""

            lat = _safe_float(r[lat_i] if (lat_i is not None and lat_i < len(r)) else None)
            lon = _safe_float(r[lon_i] if (lon_i is not None and lon_i < len(r)) else None)

            out[code] = {"code": code, "city": city, "country": country, "lat": lat, "lon": lon}
        except Exception:
            continue

    return out


def _load_xlsx_map_destland(path: str) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    if not _OPENPYXL_OK or not os.path.exists(path):
        return out

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)  # type: ignore
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        return out

    headers = [_clean_header(h) for h in header_row]
    code_i = _find_col(headers, ["dest", "code", "locationcode", "loccode", "stationcode", "facilitycode", "destcode"])
    city_i = _find_col(headers, ["city", "town", "name", "locationname"])
    country_i = _find_col(headers, ["country", "land"])

    if code_i is None:
        return out

    for r in rows:
        try:
            code = _norm_code(r[code_i] if code_i < len(r) else "")
            if not code:
                continue

            city = str(r[city_i]).strip() if (city_i is not None and city_i < len(r) and r[city_i] is not None) else ""
            country = str(r[country_i]).strip() if (country_i is not None and country_i < len(r) and r[country_i] is not None) else ""

            out[code] = {"code": code, "city": city, "country": country}
        except Exception:
            continue

    return out


def _load_destination_lookups() -> None:
    global LOCATION_BY_CODE, DESTLAND_BY_CODE
    try:
        LOCATION_BY_CODE = _load_xlsx_map_locations(LOCATIONS_XLSX)
    except Exception:
        LOCATION_BY_CODE = {}

    try:
        DESTLAND_BY_CODE = _load_xlsx_map_destland(DEST_LAND_XLSX)
    except Exception:
        DESTLAND_BY_CODE = {}


def _extract_code_from_text(v: Any) -> str:
    s = str(v or "").strip().upper()
    if not s or s.lower() in {"nan", "none", "nat"}:
        return ""

    # If it ends like "... (QAR)" take inside ()
    if "(" in s and s.endswith(")"):
        inside = s.split("(")[-1].replace(")", "").strip()
        inside = _norm_code(inside)
        if 2 <= len(inside) <= 10:
            return inside

    # If the whole thing looks like a code
    compact = _norm_code(s)
    if 2 <= len(compact) <= 10 and any(ch.isalpha() for ch in compact):
        return compact

    # Otherwise take last token if it looks like a code
    parts = [p for p in s.replace(",", " ").replace("/", " ").split() if p]
    if parts:
        last = _norm_code(parts[-1])
        if 2 <= len(last) <= 10 and any(ch.isalpha() for ch in last):
            return last

    return ""


def _first_nonempty(rec: Dict[str, Any], keys: List[str]) -> Any:
    for k in keys:
        if k in rec and _has(rec.get(k)):
            return rec.get(k)
    return None


def resolve_destination(rec: Dict[str, Any]) -> Tuple[str, Optional[float], Optional[float]]:
    """
    Returns: (destination_text, lat, lon)
    destination_text should be: "City, Country (CODE)" when possible.
    """

    # 1) Determine destination code from the snapshot (many possible field names)
    raw_code = _first_nonempty(rec, [
        "dest_code", "DestCode", "DEST_CODE",
        "destination_code", "DestinationCode", "DESTINATION_CODE",
        "dest", "Dest", "DEST",
        "destination", "Destination",
        "destination_text", "DestinationText",
        "dest_text", "DestText", "DEST_TEXT",
    ])

    code = _extract_code_from_text(raw_code)
    code_n = _norm_code(code)

    # 2) Coordinates: prefer snapshot coordinates if provided, otherwise lookup
    lat = _safe_float(_first_nonempty(rec, ["dest_lat", "DestLat", "destination_lat", "DestinationLat", "lat_dest", "LatDest"]))
    lon = _safe_float(_first_nonempty(rec, ["dest_lon", "DestLon", "destination_lon", "DestinationLon", "lon_dest", "LonDest"]))

    city = ""
    country = ""

    # 3) Lookups
    loc_row = LOCATION_BY_CODE.get(code_n) if code_n else None
    dl_row = DESTLAND_BY_CODE.get(code_n) if code_n else None

    # Coordinates: best source is FedEx_locations.xlsx
    if loc_row:
        if lat is None:
            lat = loc_row.get("lat")
        if lon is None:
            lon = loc_row.get("lon")

    # City/Country: prefer dest-land.xlsx because it contains clean city names
    if dl_row:
        city = str(dl_row.get("city") or "").strip()
        country = str(dl_row.get("country") or "").strip()

    # Fallback for city/country (if dest-land missing)
    if loc_row:
        if not city:
            city = str(loc_row.get("city") or "").strip()
            # common pattern: "ARH Depot Elst" -> remove leading "ARH "
            if code_n and city.upper().startswith(code_n + " "):
                city = city[len(code_n) + 1:].strip()
        if not country:
            country = str(loc_row.get("country") or "").strip()

    # 5) Build display text
    if city and country and code_n:
        dest_text = f"{city}, {country} ({code_n})"
    elif city and country:
        dest_text = f"{city}, {country}"
    elif code_n:
        dest_text = code_n
    else:
        # absolute fallback: keep whatever came from snapshot
        dest_text = str(_first_nonempty(rec, ["destination_text", "dest_text", "destination", "Destination"]) or "-")

    return dest_text, lat, lon


# -----------------------------
# API
# -----------------------------
@app.get("/health")
def health() -> Dict[str, Any]:
    return {
        "ok": True,
        "push_enabled": PUSH_ENABLED,
        "snapshot_loaded": bool(SNAPSHOT),
        "lookup_locations_loaded": len(LOCATION_BY_CODE),
        "lookup_destland_loaded": len(DESTLAND_BY_CODE),
        "openpyxl_ok": _OPENPYXL_OK,
    }


@app.post("/api/upload")
async def upload_snapshot(request: Request, secret: str = Query(..., min_length=8)) -> Dict[str, Any]:
    global SNAPSHOT, LAST_STATUS_KEY_BY_PLATE

    if not ADMIN_UPLOAD_SECRET:
        raise HTTPException(status_code=500, detail="Server not configured: ADMIN_UPLOAD_SECRET missing.")
    if secret != ADMIN_UPLOAD_SECRET:
        raise HTTPException(status_code=401, detail="Unauthorized.")

    try:
        body = await request.json()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid JSON: {e}")

    if not isinstance(body, dict) or "movements" not in body:
        raise HTTPException(status_code=400, detail="Snapshot must contain 'movements'.")

    SNAPSHOT = body

    # Normalize movements to a list[dict] (client might send a dict/map)
    try:
        SNAPSHOT["movements"] = _snapshot_movements()
    except Exception:
        SNAPSHOT["movements"] = []

    # Push notifications on status change (best-effort)
    if PUSH_ENABLED:
        try:
            for m in _snapshot_movements():
                plate = normalize_plate(m.get("license_plate", ""))
                if not plate:
                    continue
                st = compute_driver_status(m)
                new_key = st["status_key"]
                old_key = LAST_STATUS_KEY_BY_PLATE.get(plate)
                if old_key is None:
                    LAST_STATUS_KEY_BY_PLATE[plate] = new_key
                    continue
                if new_key != old_key:
                    LAST_STATUS_KEY_BY_PLATE[plate] = new_key
                    _push_status_change_to_plate(plate, m)
                    _maybe_admin_push_status_change(plate, m)
        except Exception:
            pass

    return {"ok": True, "count": len(_snapshot_movements()), "push_enabled": PUSH_ENABLED}

@app.post("/api/driver_message")
async def driver_message(request: Request, secret: str = Query(..., min_length=8)) -> Dict[str, Any]:
    global LAST_STATUS_KEY_BY_PLATE

    if not ADMIN_UPLOAD_SECRET:
        raise HTTPException(status_code=500, detail="Server not configured: ADMIN_UPLOAD_SECRET missing.")
    if secret != ADMIN_UPLOAD_SECRET:
        raise HTTPException(status_code=401, detail="Unauthorized.")

    try:
        body = await request.json()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid JSON: {e}")

    if not isinstance(body, dict):
        raise HTTPException(status_code=400, detail="Invalid JSON body.")

    plate = normalize_plate(str(body.get("plate", "")))
    message = str(body.get("message", "") or "").strip()

    if not plate:
        raise HTTPException(status_code=400, detail="Missing 'plate'.")
    if not message:
        raise HTTPException(status_code=400, detail="Missing 'message'.")

    # Save manual message
    MANUAL_STATUS_BY_PLATE[plate] = message

    # Force immediate push + update last key
    st = compute_driver_status({"license_plate": plate})
    try:
        LAST_STATUS_KEY_BY_PLATE[plate] = st["status_key"]
    except Exception:
        pass

    _push_driver_message_to_plate(plate, message)

    return {"ok": True, "plate": plate, "message": message}


@app.get("/api/status")
def get_status(
    plate: str = Query(..., min_length=2),
    lat: float = Query(...),
    lon: float = Query(...),
    ts: int = Query(..., description="Unix epoch seconds from the device"),
    lang: str = Query("en", description="Language: en, de, nl, fr, tr, sv, es, it, ro, ru, lt, kk, hi, pl, hu, uz, tg, ky, be"),
) -> Dict[str, Any]:
    # Enforce geofence, but we do NOT return geofence data anymore
    geofence_check(lat, lon, ts)

    rec = _get_plate_record(plate)
    if rec is None:
        try:
            p0 = normalize_plate(plate)
            _log_plate_check_event(p0)
            _maybe_admin_push_plate_checked(p0, None)
        except Exception:
            pass

        return {
            "plate": normalize_plate(plate),
            "found": False,
            "house_rules_accepted": normalize_plate(plate) in HOUSE_RULES_ACCEPTED_BY_PLATE,
            "house_rules_required": normalize_plate(plate) not in HOUSE_RULES_ACCEPTED_BY_PLATE,
            "last_refresh": (SNAPSHOT or {}).get("last_update"),
        }

    st = compute_driver_status(rec, lang=lang)

    dest_text, dlat, dlon = resolve_destination(rec)
    nav = destination_nav_url(dlat, dlon, dest_text)

    sched_raw = rec.get("scheduled_departure") or ""
    sched_disp = _format_scheduled_departure(sched_raw)


    # Mark that this plate was checked on the website (used by desktop for 👁 icon)
    try:
        p = normalize_plate(plate)
        prev = VIEWED_BY_PLATE.get(p) or {}
        VIEWED_BY_PLATE[p] = {
            "count": int(prev.get("count", 0)) + 1,
            "last_view": datetime.utcnow().isoformat() + "Z",
        }
        _log_plate_check_event(p)
        _maybe_admin_push_plate_checked(p, rec)
    except Exception:
        pass

    return {
        "plate": normalize_plate(plate),
        "found": True,
        "house_rules_accepted": normalize_plate(plate) in HOUSE_RULES_ACCEPTED_BY_PLATE,
        "house_rules_required": normalize_plate(plate) not in HOUSE_RULES_ACCEPTED_BY_PLATE,
        "status_key": st["status_key"],
        "status_text": st["status_text"],
        "destination_text": dest_text,
        "destination_nav_url": nav,
        "scheduled_departure": sched_disp,
        "report_in_office_at": st["report_in_office_at"],
        "trailer": rec.get("trailer") or "",
        "location": _clean_location_value(rec.get("location") or ""),
        "last_refresh": (SNAPSHOT or {}).get("last_update"),
        "push_enabled": PUSH_ENABLED,
        "vapid_public_key": VAPID_PUBLIC_KEY if PUSH_ENABLED else "",
    }




@app.post("/api/house_rules_accept")
def house_rules_accept(payload: Dict[str, Any] = Body(...)) -> Dict[str, Any]:
    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Invalid payload")

    plate = normalize_plate(payload.get("plate") or "")
    if len(plate) < 2:
        raise HTTPException(status_code=400, detail="Invalid plate")

    accepted_at = datetime.utcnow().isoformat() + "Z"
    HOUSE_RULES_ACCEPTED_BY_PLATE[plate] = accepted_at
    return {"ok": True, "plate": plate, "accepted_at": accepted_at}


@app.get("/api/admin/plate_flags")
def get_plate_flags(
    secret: str = Query(..., min_length=8),
    plates: str = Query("", description="Comma-separated list of plates"),
) -> Dict[str, Any]:
    if not ADMIN_UPLOAD_SECRET:
        raise HTTPException(status_code=500, detail="Server not configured: ADMIN_UPLOAD_SECRET missing.")
    if secret != ADMIN_UPLOAD_SECRET:
        raise HTTPException(status_code=401, detail="Unauthorized.")

    out: Dict[str, Any] = {}
    raw = (plates or "").strip()
    plate_list = [p for p in (raw.split(",") if raw else []) if p.strip()]

    for p in plate_list:
        np = normalize_plate(p)
        v = VIEWED_BY_PLATE.get(np) or {}
        out[np] = {
            "viewed": bool(v),
            "last_view": v.get("last_view", "") if isinstance(v, dict) else "",
            "count": int(v.get("count", 0)) if isinstance(v, dict) else 0,
            "push_enabled": bool(SUBSCRIPTIONS_BY_PLATE.get(np)),
        }

    return {"ok": True, "plates": out}



@app.get("/api/dev/summary")
def dev_summary(
    key: str = Query(..., min_length=3, description="Developer key (DEV_PLATE)"),
) -> Dict[str, Any]:
    if normalize_plate(key) != DEV_PLATE:
        raise HTTPException(status_code=401, detail="Unauthorized.")

    _prune_check_log()

    stats = _recent_plate_stats()
    items: List[Dict[str, Any]] = []

    for plate, s in stats.items():
        pn = normalize_plate(plate)
        if not pn:
            continue

        rec = _get_plate_record(pn)
        found = bool(rec)

        dest_text = "-"
        sched_disp = "-"
        status_text = "-"
        status_key = ""

        if rec:
            try:
                st = compute_driver_status(rec, lang="en")
                status_text = str(st.get("status_text", "") or "")
                status_key = str(st.get("status_key", "") or "")
            except Exception:
                pass

            try:
                dest_text, _, _ = resolve_destination(rec)
            except Exception:
                dest_text = "-"

            try:
                sched_disp = _format_scheduled_departure(rec.get("scheduled_departure") or "")
            except Exception:
                sched_disp = "-"

        items.append({
            "plate": pn,
            "last_check": str(s.get("last_iso", "") or ""),
            "count_12h": int(s.get("count", 0) or 0),
            "movement_found": found,
            "scheduled_departure": sched_disp,
            "destination_text": dest_text,
            "status_key": status_key,
            "status_text": status_text,
            "bell": bool(SUBSCRIPTIONS_BY_PLATE.get(pn)),
        })

    # Sort: newest checks first
    try:
        items.sort(key=lambda x: x.get("last_check", ""), reverse=True)
    except Exception:
        pass

    # Recent raw events (newest first) - useful for debugging
    try:
        events = sorted(CHECK_LOG, key=lambda r: int((r or {}).get("ts", 0)), reverse=True)[:200]
    except Exception:
        events = []

    return {
        "ok": True,
        "dev_plate": DEV_PLATE,
        "admin_notify_enabled": bool(ADMIN_NOTIFY_ENABLED),
        "admin_notify_changed_at": str(ADMIN_NOTIFY_CHANGED_AT or ""),
        "admin_subscribed": bool(SUBSCRIPTIONS_BY_PLATE.get(DEV_PLATE)),
        "push_enabled": bool(PUSH_ENABLED),
        "vapid_public_key": VAPID_PUBLIC_KEY if PUSH_ENABLED else "",
        "items": items,
        "events": events,
    }


@app.post("/api/dev/admin_notify")
def dev_set_admin_notify(
    key: str = Query(..., min_length=3, description="Developer key (DEV_PLATE)"),
    payload: Dict[str, Any] = Body(...),
) -> Dict[str, Any]:
    if normalize_plate(key) != DEV_PLATE:
        raise HTTPException(status_code=401, detail="Unauthorized.")

    global ADMIN_NOTIFY_ENABLED, ADMIN_NOTIFY_CHANGED_AT
    enabled = bool((payload or {}).get("enabled", False))

    ADMIN_NOTIFY_ENABLED = enabled
    ADMIN_NOTIFY_CHANGED_AT = _utc_iso_now()

    return {"ok": True, "enabled": bool(ADMIN_NOTIFY_ENABLED), "changed_at": ADMIN_NOTIFY_CHANGED_AT}


@app.post("/api/dev/subscribe_admin")
def dev_subscribe_admin(
    key: str = Query(..., min_length=3, description="Developer key (DEV_PLATE)"),
    lang: str = Query("en", description="Language for admin push notifications"),
    subscription: Dict[str, Any] = Body(...),
) -> Dict[str, Any]:
    if not PUSH_ENABLED:
        raise HTTPException(status_code=400, detail="Push is not enabled on the server.")
    if normalize_plate(key) != DEV_PLATE:
        raise HTTPException(status_code=401, detail="Unauthorized.")

    if not isinstance(subscription, dict) or "endpoint" not in subscription:
        raise HTTPException(status_code=400, detail="Invalid subscription.")

    subs = SUBSCRIPTIONS_BY_PLATE.get(DEV_PLATE, []) or []
    endpoint = subscription.get("endpoint")
    subs = [s for s in subs if s.get("endpoint") != endpoint]

    sub_rec = dict(subscription)
    sub_rec["lang"] = normalize_lang(lang)
    subs.append(sub_rec)

    SUBSCRIPTIONS_BY_PLATE[DEV_PLATE] = subs

    return {"ok": True, "plate": DEV_PLATE, "count": len(subs)}


@app.post("/api/dev/send_message")
async def dev_send_message(
    request: Request,
    key: str = Query(..., min_length=3, description="Developer key (DEV_PLATE)"),
) -> Dict[str, Any]:
    if normalize_plate(key) != DEV_PLATE:
        raise HTTPException(status_code=401, detail="Unauthorized.")

    try:
        body = await request.json()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid JSON: {e}")

    if not isinstance(body, dict):
        raise HTTPException(status_code=400, detail="Invalid JSON body.")

    plate = normalize_plate(str(body.get("plate", "")))
    message = str(body.get("message", "") or "").strip()

    if not plate:
        raise HTTPException(status_code=400, detail="Missing 'plate'.")
    if not message:
        raise HTTPException(status_code=400, detail="Missing 'message'.")
    if len(message) > 500:
        raise HTTPException(status_code=400, detail="Message too long (max 500 characters).")
    if not SUBSCRIPTIONS_BY_PLATE.get(plate):
        raise HTTPException(status_code=400, detail="This movement has no active push subscribers.")

    MANUAL_STATUS_BY_PLATE[plate] = message

    st = compute_driver_status({"license_plate": plate})
    try:
        LAST_STATUS_KEY_BY_PLATE[plate] = st["status_key"]
    except Exception:
        pass

    _push_driver_message_to_plate(plate, message)

    return {
        "ok": True,
        "plate": plate,
        "message": message,
        "subscriber_count": len(SUBSCRIPTIONS_BY_PLATE.get(plate) or []),
    }




@app.get("/api/traffic_delay")
def traffic_delay(
    secret: str = Query(..., min_length=8),
    o: str = Query(..., description="origin as 'lat,lon'"),
    d: str = Query(..., description="destination as 'lat,lon'"),
    depart: str = Query("", description="optional ISO-8601 departure time"),
) -> Dict[str, Any]:
    """Return live-traffic delay minutes (server-side)."""
    if not ADMIN_UPLOAD_SECRET:
        raise HTTPException(status_code=500, detail="Server not configured: ADMIN_UPLOAD_SECRET missing.")
    if secret != ADMIN_UPLOAD_SECRET:
        raise HTTPException(status_code=401, detail="Unauthorized.")

    try:
        o_lat_s, o_lon_s = [x.strip() for x in (o or "").split(",", 1)]
        d_lat_s, d_lon_s = [x.strip() for x in (d or "").split(",", 1)]
        o_lat = float(o_lat_s)
        o_lon = float(o_lon_s)
        d_lat = float(d_lat_s)
        d_lon = float(d_lon_s)
    except Exception:
        return {"ok": False, "error": "invalid o/d coords"}

    depart_bucket = (depart or "").strip()[:16]  # minute bucket
    cache_key = (round(o_lat, 4), round(o_lon, 4), round(d_lat, 4), round(d_lon, 4), depart_bucket)

    cached = _traffic_cache_get(cache_key)
    if cached is not None:
        return cached

    delay_min, err = _here_fetch_delay_minutes(o_lat, o_lon, d_lat, d_lon, (depart or "").strip())
    if err:
        payload = {"ok": False, "error": err}
        _traffic_cache_set(cache_key, payload)
        return payload

    payload = {"ok": True, "delay_min": int(delay_min or 0)}
    _traffic_cache_set(cache_key, payload)
    return payload


@app.get("/api/route")
def get_route(
    plate: str = Query(..., min_length=2),
    lat: float = Query(...),
    lon: float = Query(...),
    ts: int = Query(..., description="Unix epoch seconds from the device"),
    lang: str = Query("en", description="Language: en, de, nl, fr, tr, sv, es, it, ro, ru, lt, kk, hi, pl, hu, uz, tg, ky, be"),
) -> Dict[str, Any]:
    """Return a zoomable route map polyline for the website."""
    geofence_check(lat, lon, ts)

    rec = _get_plate_record(plate)
    if rec is None:
        raise HTTPException(status_code=404, detail="No movement found for this plate.")

    dest_text, dlat, dlon = resolve_destination(rec)
    if dlat is None or dlon is None:
        raise HTTPException(status_code=404, detail="Destination coordinates not available for this movement.")

    origin_lat = float(HUB_LAT)
    origin_lon = float(HUB_LON)
    dest_lat = float(dlat)
    dest_lon = float(dlon)

    route_pts, route_key = build_route_points(origin_lat, origin_lon, dest_lat, dest_lon)
    note = route_note_text(route_key, lang)

    return {
        "plate": normalize_plate(plate),
        "origin": {"lat": origin_lat, "lon": origin_lon},
        "dest": {"lat": dest_lat, "lon": dest_lon},
        "route": route_pts,
        "note": note,
        "destination_text": dest_text,
        "last_refresh": (SNAPSHOT or {}).get("last_update"),
    }



@app.post("/api/subscribe")
def subscribe(
    plate: str = Query(..., min_length=2),
    lat: float = Query(...),
    lon: float = Query(...),
    ts: int = Query(...),
    lang: str = Query("en", description="Language: en, de, nl, fr, tr, sv, es, it, ro, ru, lt, kk, hi, pl, hu, uz, tg, ky, be"),
    subscription: Dict[str, Any] = Body(...),
) -> Dict[str, Any]:
    if not PUSH_ENABLED:
        raise HTTPException(status_code=400, detail="Push is not enabled on the server.")

    geofence_check(lat, lon, ts)

    if not isinstance(subscription, dict) or "endpoint" not in subscription:
        raise HTTPException(status_code=400, detail="Invalid subscription.")

    plate_n = normalize_plate(plate)

    subs = SUBSCRIPTIONS_BY_PLATE.get(plate_n, []) or []
    endpoint = subscription.get("endpoint")
    subs = [s for s in subs if s.get("endpoint") != endpoint]

    sub_rec = dict(subscription)
    sub_rec["lang"] = normalize_lang(lang)
    subs.append(sub_rec)

    SUBSCRIPTIONS_BY_PLATE[plate_n] = subs

    return {"ok": True, "plate": plate_n, "count": len(subs)}



@app.get("/favicon.ico", include_in_schema=False)
def favicon() -> FileResponse:
    """Serve the browser tab icon.

    Browsers often request /favicon.ico directly (not /static/favicon.ico).
    Put the icon file here: static/favicon.ico
    """
    path = os.path.join(STATIC_DIR, "favicon.ico")
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="favicon.ico not found in static/")
    return FileResponse(path, media_type="image/x-icon")


@app.get("/sw.js")
def sw() -> Response:
    return Response(content=SERVICE_WORKER_JS, media_type="application/javascript")


@app.get("/")
def index() -> HTMLResponse:
    return HTMLResponse(INDEX_HTML)


# -----------------------------
# Website (no geofence shown + background image + destination is server-calculated)
# -----------------------------
INDEX_HTML = r"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width,initial-scale=1" />
  <title>Driver Status</title>
  <link rel="icon" href="/favicon.ico" type="image/x-icon" />
  <link rel="icon" href="/static/favicon.ico" type="image/x-icon" />

    <link rel="manifest" href="/static/manifest.webmanifest" />
  <link rel="apple-touch-icon" href="/static/apple-touch-icon.png" />
  <meta name="theme-color" content="#4D148C" />
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css" />
  <style>
    body {
      font-family: Arial, sans-serif;
      margin: 0;
      min-height: 100vh;
      background: url('/static/bg.png') no-repeat center center fixed;
      background-size: cover;
    }
    .wrap {
      width: min(860px, calc(100% - 24px));
      margin: 0 auto;
      padding: 18px 0 24px;
    }
    .topcard {
      background: rgba(255,255,255,0.45);
      border-radius: 16px;
      padding: 14px;
      border: 1px solid rgba(0,0,0,0.08);
      box-shadow: 0 10px 30px rgba(0,0,0,0.10);
    }
    input {
      font-size: 16px;
      padding: 10px 12px;
      border-radius: 12px;
      border: 1px solid rgba(0,0,0,0.18);
      background: rgba(255,255,255,0.60);
      outline: none;
    }
    input:focus {
      border-color: rgba(77,20,140,0.55);
      box-shadow: 0 0 0 3px rgba(77,20,140,0.18);
    }

    .btn {
      font-size: 16px;
      padding: 10px 16px;
      border-radius: 12px;
      border: none;
      cursor: pointer;
      font-weight: 700;
      letter-spacing: 0.2px;
      box-shadow: 0 10px 18px rgba(0,0,0,0.18);
      transition: transform 0.08s ease, filter 0.15s ease, box-shadow 0.15s ease;
      user-select: none;
      -webkit-tap-highlight-color: transparent;
    }
    .btn:hover {
      filter: brightness(1.08);
      transform: translateY(-1px);
      box-shadow: 0 12px 22px rgba(0,0,0,0.22);
    }
    .btn:active {
      transform: translateY(0px);
      filter: brightness(0.98);
      box-shadow: 0 8px 14px rgba(0,0,0,0.18);
    }
    .btn:focus-visible {
      outline: none;
      box-shadow: 0 0 0 3px rgba(77,20,140,0.24), 0 10px 18px rgba(0,0,0,0.18);
    }

    .btn-primary {
      color: #ffffff;
      background: linear-gradient(180deg, rgba(104,68,232,1) 0%, rgba(60,32,170,1) 100%);
    }

    .btn-secondary {
      color: rgba(25,25,35,1);
      background: linear-gradient(180deg, rgba(245,245,255,0.70) 0%, rgba(220,220,235,0.45) 100%);
      border: 1px solid rgba(255,255,255,0.35);
    }
    .row { display: flex; gap: 8px; }
    .row > * { flex: 1; }
    .row-main > input { flex: 1 1 auto; }
    .row-main > button { flex: 0 0 120px; }
    .card { border: 1px solid #ddd; border-radius: 12px; padding: 14px; margin-top: 12px; background: rgba(255,255,255,0.45); }
    .muted { color: #666; }
    .status-big { font-size: 22px; font-weight: 700; line-height: 1.25; }
    .ok { border-color: #bfe6c3; }
    .warn { border-color: #ffd18a; }
    .err { border-color: #f5b5b5; }

    #map {
      height: clamp(240px, 42vh, 460px);
      width: 100%;
      border-radius: 12px;
      border: 1px solid rgba(0,0,0,0.18);
      overflow: hidden;
      background: rgba(255,255,255,0.35);
    }

    .langbar {
      display: flex;
      gap: 10px;
      align-items: center;
      margin: 8px 0 10px;
      flex-wrap: wrap;
    }
    .flagbtn {
      width: 38px;
      height: 38px;
      border-radius: 999px;
      border: 1px solid rgba(0,0,0,0.16);
      background: rgba(255,255,255,0.55);
      cursor: pointer;
      font-size: 22px;
      line-height: 1;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      box-shadow: 0 8px 16px rgba(0,0,0,0.12);
      transition: transform 0.08s ease, filter 0.15s ease, box-shadow 0.15s ease;
      user-select: none;
      -webkit-tap-highlight-color: transparent;
      padding: 0;
    }
    .flagbtn:hover { filter: brightness(1.08); transform: translateY(-1px); box-shadow: 0 10px 18px rgba(0,0,0,0.16); }
    .flagbtn:active { transform: translateY(0px); filter: brightness(0.98); box-shadow: 0 7px 14px rgba(0,0,0,0.12); }
    .flagbtn.active {
      outline: none;
      border-color: rgba(77,20,140,0.55);
      box-shadow: 0 0 0 3px rgba(77,20,140,0.18), 0 8px 16px rgba(0,0,0,0.12);
    }

    @media (max-width: 420px) {
      .row { flex-direction: column; }
      .row-main > button { flex: 0 0 auto; width: 100%; }
      body { background-attachment: scroll; }
    }
    a { color: inherit; }
  
    /* House rules modal */
    .modal-backdrop {
      position: fixed;
      inset: 0;
      background: rgba(0,0,0,0.45);
      display: flex;
      align-items: center;
      justify-content: center;
      padding: 18px;
      z-index: 9999;
    }
    .modal {
      width: min(920px, 100%);
      max-height: calc(100vh - 36px);
      overflow: auto;
      background: rgba(255,255,255,0.95);
      border-radius: 16px;
      border: 1px solid rgba(0,0,0,0.12);
      box-shadow: 0 20px 60px rgba(0,0,0,0.35);
    }
    .modal-header {
      padding: 12px 14px;
      border-bottom: 1px solid rgba(0,0,0,0.10);
    }
    .modal-title { font-size: 18px; font-weight: 700; }
    .modal-body { padding: 12px 14px 14px; }

    .hr-text {
      margin-top: 10px;
      padding: 10px 12px;
      border-radius: 12px;
      background: rgba(255,255,255,0.65);
      border: 1px solid rgba(0,0,0,0.08);
    }
    .hr-text h3 { margin: 10px 0 6px; font-size: 15px; }
    .hr-text ul { margin: 6px 0 10px 18px; }
    .hr-text li { margin: 2px 0; }

    .hr-img {
      width: 100%;
      height: auto;
      display: block;
      border-radius: 12px;
      border: 1px solid rgba(0,0,0,0.12);
      background: rgba(255,255,255,0.85);
      margin: 8px 0 10px;
    }

    .hr-accept { margin-top: 12px; display: flex; gap: 10px; align-items: center; }
    .hr-accept-label { font-weight: 700; }

    .switch { position: relative; display: inline-block; width: 50px; height: 28px; }
    .switch input { opacity: 0; width: 0; height: 0; }
    .slider {
      position: absolute;
      cursor: pointer;
      top: 0; left: 0; right: 0; bottom: 0;
      background: #bbb;
      transition: .2s;
      border-radius: 999px;
    }
    .slider:before {
      position: absolute;
      content: "";
      height: 22px;
      width: 22px;
      left: 3px;
      top: 3px;
      background: white;
      transition: .2s;
      border-radius: 50%;
    }
    .switch input:checked + .slider { background: #4D148C; }
    .switch input:checked + .slider:before { transform: translateX(22px); }

  </style>
</head>
<body>
  <div class="wrap">
    <div class="topcard">
      <h2 id="titleH2" style="margin: 6px 0 6px;">Movement status by license plate</h2>
      <div class="langbar" id="langbar" aria-label="Language">
        <button class="flagbtn" data-lang="en" title="English" aria-label="English">🇬🇧</button>
        <button class="flagbtn" data-lang="de" title="Deutsch" aria-label="Deutsch">🇩🇪</button>
        <button class="flagbtn" data-lang="nl" title="Nederlands" aria-label="Nederlands">🇳🇱</button>
        <button class="flagbtn" data-lang="fr" title="Français" aria-label="Français">🇫🇷</button>
        <button class="flagbtn" data-lang="tr" title="Türkçe" aria-label="Türkçe">🇹🇷</button>
        <button class="flagbtn" data-lang="sv" title="Svenska" aria-label="Svenska">🇸🇪</button>
        <button class="flagbtn" data-lang="es" title="Español" aria-label="Español">🇪🇸</button>
        <button class="flagbtn" data-lang="it" title="Italiano" aria-label="Italiano">🇮🇹</button>
        <button class="flagbtn" data-lang="ro" title="Română" aria-label="Română">🇷🇴</button>
        <button class="flagbtn" data-lang="ru" title="Русский" aria-label="Русский">🇷🇺</button>
        <button class="flagbtn" data-lang="lt" title="Lietuvių" aria-label="Lietuvių">🇱🇹</button>
        <button class="flagbtn" data-lang="kk" title="Қазақша" aria-label="Қазақша">🇰🇿</button>
        <button class="flagbtn" data-lang="hi" title="हिन्दी" aria-label="हिन्दी">🇮🇳</button>
        <button class="flagbtn" data-lang="pl" title="Polski" aria-label="Polski">🇵🇱</button>
        <button class="flagbtn" data-lang="hu" title="Magyar" aria-label="Magyar">🇭🇺</button>
        <button class="flagbtn" data-lang="uz" title="O‘zbek" aria-label="O‘zbek">🇺🇿</button>
        <button class="flagbtn" data-lang="tg" title="Тоҷикӣ" aria-label="Тоҷикӣ">🇹🇯</button>
        <button class="flagbtn" data-lang="ky" title="Кыргызча" aria-label="Кыргызча">🇰🇬</button>
        <button class="flagbtn" data-lang="be" title="Беларуская" aria-label="Беларуская">🇧🇾</button>
      </div>

      <div class="row row-main">
        <input id="plate" placeholder="Enter license plate (e.g. AB-123-CD)" />
        <button id="btn" class="btn btn-primary">Check</button>
      </div>

      <div class="row" style="margin-top: 8px;">
        <button id="btnNotify" class="btn btn-secondary" style="display:none;">Enable notifications</button>
      </div>
      <div id="notifyMsg" class="muted" style="margin-top:6px; display:none;"></div>

      <div id="out" class="card" style="display:none;"></div>
    </div>
  </div>

  <!-- House rules / routes modal (shown once per server session per license plate) -->
  <div id="hrBackdrop" class="modal-backdrop" style="display:none;">
    <div class="modal" role="dialog" aria-modal="true" aria-labelledby="hrTitle">
      <div class="modal-header">
        <div class="modal-title" id="hrTitle">House rules</div>
      </div>
      <div class="modal-body">
        <div id="hrIntro" class="muted" style="margin-bottom:10px;"></div>
        </div>

        <div id="hrText" class="hr-text"></div>

        <div class="hr-accept">
          <label class="switch" title="Accept">
            <input id="hrAccept" type="checkbox" />
            <span class="slider"></span>
          </label>
          <div id="hrAcceptLabel" class="hr-accept-label"></div>
        </div>

        <div style="display:flex; gap:10px; justify-content:flex-end; margin-top:12px;">
          <button id="hrContinue" class="btn btn-primary" disabled>Continue</button>
        </div>
      </div>
    </div>
  </div>

  <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>

  <script>
    const API_BASE = window.location.origin;
    const DEV_PLATE = "KLETH743";
    const SUPPORTED_LANGS = ["en", "de", "nl", "fr", "tr", "sv", "es", "it", "ro", "ru", "lt", "kk", "hi", "pl", "hu", "uz", "tg", "ky", "be"];
    const UI = {
      en: {
        title: "Movement status by license plate",
        plate_ph: "Enter license plate (e.g. AB-123-CD)",
        btn_check: "Check",
        btn_notify: "Enable notifications",
        btn_enabling: "Enabling...",
        btn_enabled: "Notifications enabled",

        getting_location: "Getting location…",
        loading_status: "Loading status…",
        loading_route: "Loading route…",

        no_movement: "No movement found",
        last_refresh: "Last refresh",
        destination: "Destination",
        departure_time: "Departure time",
        report_office: "Report in the office",
        trailer: "Trailer",
        place: "Place",
        route_map: "Route map",
        origin: "Origin",
        destination_pin: "Destination",

        parking: "Parking",
        dock: "Dock",

        err_location: "Location error",
        err_network: "Network error",
        err_error: "Error",
        help_location: "Enable GPS and allow location permission.",

        notify_not_supported: "Notifications not supported",
        notify_not_supported_help: "Use Chrome/Edge on Android. iOS requires adding the site to Home Screen.",
        notify_denied: "Notifications denied",
        notify_denied_help: "Allow notifications in browser settings.",
        notify_failed: "Subscribe failed",
        notify_enabled_msg: "Notifications enabled",
        notify_enabled_help: "You will receive a push when your status changes.",
        subscribe_error: "Subscribe error",
        route_error: "Route error"
      },
      de: {
        title: "Bewegungsstatus nach Kennzeichen",
        plate_ph: "Kennzeichen eingeben (z. B. AB-123-CD)",
        btn_check: "Prüfen",
        btn_notify: "Benachrichtigungen aktivieren",
        btn_enabling: "Aktiviere…",
        btn_enabled: "Benachrichtigungen aktiv",

        getting_location: "Standort wird abgerufen…",
        loading_status: "Status wird geladen…",
        loading_route: "Route wird geladen…",

        no_movement: "Keine Bewegung gefunden",
        last_refresh: "Letzte Aktualisierung",
        destination: "Ziel",
        departure_time: "Abfahrtszeit",
        report_office: "Im Büro melden",
        trailer: "Anhänger",
        place: "Ort",
        route_map: "Routenkarte",
        origin: "Start",
        destination_pin: "Ziel",

        parking: "Parkplatz",
        dock: "Tor",

        err_location: "Standortfehler",
        err_network: "Netzwerkfehler",
        err_error: "Fehler",
        help_location: "GPS aktivieren und Standortzugriff erlauben.",

        notify_not_supported: "Benachrichtigungen nicht unterstützt",
        notify_not_supported_help: "Nutze Chrome/Edge auf Android. Unter iOS muss die Seite zum Home-Bildschirm hinzugefügt werden.",
        notify_denied: "Benachrichtigungen abgelehnt",
        notify_denied_help: "Benachrichtigungen in den Browser-Einstellungen erlauben.",
        notify_failed: "Abonnement fehlgeschlagen",
        notify_enabled_msg: "Benachrichtigungen aktiv",
        notify_enabled_help: "Du erhältst eine Push-Nachricht, wenn sich dein Status ändert.",
        subscribe_error: "Abo-Fehler",
        route_error: "Routenfehler"
      },
      nl: {
        title: "Bewegingsstatus op kenteken",
        plate_ph: "Kenteken invoeren (bv. AB-123-CD)",
        btn_check: "Check",
        btn_notify: "Meldingen inschakelen",
        btn_enabling: "Inschakelen…",
        btn_enabled: "Meldingen ingeschakeld",

        getting_location: "Locatie ophalen…",
        loading_status: "Status laden…",
        loading_route: "Route laden…",

        no_movement: "Geen beweging gevonden",
        last_refresh: "Laatste update",
        destination: "Bestemming",
        departure_time: "Vertrektijd",
        report_office: "Melden op kantoor",
        trailer: "Trailer",
        place: "Plek",
        route_map: "Routekaart",
        origin: "Start",
        destination_pin: "Bestemming",

        parking: "Parkeerplaats",
        dock: "Dock",

        err_location: "Locatiefout",
        err_network: "Netwerkfout",
        err_error: "Fout",
        help_location: "Zet GPS aan en sta locatie-toestemming toe.",

        notify_not_supported: "Meldingen niet ondersteund",
        notify_not_supported_help: "Gebruik Chrome/Edge op Android. Op iOS moet de site aan het beginscherm worden toegevoegd.",
        notify_denied: "Meldingen geweigerd",
        notify_denied_help: "Sta meldingen toe in de browserinstellingen.",
        notify_failed: "Abonneren mislukt",
        notify_enabled_msg: "Meldingen ingeschakeld",
        notify_enabled_help: "Je ontvangt een push als je status verandert.",
        subscribe_error: "Abonneerfout",
        route_error: "Routefout"
      },

      es: {
        title: "Estado del movimiento por matrícula",
        plate_ph: "Introduce la matrícula (p. ej. AB-123-CD)",
        btn_check: "Comprobar",
        btn_notify: "Activar notificaciones",
        btn_enabling: "Activando…",
        btn_enabled: "Notificaciones activadas",

        getting_location: "Obteniendo ubicación…",
        loading_status: "Cargando estado…",
        loading_route: "Cargando ruta…",

        no_movement: "No se encontró movimiento",
        last_refresh: "Última actualización",
        destination: "Destino",
        departure_time: "Hora de salida",
        report_office: "Presentarse en la oficina",
        trailer: "Remolque",
        place: "Lugar",
        route_map: "Mapa de ruta",
        origin: "Origen",
        destination_pin: "Destino",

        parking: "Parking",
        dock: "Muelle",

        err_location: "Error de ubicación",
        err_network: "Error de red",
        err_error: "Error",
        help_location: "Activa el GPS y permite el acceso a la ubicación.",

        notify_not_supported: "Notificaciones no compatibles",
        notify_not_supported_help: "Usa Chrome/Edge en Android. En iOS hay que añadir el sitio a la pantalla de inicio.",
        notify_denied: "Notificaciones denegadas",
        notify_denied_help: "Permite las notificaciones en la configuración del navegador.",
        notify_failed: "Fallo al suscribirse",
        notify_enabled_msg: "Notificaciones activadas",
        notify_enabled_help: "Recibirás un push cuando cambie tu estado.",
        subscribe_error: "Error de suscripción",
        route_error: "Error de ruta"
      },
      it: {
        title: "Stato del movimento per targa",
        plate_ph: "Inserisci la targa (es. AB-123-CD)",
        btn_check: "Verifica",
        btn_notify: "Abilita notifiche",
        btn_enabling: "Abilitazione…",
        btn_enabled: "Notifiche abilitate",

        getting_location: "Rilevamento posizione…",
        loading_status: "Caricamento stato…",
        loading_route: "Caricamento percorso…",

        no_movement: "Nessun movimento trovato",
        last_refresh: "Ultimo aggiornamento",
        destination: "Destinazione",
        departure_time: "Ora di partenza",
        report_office: "Presentarsi in ufficio",
        trailer: "Rimorchio",
        place: "Luogo",
        route_map: "Mappa percorso",
        origin: "Origine",
        destination_pin: "Destinazione",

        parking: "Parcheggio",
        dock: "Dock",

        err_location: "Errore posizione",
        err_network: "Errore di rete",
        err_error: "Errore",
        help_location: "Attiva il GPS e consenti l'accesso alla posizione.",

        notify_not_supported: "Notifiche non supportate",
        notify_not_supported_help: "Usa Chrome/Edge su Android. Su iOS aggiungi il sito alla schermata Home.",
        notify_denied: "Notifiche negate",
        notify_denied_help: "Consenti le notifiche nelle impostazioni del browser.",
        notify_failed: "Iscrizione non riuscita",
        notify_enabled_msg: "Notifiche abilitate",
        notify_enabled_help: "Riceverai un push quando cambia lo stato.",
        subscribe_error: "Errore iscrizione",
        route_error: "Errore percorso"
      },
      ro: {
        title: "Starea mișcării după numărul de înmatriculare",
        plate_ph: "Introdu numărul (ex. AB-123-CD)",
        btn_check: "Verifică",
        btn_notify: "Activează notificări",
        btn_enabling: "Se activează…",
        btn_enabled: "Notificări active",

        getting_location: "Se obține locația…",
        loading_status: "Se încarcă statusul…",
        loading_route: "Se încarcă ruta…",

        no_movement: "Nu s-a găsit mișcarea",
        last_refresh: "Ultima actualizare",
        destination: "Destinație",
        departure_time: "Ora plecării",
        report_office: "Prezintă-te la birou",
        trailer: "Remorcă",
        place: "Loc",
        route_map: "Harta rutei",
        origin: "Origine",
        destination_pin: "Destinație",

        parking: "Parcare",
        dock: "Rampă",

        err_location: "Eroare locație",
        err_network: "Eroare de rețea",
        err_error: "Eroare",
        help_location: "Activează GPS-ul și permite accesul la locație.",

        notify_not_supported: "Notificări neacceptate",
        notify_not_supported_help: "Folosește Chrome/Edge pe Android. Pe iOS trebuie adăugat site-ul pe ecranul principal.",
        notify_denied: "Notificări refuzate",
        notify_denied_help: "Permite notificările în setările browserului.",
        notify_failed: "Abonarea a eșuat",
        notify_enabled_msg: "Notificări activate",
        notify_enabled_help: "Vei primi un push când se schimbă statusul.",
        subscribe_error: "Eroare abonare",
        route_error: "Eroare rută"
      },
      ru: {
        title: "Статус рейса по номеру",
        plate_ph: "Введите номер (например AB-123-CD)",
        btn_check: "Проверить",
        btn_notify: "Включить уведомления",
        btn_enabling: "Включение…",
        btn_enabled: "Уведомления включены",

        getting_location: "Получаем геолокацию…",
        loading_status: "Загружаем статус…",
        loading_route: "Загружаем маршрут…",

        no_movement: "Рейс не найден",
        last_refresh: "Последнее обновление",
        destination: "Пункт назначения",
        departure_time: "Время выезда",
        report_office: "Подойти в офис",
        trailer: "Прицеп",
        place: "Место",
        route_map: "Карта маршрута",
        origin: "Старт",
        destination_pin: "Назначение",

        parking: "Парковка",
        dock: "Док",

        err_location: "Ошибка геолокации",
        err_network: "Ошибка сети",
        err_error: "Ошибка",
        help_location: "Включите GPS и разрешите доступ к геолокации.",

        notify_not_supported: "Уведомления не поддерживаются",
        notify_not_supported_help: "Используйте Chrome/Edge на Android. На iOS добавьте сайт на главный экран.",
        notify_denied: "Уведомления запрещены",
        notify_denied_help: "Разрешите уведомления в настройках браузера.",
        notify_failed: "Подписка не удалась",
        notify_enabled_msg: "Уведомления включены",
        notify_enabled_help: "Вы получите push, когда статус изменится.",
        subscribe_error: "Ошибка подписки",
        route_error: "Ошибка маршрута"
      },
      lt: {
        title: "Judėjimo būsena pagal valstybinį numerį",
        plate_ph: "Įveskite numerį (pvz. AB-123-CD)",
        btn_check: "Tikrinti",
        btn_notify: "Įjungti pranešimus",
        btn_enabling: "Įjungiama…",
        btn_enabled: "Pranešimai įjungti",

        getting_location: "Gaunama vieta…",
        loading_status: "Įkeliama būsena…",
        loading_route: "Įkeliama trasa…",

        no_movement: "Judėjimas nerastas",
        last_refresh: "Paskutinis atnaujinimas",
        destination: "Paskirtis",
        departure_time: "Išvykimo laikas",
        report_office: "Atsižymėti biure",
        trailer: "Priekaba",
        place: "Vieta",
        route_map: "Maršruto žemėlapis",
        origin: "Pradžia",
        destination_pin: "Paskirtis",

        parking: "Parkingas",
        dock: "Dokas",

        err_location: "Vietos klaida",
        err_network: "Tinklo klaida",
        err_error: "Klaida",
        help_location: "Įjunkite GPS ir leiskite vietos leidimą.",

        notify_not_supported: "Pranešimai nepalaikomi",
        notify_not_supported_help: "Naudokite Chrome/Edge Android. iOS reikalauja pridėti svetainę į pagrindinį ekraną.",
        notify_denied: "Pranešimai atmesti",
        notify_denied_help: "Leiskite pranešimus naršyklės nustatymuose.",
        notify_failed: "Prenumerata nepavyko",
        notify_enabled_msg: "Pranešimai įjungti",
        notify_enabled_help: "Gausite push pranešimą, kai pasikeis būsena.",
        subscribe_error: "Prenumeratos klaida",
        route_error: "Maršruto klaida"
      },
      kk: {
        title: "Көлік нөмірі бойынша қозғалыс күйі",
        plate_ph: "Нөмірді енгізіңіз (мысалы AB-123-CD)",
        btn_check: "Тексеру",
        btn_notify: "Хабарландыруларды қосу",
        btn_enabling: "Қосылуда…",
        btn_enabled: "Хабарландырулар қосулы",

        getting_location: "Орналасу анықталуда…",
        loading_status: "Күй жүктелуде…",
        loading_route: "Маршрут жүктелуде…",

        no_movement: "Қозғалыс табылмады",
        last_refresh: "Соңғы жаңарту",
        destination: "Бағыт",
        departure_time: "Жөнелу уақыты",
        report_office: "Кеңсеге келу",
        trailer: "Тіркеме",
        place: "Орын",
        route_map: "Маршрут картасы",
        origin: "Бастау",
        destination_pin: "Бағыт",

        parking: "Тұрақ",
        dock: "Док",

        err_location: "Орналасу қатесі",
        err_network: "Желі қатесі",
        err_error: "Қате",
        help_location: "GPS-ті қосыңыз және геолокацияға рұқсат беріңіз.",

        notify_not_supported: "Хабарландырулар қолдау көрсетілмейді",
        notify_not_supported_help: "Android-та Chrome/Edge қолданыңыз. iOS-та сайтты Home Screen-ге қосу керек.",
        notify_denied: "Хабарландыруларға тыйым салынған",
        notify_denied_help: "Браузер баптауларында хабарландыруларды рұқсат етіңіз.",
        notify_failed: "Жазылу сәтсіз",
        notify_enabled_msg: "Хабарландырулар қосылды",
        notify_enabled_help: "Күй өзгерсе, push хабарлама аласыз.",
        subscribe_error: "Жазылу қатесі",
        route_error: "Маршрут қатесі"
      },
      hi: {
        title: "लाइसेंस प्लेट के अनुसार मूवमेंट स्टेटस",
        plate_ph: "लाइसेंस प्लेट दर्ज करें (जैसे AB-123-CD)",
        btn_check: "जाँचें",
        btn_notify: "सूचनाएँ सक्षम करें",
        btn_enabling: "सक्षम किया जा रहा है…",
        btn_enabled: "सूचनाएँ सक्षम",

        getting_location: "लोकेशन प्राप्त की जा रही है…",
        loading_status: "स्टेटस लोड हो रहा है…",
        loading_route: "रूट लोड हो रहा है…",

        no_movement: "कोई मूवमेंट नहीं मिला",
        last_refresh: "अंतिम अपडेट",
        destination: "गंतव्य",
        departure_time: "प्रस्थान समय",
        report_office: "ऑफिस में रिपोर्ट करें",
        trailer: "ट्रेलर",
        place: "स्थान",
        route_map: "रूट मैप",
        origin: "प्रारंभ",
        destination_pin: "गंतव्य",

        parking: "पार्किंग",
        dock: "डॉक",

        err_location: "लोकेशन त्रुटि",
        err_network: "नेटवर्क त्रुटि",
        err_error: "त्रुटि",
        help_location: "GPS चालू करें और लोकेशन अनुमति दें।",

        notify_not_supported: "सूचनाएँ समर्थित नहीं हैं",
        notify_not_supported_help: "Android पर Chrome/Edge उपयोग करें। iOS के लिए साइट को Home Screen पर जोड़ना आवश्यक है।",
        notify_denied: "सूचनाएँ अस्वीकृत",
        notify_denied_help: "ब्राउज़र सेटिंग्स में सूचनाएँ अनुमति दें।",
        notify_failed: "सब्सक्राइब विफल",
        notify_enabled_msg: "सूचनाएँ सक्षम",
        notify_enabled_help: "स्टेटस बदलने पर आपको push सूचना मिलेगी।",
        subscribe_error: "सब्सक्राइब त्रुटि",
        route_error: "रूट त्रुटि"
      },
      pl: {
        title: "Status ruchu według tablicy rejestracyjnej",
        plate_ph: "Wpisz rejestrację (np. AB-123-CD)",
        btn_check: "Sprawdź",
        btn_notify: "Włącz powiadomienia",
        btn_enabling: "Włączanie…",
        btn_enabled: "Powiadomienia włączone",

        getting_location: "Pobieranie lokalizacji…",
        loading_status: "Ładowanie statusu…",
        loading_route: "Ładowanie trasy…",

        no_movement: "Nie znaleziono ruchu",
        last_refresh: "Ostatnie odświeżenie",
        destination: "Cel",
        departure_time: "Czas odjazdu",
        report_office: "Zgłoś się do biura",
        trailer: "Naczepa",
        place: "Miejsce",
        route_map: "Mapa trasy",
        origin: "Start",
        destination_pin: "Cel",

        parking: "Parking",
        dock: "Dok",

        err_location: "Błąd lokalizacji",
        err_network: "Błąd sieci",
        err_error: "Błąd",
        help_location: "Włącz GPS i zezwól na dostęp do lokalizacji.",

        notify_not_supported: "Powiadomienia nieobsługiwane",
        notify_not_supported_help: "Użyj Chrome/Edge na Androidzie. iOS wymaga dodania strony do ekranu początkowego.",
        notify_denied: "Powiadomienia odrzucone",
        notify_denied_help: "Zezwól na powiadomienia w ustawieniach przeglądarki.",
        notify_failed: "Subskrypcja nie powiodła się",
        notify_enabled_msg: "Powiadomienia włączone",
        notify_enabled_help: "Otrzymasz push, gdy status się zmieni.",
        subscribe_error: "Błąd subskrypcji",
        route_error: "Błąd trasy"
      },
      hu: {
        title: "Mozgás státusz rendszám alapján",
        plate_ph: "Add meg a rendszámot (pl. AB-123-CD)",
        btn_check: "Ellenőrzés",
        btn_notify: "Értesítések bekapcsolása",
        btn_enabling: "Bekapcsolás…",
        btn_enabled: "Értesítések bekapcsolva",

        getting_location: "Helyzet lekérése…",
        loading_status: "Státusz betöltése…",
        loading_route: "Útvonal betöltése…",

        no_movement: "Nincs találat",
        last_refresh: "Utolsó frissítés",
        destination: "Célállomás",
        departure_time: "Indulási idő",
        report_office: "Jelentkezz az irodában",
        trailer: "Pótkocsi",
        place: "Hely",
        route_map: "Útvonal térkép",
        origin: "Kiindulás",
        destination_pin: "Cél",

        parking: "Parkoló",
        dock: "Dokk",

        err_location: "Helymeghatározási hiba",
        err_network: "Hálózati hiba",
        err_error: "Hiba",
        help_location: "Kapcsold be a GPS-t és engedélyezd a helyhozzáférést.",

        notify_not_supported: "Értesítések nem támogatottak",
        notify_not_supported_help: "Androidon Chrome/Edge ajánlott. iOS-en add a weboldalt a Főképernyőhöz.",
        notify_denied: "Értesítések letiltva",
        notify_denied_help: "Engedélyezd az értesítéseket a böngésző beállításaiban.",
        notify_failed: "Feliratkozás sikertelen",
        notify_enabled_msg: "Értesítések bekapcsolva",
        notify_enabled_help: "Push értesítést kapsz, ha a státusz változik.",
        subscribe_error: "Feliratkozási hiba",
        route_error: "Útvonal hiba"
      },
      uz: {
  title: "Davlat raqami bo‘yicha harakat holati",
  plate_ph: "Davlat raqamini kiriting (masalan AB-123-CD)",
  btn_check: "Tekshirish",
  btn_notify: "Bildirishnomalarni yoqish",
  btn_enabling: "Yoqilmoqda…",
  btn_enabled: "Bildirishnomalar yoqildi",

  getting_location: "Joylashuv olinmoqda…",
  loading_status: "Holat yuklanmoqda…",
  loading_route: "Marshrut yuklanmoqda…",

  no_movement: "Harakat topilmadi",
  last_refresh: "Oxirgi yangilanish",
  destination: "Manzil",
  departure_time: "Jo‘nash vaqti",
  report_office: "Ofisga murojaat qiling",
  trailer: "Treyler",
  place: "Joy",
  route_map: "Marshrut xaritasi",
  origin: "Boshlanish",
  destination_pin: "Manzil",

  parking: "Parkovka",
  dock: "Dok",

  err_location: "Joylashuv xatosi",
  err_network: "Tarmoq xatosi",
  err_error: "Xato",
  help_location: "GPS-ni yoqing va joylashuv ruxsatini bering.",

  notify_not_supported: "Bildirishnomalar qo‘llab-quvvatlanmaydi",
  notify_not_supported_help: "Androidda Chrome/Edge’dan foydalaning. iOS’da saytni Home Screen’ga qo‘shish kerak.",
  notify_denied: "Bildirishnomalar rad etildi",
  notify_denied_help: "Brauzer sozlamalarida bildirishnomalarga ruxsat bering.",
  notify_failed: "Obuna bo‘lish muvaffaqiyatsiz",
  notify_enabled_msg: "Bildirishnomalar yoqildi",
  notify_enabled_help: "Holat o‘zgarsa push xabar olasiz.",
  subscribe_error: "Obuna xatosi",
  route_error: "Marshrut xatosi"
      },
      tg: {
  title: "Ҳолати ҳаракат аз рӯи рақами мошин",
  plate_ph: "Рақамро ворид кунед (масалан AB-123-CD)",
  btn_check: "Санҷидан",
  btn_notify: "Фаъол кардани огоҳиномаҳо",
  btn_enabling: "Фаъол мешавад…",
  btn_enabled: "Огоҳиномаҳо фаъол шуданд",

  getting_location: "Ҷойгиршавӣ гирифта мешавад…",
  loading_status: "Ҳолат бор мешавад…",
  loading_route: "Масир бор мешавад…",

  no_movement: "Ҳаракат ёфт нашуд",
  last_refresh: "Охирин навсозӣ",
  destination: "Самт",
  departure_time: "Вақти баромад",
  report_office: "Ба офис ҳозир шавед",
  trailer: "Прицеп",
  place: "Ҷой",
  route_map: "Харитаи масир",
  origin: "Оғоз",
  destination_pin: "Самт",

  parking: "Парковка",
  dock: "Док",

  err_location: "Хатои ҷойгиршавӣ",
  err_network: "Хатои шабака",
  err_error: "Хато",
  help_location: "GPS-ро фаъол кунед ва иҷозати ҷойгиршавиро диҳед.",

  notify_not_supported: "Огоҳиномаҳо дастгирӣ намешаванд",
  notify_not_supported_help: "Дар Android Chrome/Edge истифода баред. Дар iOS сайтро ба Home Screen илова кунед.",
  notify_denied: "Огоҳиномаҳо рад шуданд",
  notify_denied_help: "Дар танзимоти браузер огоҳиномаҳоро иҷозат диҳед.",
  notify_failed: "Обуна шудан ноком шуд",
  notify_enabled_msg: "Огоҳиномаҳо фаъол шуданд",
  notify_enabled_help: "Ҳангоми тағйири ҳолат push мегиред.",
  subscribe_error: "Хатои обуна",
  route_error: "Хатои масир"
      },
      ky: {
  title: "Мамлекеттик номер боюнча кыймылдын абалы",
  plate_ph: "Номерди киргизиңиз (мисалы AB-123-CD)",
  btn_check: "Текшерүү",
  btn_notify: "Билдирмелерди күйгүзүү",
  btn_enabling: "Күйгүзүлүүдө…",
  btn_enabled: "Билдирмелер күйгүзүлдү",

  getting_location: "Жайгашкан жер алынууда…",
  loading_status: "Абалы жүктөлүүдө…",
  loading_route: "Маршрут жүктөлүүдө…",

  no_movement: "Кыймыл табылган жок",
  last_refresh: "Акыркы жаңыртуу",
  destination: "Багыт",
  departure_time: "Жөнөө убактысы",
  report_office: "Кеңсеге кайрылыңыз",
  trailer: "Чиркегич",
  place: "Жай",
  route_map: "Маршрут картасы",
  origin: "Башталыш",
  destination_pin: "Багыт",

  parking: "Токтотмо",
  dock: "Док",

  err_location: "Жайгашуу катасы",
  err_network: "Тармак катасы",
  err_error: "Ката",
  help_location: "GPSти күйгүзүп, геолокацияга уруксат бериңиз.",

  notify_not_supported: "Билдирмелер колдоого алынбайт",
  notify_not_supported_help: "Android'де Chrome/Edge колдонуңуз. iOS'то сайтты Home Screen'ге кошуңуз.",
  notify_denied: "Билдирмелерге тыюу салынды",
  notify_denied_help: "Браузердин жөндөөлөрүнөн билдирмелерге уруксат бериңиз.",
  notify_failed: "Жазылуу ийгиликсиз",
  notify_enabled_msg: "Билдирмелер күйгүзүлдү",
  notify_enabled_help: "Абалы өзгөрсө, push билдирүү аласыз.",
  subscribe_error: "Жазылуу катасы",
  route_error: "Маршрут катасы"
      },
      be: {
  title: "Статус руху pa нумары",
  plate_ph: "Увядзіце нумар (напрыклад AB-123-CD)",
  btn_check: "Праверыць",
  btn_notify: "Уключыць апавяшчэнні",
  btn_enabling: "Уключэнне…",
  btn_enabled: "Апавяшчэнні ўключаны",

  getting_location: "Атрымліваем месцазнаходжанне…",
  loading_status: "Загружаем статус…",
  loading_route: "Загружаем маршрут…",

  no_movement: "Рух не знойдзены",
  last_refresh: "Апошняе абнаўленне",
  destination: "Пункт прызначэння",
  departure_time: "Час выезду",
  report_office: "Зайдзіце ў офіс",
  trailer: "Прычэп",
  place: "Месца",
  route_map: "Карта маршруту",
  origin: "Старт",
  destination_pin: "Прызначэнне",

  parking: "Паркоўка",
  dock: "Док",

  err_location: "Памылка месцазнаходжання",
  err_network: "Памылка сеткі",
  err_error: "Памылка",
  help_location: "Уключыце GPS і дазвольце доступ да месцазнаходжання.",

  notify_not_supported: "Апавяшчэнні не падтрымліваюцца",
  notify_not_supported_help: "Выкарыстоўвайце Chrome/Edge на Android. На iOS дадайце сайт на Home Screen.",
  notify_denied: "Апавяшчэнні забароненыя",
  notify_denied_help: "Дазвольце апавяшчэнні ў наладах браўзера.",
  notify_failed: "Падпіска не атрымалася",
  notify_enabled_msg: "Апавяшчэнні ўключаны",
  notify_enabled_help: "Вы атрымаеце push, калі статус зменіцца.",
  subscribe_error: "Памылка падпіскі",
  route_error: "Памылка маршруту"
},

      fr: {
        title: "Statut du mouvement par plaque",
        plate_ph: "Saisir la plaque (ex. AB-123-CD)",
        btn_check: "Vérifier",
        btn_notify: "Activer les notifications",
        btn_enabling: "Activation…",
        btn_enabled: "Notifications activées",

        getting_location: "Récupération de la position…",
        loading_status: "Chargement du statut…",
        loading_route: "Chargement de l’itinéraire…",

        no_movement: "Aucun mouvement trouvé",
        last_refresh: "Dernière mise à jour",
        destination: "Destination",
        departure_time: "Heure de départ",
        report_office: "Se présenter au bureau",
        trailer: "Remorque",
        place: "Emplacement",
        route_map: "Carte de l’itinéraire",
        origin: "Départ",
        destination_pin: "Destination",

        parking: "Parking",
        dock: "Quai",

        err_location: "Erreur de localisation",
        err_network: "Erreur réseau",
        err_error: "Erreur",
        help_location: "Activez le GPS et autorisez l’accès à la localisation.",

        notify_not_supported: "Notifications non prises en charge",
        notify_not_supported_help: "Utilisez Chrome/Edge sur Android. Sur iOS, ajoutez le site à l’écran d’accueil.",
        notify_denied: "Notifications refusées",
        notify_denied_help: "Autorisez les notifications dans les paramètres du navigateur.",
        notify_failed: "Échec de l’abonnement",
        notify_enabled_msg: "Notifications activées",
        notify_enabled_help: "Vous recevrez une notification push lorsque votre statut change.",
        subscribe_error: "Erreur d’abonnement",
        route_error: "Erreur d’itinéraire"
      },
      tr: {
        title: "Plakaya göre hareket durumu",
        plate_ph: "Plakayı girin (örn. AB-123-CD)",
        btn_check: "Kontrol et",
        btn_notify: "Bildirimleri etkinleştir",
        btn_enabling: "Etkinleştiriliyor…",
        btn_enabled: "Bildirimler etkin",

        getting_location: "Konum alınıyor…",
        loading_status: "Durum yükleniyor…",
        loading_route: "Rota yükleniyor…",

        no_movement: "Hareket bulunamadı",
        last_refresh: "Son yenileme",
        destination: "Varış",
        departure_time: "Çıkış saati",
        report_office: "Ofise bildirin",
        trailer: "Dorse",
        place: "Yer",
        route_map: "Rota haritası",
        origin: "Başlangıç",
        destination_pin: "Varış",

        parking: "Park",
        dock: "Kapı",

        err_location: "Konum hatası",
        err_network: "Ağ hatası",
        err_error: "Hata",
        help_location: "GPS’i açın ve konum izni verin.",

        notify_not_supported: "Bildirimler desteklenmiyor",
        notify_not_supported_help: "Android’de Chrome/Edge kullanın. iOS’ta siteyi Ana Ekran’a eklemek gerekir.",
        notify_denied: "Bildirimler engellendi",
        notify_denied_help: "Tarayıcı ayarlarından bildirimlere izin verin.",
        notify_failed: "Abonelik başarısız",
        notify_enabled_msg: "Bildirimler etkin",
        notify_enabled_help: "Durumunuz değiştiğinde push bildirimi alacaksınız.",
        subscribe_error: "Abonelik hatası",
        route_error: "Rota hatası"
      },
      sv: {
        title: "Rörelsestatus per registreringsnummer",
        plate_ph: "Ange registreringsnummer (t.ex. AB-123-CD)",
        btn_check: "Kontrollera",
        btn_notify: "Aktivera aviseringar",
        btn_enabling: "Aktiverar…",
        btn_enabled: "Aviseringar aktiverade",

        getting_location: "Hämtar position…",
        loading_status: "Laddar status…",
        loading_route: "Laddar rutt…",

        no_movement: "Ingen rörelse hittades",
        last_refresh: "Senast uppdaterad",
        destination: "Destination",
        departure_time: "Avgångstid",
        report_office: "Anmäl dig på kontoret",
        trailer: "Släp",
        place: "Plats",
        route_map: "Ruttkarta",
        origin: "Start",
        destination_pin: "Destination",

        parking: "Parkering",
        dock: "Port",

        err_location: "Positionsfel",
        err_network: "Nätverksfel",
        err_error: "Fel",
        help_location: "Aktivera GPS och tillåt platsbehörighet.",

        notify_not_supported: "Aviseringar stöds inte",
        notify_not_supported_help: "Använd Chrome/Edge på Android. iOS kräver att du lägger till sidan på hemskärmen.",
        notify_denied: "Aviseringar nekade",
        notify_denied_help: "Tillåt aviseringar i webbläsarens inställningar.",
        notify_failed: "Prenumeration misslyckades",
        notify_enabled_msg: "Aviseringar aktiverade",
        notify_enabled_help: "Du får en push-notis när din status ändras.",
        subscribe_error: "Prenumerationsfel",
        route_error: "Ruttfel"
      }

    };


    function tryLang(v) {
      try {
        const s0 = String(v || "").trim().toLowerCase().replaceAll("_", "-");
        const base = s0.split("-", 1)[0];
        if (SUPPORTED_LANGS.includes(base)) return base;
        // common aliases
        if (base === "kz" || base === "kaz") return "kk";
        if (base === "uzb") return "uz";
        if (base === "tgk" || base === "taj" || base === "tj") return "tg";
        if (base === "kir" || base === "kg") return "ky";
        if (base === "bel" || base === "by") return "be";
                if (base === "fre" || base === "fra") return "fr";
        if (base === "tur") return "tr";
        if (base === "swe") return "sv";
        return "";
      } catch (e) {
        return "";
      }
    }

    function normLang(v) {
      return tryLang(v) || "en";
    }

    let CURRENT_LANG = "en";

    const HOUSE_RULES = {
      en: {
        title: "QAR Duiven | Driving and Walking Routes",
        intro: "Please read the site map and safety rules. You must accept the house rules to continue.",
        accept: "I accept the house rules",
        cont: "Continue",
        open_pdf: "Open PDF in a new tab",
        html: `<img class="hr-img" src="/static/house_rules/01_site_map.png" alt="Site map" />
<h3>Site map – key locations</h3>
<ul>
  <li><b>Offices/driver facilities</b>: A – Gatehouse / driver lounge; B – Driver briefing &amp; paperwork office</li>
  <li><b>Parking areas</b>: C – Driver rest parking; D – Trailer parking; E – Relay parking → truck trailers; F – Relay parking → solo trucks; G – Visitor parking</li>
  <li><b>Pedestrians</b>: H – Entry to driver lounge; I – Stairs to office walkway; J – Pedestrian gate to office</li>
  <li><b>Barrier gates</b>: K – Truck entrance barrier; L – Exit yard to driver parking; M – Entry to yard from driver area; N – Visitor parking barrier; O – Truck exit barrier</li>
</ul>

<h3>DRIVING – enter/exit the driver rest parking area</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/02_after_gatehouse.png" alt="After passing the gatehouse" />
  After passing the gatehouse (A), drive straight to the driver rest parking area (C).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/03_enter_driver_rest_parking.png" alt="Enter the driver rest parking area" />
  Enter the driver rest parking area through the barrier gate (L).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/04_go_through_barrier_gate.png" alt="Go through the barrier gate" />
  Go through the barrier gate (M) to exit the driver rest parking area and access the yard.
</li>
</ul>

<h3>WALKING</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/05_stairs_leading_from.png" alt="Stairs leading from" />
  <img class="hr-img" src="/static/house_rules/06_walk_along_fence.png" alt="Walk along the fence" />
  <img class="hr-img" src="/static/house_rules/07_pass_through_gate.png" alt="Pass through the gate" />
  <b>From the driver lounge to the office</b>: take the stairs (I) from the driver lounge (A) to the walkway, walk along the fence and follow the signs, then pass through gate (J) to reach the briefing office (B).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/08_walk_along_path.png" alt="Walk along the path" />
  <img class="hr-img" src="/static/house_rules/09_continue_straight.png" alt="Continue straight" />
  <b>From the relay parking area to the office</b>: walk along the path with green markings next to the fence, continue straight, then pass through gate (J) to reach the briefing office (B).
</li>
</ul>

<h3>In case of emergency</h3>
<ul>
  <li>Stay calm.</li>
  <li>Call <b>112</b> first.</li>
  <li>Call FedEx security: <b>0316 799 476</b>.</li>
  <li>In case of fire or a fire alarm: leave via the nearest emergency exit and go to the assembly point.</li>
</ul>

<h3>Site safety rules</h3>
<ul>
  <li>Speed limit on site: <b>15 km/h</b>.</li>
  <li>One-way traffic on the yard.</li>
  <li>Walk only on designated pedestrian paths.</li>
  <li>Wear a closed high-visibility safety vest and safety shoes (class S3) on the yard.</li>
  <li>Use appropriate safety equipment when docking or handling trailers.</li>
  <li>Always follow local site rules and the instructions of security personnel and shunters.</li>
</ul>
<div class="muted">For full site rules, see the FedEx Duiven–QAR site yard rules leaflet (available at the gatehouse and the office).</div>`
      },
      de: {
        title: "QAR Duiven | Fahr- und Fußwege",
        intro: "Bitte lesen Sie den Lageplan und die Sicherheitsregeln. Sie müssen die Hofregeln akzeptieren, um fortzufahren.",
        accept: "Ich akzeptiere die Hofregeln",
        cont: "Weiter",
        open_pdf: "PDF in neuem Tab öffnen",
        html: `<img class="hr-img" src="/static/house_rules/01_site_map.png" alt="Site map" />
<h3>Lageplan – wichtige Orte</h3>
<ul>
  <li><b>Büros/Fahrerbereiche</b>: A – Pförtnerhaus / Fahrerlounge; B – Fahrerbriefing &amp; Büro für Papiere</li>
  <li><b>Parkflächen</b>: C – Fahrer-Ruheparkplatz; D – Trailer-Parkplatz; E – Relay-Parkplatz → Lkw mit Trailern; F – Relay-Parkplatz → Solo-Trucks; G – Besucherparkplatz</li>
  <li><b>Fußgänger</b>: H – Eingang Fahrerlounge; I – Treppe zum Fußweg Richtung Büro; J – Fußgängertor zum Büro</li>
  <li><b>Schranken</b>: K – Lkw-Einfahrtsschranke; L – Ausfahrt Hof zum Fahrerparkplatz; M – Einfahrt Hof vom Fahrerbereich; N – Schranke Besucherparkplatz; O – Lkw-Ausfahrtsschranke</li>
</ul>

<h3>FAHREN – Ein-/Ausfahrt Fahrer-Ruheparkplatz</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/02_after_gatehouse.png" alt="After passing the gatehouse" />
  Nach dem Pförtnerhaus (A) geradeaus zum Fahrer-Ruheparkplatz (C) fahren.
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/03_enter_driver_rest_parking.png" alt="Enter the driver rest parking area" />
  Durch die Schranke (L) in den Fahrer-Ruheparkplatz einfahren.
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/04_go_through_barrier_gate.png" alt="Go through the barrier gate" />
  Durch die Schranke (M) ausfahren, um den Hof zu erreichen.
</li>
</ul>

<h3>ZU FUß</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/05_stairs_leading_from.png" alt="Stairs leading from" />
  <img class="hr-img" src="/static/house_rules/06_walk_along_fence.png" alt="Walk along the fence" />
  <img class="hr-img" src="/static/house_rules/07_pass_through_gate.png" alt="Pass through the gate" />
  <b>Von der Fahrerlounge zum Büro</b>: über die Treppe (I) von der Lounge (A) auf den Fußweg, am Zaun entlanggehen, Beschilderung folgen und durch Tor (J) zum Briefingbüro (B).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/08_walk_along_path.png" alt="Walk along the path" />
  <img class="hr-img" src="/static/house_rules/09_continue_straight.png" alt="Continue straight" />
  <b>Vom Relay-Parkplatz zum Büro</b>: dem Weg mit grünen Markierungen am Zaun folgen, geradeaus weiter und durch Tor (J) zum Briefingbüro (B).
</li>
</ul>

<h3>Im Notfall</h3>
<ul>
  <li>Ruhig bleiben.</li>
  <li>Zuerst <b>112</b> anrufen.</li>
  <li>FedEx Security anrufen: <b>0316 799 476</b>.</li>
  <li>Bei Feuer oder Feueralarm: über den nächsten Notausgang verlassen und zum Sammelplatz gehen.</li>
</ul>

<h3>Sicherheitsregeln auf dem Gelände</h3>
<ul>
  <li>Höchstgeschwindigkeit: <b>15 km/h</b>.</li>
  <li>Einbahnverkehr auf dem Hof.</li>
  <li>Nur auf ausgewiesenen Fußwegen gehen.</li>
  <li>Auf dem Hof geschlossene Warnweste und Sicherheitsschuhe (Klasse S3) tragen.</li>
  <li>Beim Andocken/Handling von Trailern geeignete Schutzausrüstung verwenden.</li>
  <li>Lokale Regeln sowie Anweisungen von Security und Rangierern immer befolgen.</li>
</ul>
<div class="muted">Vollständige Regeln: „FedEx Duiven–QAR site yard rules“ (erhältlich am Pförtnerhaus und im Büro).</div>`
      },
      nl: {
        title: "QAR Duiven | Rij- en looproutes",
        intro: "Lees de plattegrond en veiligheidsregels. Je moet de huisregels accepteren om verder te gaan.",
        accept: "Ik accepteer de huisregels",
        cont: "Doorgaan",
        open_pdf: "PDF openen in een nieuw tabblad",
        html: `<img class="hr-img" src="/static/house_rules/01_site_map.png" alt="Site map" />
<h3>Plattegrond – belangrijke locaties</h3>
<ul>
  <li><b>Kantoren/chauffeursfaciliteiten</b>: A – Poortgebouw / chauffeurslounge; B – Chauffeursbriefing &amp; papierwerk-kantoor</li>
  <li><b>Parkeerplaatsen</b>: C – Rustparking chauffeurs; D – Trailerparking; E – Relay parking → trucks met trailers; F – Relay parking → solo trucks; G – Bezoekersparking</li>
  <li><b>Voetgangers</b>: H – Ingang chauffeurslounge; I – Trap naar looproute richting kantoor; J – Voetgangerspoort naar kantoor</li>
  <li><b>Slagbomen</b>: K – Slagboom truck-ingang; L – Uitrit terrein naar chauffeursparking; M – Inrit terrein vanaf chauffeursgebied; N – Slagboom bezoekersparking; O – Slagboom truck-uitgang</li>
</ul>

<h3>RIJDEN – in-/uitrijden rustparking chauffeurs</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/02_after_gatehouse.png" alt="After passing the gatehouse" />
  Na het poortgebouw (A) rechtdoor naar de rustparking (C).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/03_enter_driver_rest_parking.png" alt="Enter the driver rest parking area" />
  Rijd via de slagboom (L) de rustparking binnen.
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/04_go_through_barrier_gate.png" alt="Go through the barrier gate" />
  Ga via de slagboom (M) naar buiten om het terrein op te rijden.
</li>
</ul>

<h3>LOPEN</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/05_stairs_leading_from.png" alt="Stairs leading from" />
  <img class="hr-img" src="/static/house_rules/06_walk_along_fence.png" alt="Walk along the fence" />
  <img class="hr-img" src="/static/house_rules/07_pass_through_gate.png" alt="Pass through the gate" />
  <b>Van de lounge naar het kantoor</b>: neem de trap (I) vanuit de lounge (A) naar de looproute, loop langs het hek en volg de borden, ga daarna door poort (J) naar het briefingkantoor (B).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/08_walk_along_path.png" alt="Walk along the path" />
  <img class="hr-img" src="/static/house_rules/09_continue_straight.png" alt="Continue straight" />
  <b>Van de relay parking naar het kantoor</b>: volg het pad met groene markeringen langs het hek, ga rechtdoor en ga door poort (J) naar het briefingkantoor (B).
</li>
</ul>

<h3>In geval van nood</h3>
<ul>
  <li>Blijf kalm.</li>
  <li>Bel eerst <b>112</b>.</li>
  <li>Bel FedEx security: <b>0316 799 476</b>.</li>
  <li>Bij brand of brandalarm: via de dichtstbijzijnde nooduitgang naar buiten en naar het verzamelpunt.</li>
</ul>

<h3>Veiligheidsregels op het terrein</h3>
<ul>
  <li>Snelheidslimiet: <b>15 km/u</b>.</li>
  <li>Eenrichtingsverkeer op het terrein.</li>
  <li>Loop alleen op aangewezen voetpaden.</li>
  <li>Draag een gesloten veiligheidsvest (hi-vis) en veiligheidsschoenen (klasse S3) op het terrein.</li>
  <li>Gebruik passende veiligheidsmiddelen bij het docken of verplaatsen van trailers.</li>
  <li>Volg altijd de lokale regels en instructies van security en shunters.</li>
</ul>
<div class="muted">Voor volledige regels: FedEx Duiven–QAR terreinregels leaflet (verkrijgbaar bij het poortgebouw en het kantoor).</div>`
      },
      fr: {
        title: "QAR Duiven | Itinéraires en véhicule et à pied",
        intro: "Veuillez lire le plan du site et les règles de sécurité. Vous devez accepter le règlement intérieur pour continuer.",
        accept: "J’accepte le règlement intérieur",
        cont: "Continuer",
        open_pdf: "Ouvrir le PDF dans un nouvel onglet",
        html: `<img class="hr-img" src="/static/house_rules/01_site_map.png" alt="Site map" />
<h3>Plan du site – lieux clés</h3>
<ul>
  <li><b>Bureaux / installations chauffeurs</b> : A – Poste de garde / lounge chauffeurs ; B – Briefing chauffeurs &amp; bureau des documents</li>
  <li><b>Zones de parking</b> : C – Parking repos chauffeurs ; D – Parking remorques ; E – Parking relais → tracteurs + remorques ; F – Parking relais → tracteurs seuls ; G – Parking visiteurs</li>
  <li><b>Piétons</b> : H – Entrée lounge chauffeurs ; I – Escaliers vers la passerelle du bureau ; J – Portillon piéton vers le bureau</li>
  <li><b>Barrières</b> : K – Barrière entrée camions ; L – Sortie cour vers parking chauffeurs ; M – Entrée cour depuis zone chauffeurs ; N – Barrière parking visiteurs ; O – Barrière sortie camions</li>
</ul>

<h3>CONDUITE – entrer/sortir du parking repos chauffeurs</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/02_after_gatehouse.png" alt="After passing the gatehouse" />
  Après le poste de garde (A), roulez tout droit jusqu’au parking repos chauffeurs (C).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/03_enter_driver_rest_parking.png" alt="Enter the driver rest parking area" />
  Entrez dans le parking repos chauffeurs via la barrière (L).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/04_go_through_barrier_gate.png" alt="Go through the barrier gate" />
  Passez la barrière (M) pour sortir du parking repos et accéder à la cour.
</li>
</ul>

<h3>À PIED</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/05_stairs_leading_from.png" alt="Stairs leading from" />
  <img class="hr-img" src="/static/house_rules/06_walk_along_fence.png" alt="Walk along the fence" />
  <img class="hr-img" src="/static/house_rules/07_pass_through_gate.png" alt="Pass through the gate" />
  <b>Du lounge au bureau</b> : prenez les escaliers (I) depuis le lounge (A) vers la passerelle, marchez le long de la clôture et suivez la signalisation, puis passez le portillon (J) pour atteindre le bureau de briefing (B).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/08_walk_along_path.png" alt="Walk along the path" />
  <img class="hr-img" src="/static/house_rules/09_continue_straight.png" alt="Continue straight" />
  <b>Du parking relais au bureau</b> : suivez le chemin avec les marquages verts le long de la clôture, continuez tout droit, puis passez le portillon (J) pour atteindre le bureau de briefing (B).
</li>
</ul>

<h3>En cas d’urgence</h3>
<ul>
  <li>Restez calme.</li>
  <li>Appelez d’abord le <b>112</b>.</li>
  <li>Appelez la sécurité FedEx : <b>0316 799 476</b>.</li>
  <li>En cas d’incendie ou d’alarme incendie : sortez par l’issue de secours la plus proche et rendez-vous au point de rassemblement.</li>
</ul>

<h3>Règles de sécurité sur site</h3>
<ul>
  <li>Limite de vitesse : <b>15 km/h</b>.</li>
  <li>Circulation à sens unique dans la cour.</li>
  <li>Marchez uniquement sur les cheminements piétons dédiés.</li>
  <li>Portez un gilet haute visibilité fermé et des chaussures de sécurité (classe S3) dans la cour.</li>
  <li>Utilisez l’équipement de sécurité approprié lors de l’accostage/manutention des remorques.</li>
  <li>Respectez toujours les règles locales et les consignes du personnel de sécurité et des shunters.</li>
</ul>
<div class="muted">Pour les règles complètes, consultez la brochure « FedEx Duiven–QAR site yard rules » (disponible au poste de garde et au bureau).</div>`
      },
      tr: {
        title: "QAR Duiven | Araç ve Yaya Güzergâhları",
        intro: "Lütfen saha haritasını ve güvenlik kurallarını okuyun. Devam etmek için tesis kurallarını kabul etmelisiniz.",
        accept: "Tesis kurallarını kabul ediyorum",
        cont: "Devam",
        open_pdf: "PDF’yi yeni sekmede aç",
        html: `<img class="hr-img" src="/static/house_rules/01_site_map.png" alt="Site map" />
<h3>Saha haritası – önemli noktalar</h3>
<ul>
  <li><b>Ofisler/şoför alanları</b>: A – Güvenlik kulübesi / şoför dinlenme alanı; B – Şoför brifingi &amp; evrak ofisi</li>
  <li><b>Park alanları</b>: C – Şoför dinlenme parkı; D – Dorse parkı; E – Röle parkı → dorse bağlı çekiciler; F – Röle parkı → solo çekiciler; G – Ziyaretçi parkı</li>
  <li><b>Yaya</b>: H – Şoför alanına giriş; I – Ofise giden yürüyüş yoluna merdiven; J – Ofise yaya kapısı</li>
  <li><b>Bariyer kapıları</b>: K – Kamyon giriş bariyeri; L – Saha çıkışı → şoför parkı; M – Şoför alanından sahaya giriş; N – Ziyaretçi parkı bariyeri; O – Kamyon çıkış bariyeri</li>
</ul>

<h3>SÜRÜŞ – şoför dinlenme parkına giriş/çıkış</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/02_after_gatehouse.png" alt="After passing the gatehouse" />
  Güvenlik kulübesini (A) geçtikten sonra düz devam edip şoför dinlenme parkına (C) gidin.
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/03_enter_driver_rest_parking.png" alt="Enter the driver rest parking area" />
  Bariyer kapısından (L) geçerek dinlenme parkına girin.
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/04_go_through_barrier_gate.png" alt="Go through the barrier gate" />
  Dinlenme parkından çıkıp sahaya erişmek için bariyer kapısından (M) geçin.
</li>
</ul>

<h3>YÜRÜYÜŞ</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/05_stairs_leading_from.png" alt="Stairs leading from" />
  <img class="hr-img" src="/static/house_rules/06_walk_along_fence.png" alt="Walk along the fence" />
  <img class="hr-img" src="/static/house_rules/07_pass_through_gate.png" alt="Pass through the gate" />
  <b>Şoför alanından ofise</b>: şoför alanından (A) merdivenleri (I) kullanarak yürüyüş yoluna çıkın, çit boyunca ilerleyip tabelaları takip edin, ardından kapıdan (J) geçerek brifing ofisine (B) ulaşın.
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/08_walk_along_path.png" alt="Walk along the path" />
  <img class="hr-img" src="/static/house_rules/09_continue_straight.png" alt="Continue straight" />
  <b>Röle parkından ofise</b>: çitin yanındaki yeşil işaretli yolu takip edin, düz devam edin, ardından kapıdan (J) geçerek brifing ofisine (B) ulaşın.
</li>
</ul>

<h3>Acil durumda</h3>
<ul>
  <li>Sakin olun.</li>
  <li>Önce <b>112</b>’yi arayın.</li>
  <li>FedEx güvenliğini arayın: <b>0316 799 476</b>.</li>
  <li>Yangın veya yangın alarmında: en yakın acil çıkıştan çıkın ve toplanma noktasına gidin.</li>
</ul>

<h3>Saha güvenlik kuralları</h3>
<ul>
  <li>Hız limiti: <b>15 km/sa</b>.</li>
  <li>Sahada tek yön trafik vardır.</li>
  <li>Sadece belirlenmiş yaya yollarını kullanın.</li>
  <li>Sahada kapalı reflektörlü yelek ve güvenlik ayakkabısı (S3) giyin.</li>
  <li>Dorseleri yanaştırırken/taşırken uygun güvenlik ekipmanı kullanın.</li>
  <li>Yerel kurallara ve güvenlik personeli ile shunter talimatlarına her zaman uyun.</li>
</ul>
<div class="muted">Tüm kurallar için: FedEx Duiven–QAR saha kuralları broşürü (güvenlik kulübesi ve ofiste mevcut).</div>`
      },
      sv: {
        title: "QAR Duiven | Kör- och gångvägar",
        intro: "Läs kartan och säkerhetsreglerna. Du måste godkänna platsreglerna för att fortsätta.",
        accept: "Jag godkänner platsreglerna",
        cont: "Fortsätt",
        open_pdf: "Öppna PDF i ny flik",
        html: `<img class="hr-img" src="/static/house_rules/01_site_map.png" alt="Site map" />
<h3>Karta – viktiga platser</h3>
<ul>
  <li><b>Kontor/förarfaciliteter</b>: A – Vakthus / förarlounge; B – Förarbriefing &amp; kontor för dokument</li>
  <li><b>Parkeringsområden</b>: C – Förarviloparkering; D – Släpparkering; E – Relay-parkering → dragbil + släp; F – Relay-parkering → enbart dragbil; G – Besöksparkering</li>
  <li><b>Gående</b>: H – Ingång till förarlounge; I – Trappor till gångvägen mot kontoret; J – Gånggrind till kontoret</li>
  <li><b>Bommar</b>: K – Bom vid lastbilsinfart; L – Utfart gård → förarparkering; M – Infart gård från förarområdet; N – Bom vid besöksparkering; O – Bom vid lastbilsutfart</li>
</ul>

<h3>KÖRNING – in/ut från förarviloparkeringen</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/02_after_gatehouse.png" alt="After passing the gatehouse" />
  Efter vakthuset (A), kör rakt fram till förarviloparkeringen (C).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/03_enter_driver_rest_parking.png" alt="Enter the driver rest parking area" />
  Kör in på förarviloparkeringen via bommen (L).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/04_go_through_barrier_gate.png" alt="Go through the barrier gate" />
  Passera bommen (M) för att lämna förarviloparkeringen och komma ut på gården.
</li>
</ul>

<h3>GÅNG</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/05_stairs_leading_from.png" alt="Stairs leading from" />
  <img class="hr-img" src="/static/house_rules/06_walk_along_fence.png" alt="Walk along the fence" />
  <img class="hr-img" src="/static/house_rules/07_pass_through_gate.png" alt="Pass through the gate" />
  <b>Från loungen till kontoret</b>: ta trapporna (I) från loungen (A) till gångvägen, gå längs stängslet och följ skyltarna, passera sedan grinden (J) till briefingkontoret (B).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/08_walk_along_path.png" alt="Walk along the path" />
  <img class="hr-img" src="/static/house_rules/09_continue_straight.png" alt="Continue straight" />
  <b>Från relay-parkeringen till kontoret</b>: följ vägen med gröna markeringar längs stängslet, fortsätt rakt fram och passera grinden (J) till briefingkontoret (B).
</li>
</ul>

<h3>Vid nödsituation</h3>
<ul>
  <li>Behåll lugnet.</li>
  <li>Ring först <b>112</b>.</li>
  <li>Ring FedEx säkerhet: <b>0316 799 476</b>.</li>
  <li>Vid brand eller brandlarm: lämna via närmaste nödutgång och gå till återsamlingsplatsen.</li>
</ul>

<h3>Säkerhetsregler på området</h3>
<ul>
  <li>Hastighetsgräns: <b>15 km/h</b>.</li>
  <li>Enkelriktad trafik på gården.</li>
  <li>Gå endast på markerade gångvägar.</li>
  <li>Använd stängd varselväst och skyddsskor (klass S3) på gården.</li>
  <li>Använd lämplig skyddsutrustning vid dockning eller hantering av släp.</li>
  <li>Följ alltid lokala regler och instruktioner från säkerhetspersonal och shunters.</li>
</ul>
<div class="muted">För fullständiga regler: FedEx Duiven–QAR område-regler (broschyr finns vid vakthuset och kontoret).</div>`
      },
      es: {
        title: "QAR Duiven | Rutas de conducción y a pie",
        intro: "Lea el plano del sitio y las normas de seguridad. Debe aceptar las normas del sitio para continuar.",
        accept: "Acepto las normas del sitio",
        cont: "Continuar",
        open_pdf: "Abrir PDF en una pestaña nueva",
        html: `<img class="hr-img" src="/static/house_rules/01_site_map.png" alt="Site map" />
<h3>Mapa del sitio – ubicaciones clave</h3>
<ul>
  <li><b>Oficinas/instalaciones para conductores</b>: A – Garita / sala de conductores; B – Oficina de briefing y documentación</li>
  <li><b>Zonas de aparcamiento</b>: C – Aparcamiento de descanso; D – Aparcamiento de remolques; E – Aparcamiento relay → camión con remolque; F – Aparcamiento relay → camión solo; G – Aparcamiento de visitantes</li>
  <li><b>Peatones</b>: H – Entrada a la sala; I – Escaleras a la pasarela hacia la oficina; J – Puerta peatonal a la oficina</li>
  <li><b>Barreras</b>: K – Barrera de entrada camiones; L – Salida del patio al aparcamiento de conductores; M – Entrada al patio desde zona de conductores; N – Barrera de visitantes; O – Barrera de salida camiones</li>
</ul>

<h3>CONDUCCIÓN – entrar/salir del aparcamiento de descanso</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/02_after_gatehouse.png" alt="After passing the gatehouse" />
  Después de la garita (A), conduzca recto hasta el aparcamiento de descanso (C).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/03_enter_driver_rest_parking.png" alt="Enter the driver rest parking area" />
  Entre al aparcamiento de descanso por la barrera (L).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/04_go_through_barrier_gate.png" alt="Go through the barrier gate" />
  Pase por la barrera (M) para salir del aparcamiento de descanso y acceder al patio.
</li>
</ul>

<h3>A PIE</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/05_stairs_leading_from.png" alt="Stairs leading from" />
  <img class="hr-img" src="/static/house_rules/06_walk_along_fence.png" alt="Walk along the fence" />
  <img class="hr-img" src="/static/house_rules/07_pass_through_gate.png" alt="Pass through the gate" />
  <b>De la sala a la oficina</b>: use las escaleras (I) desde la sala (A) hacia la pasarela, camine junto a la valla y siga las señales, luego pase por la puerta (J) para llegar a la oficina de briefing (B).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/08_walk_along_path.png" alt="Walk along the path" />
  <img class="hr-img" src="/static/house_rules/09_continue_straight.png" alt="Continue straight" />
  <b>Del aparcamiento relay a la oficina</b>: siga el camino con marcas verdes junto a la valla, continúe recto y pase por la puerta (J) para llegar a la oficina de briefing (B).
</li>
</ul>

<h3>En caso de emergencia</h3>
<ul>
  <li>Mantenga la calma.</li>
  <li>Llame primero al <b>112</b>.</li>
  <li>Llame a seguridad de FedEx: <b>0316 799 476</b>.</li>
  <li>En caso de incendio o alarma: salga por la salida de emergencia más cercana y vaya al punto de reunión.</li>
</ul>

<h3>Normas de seguridad del sitio</h3>
<ul>
  <li>Límite de velocidad: <b>15 km/h</b>.</li>
  <li>Tráfico de sentido único en el patio.</li>
  <li>Caminar solo por los caminos peatonales señalizados.</li>
  <li>Use chaleco de alta visibilidad cerrado y calzado de seguridad (clase S3) en el patio.</li>
  <li>Use el equipo de seguridad adecuado al acoplar o manipular remolques.</li>
  <li>Siga siempre las normas locales y las instrucciones del personal de seguridad y shunters.</li>
</ul>
<div class="muted">Para las normas completas, consulte el folleto de reglas del patio FedEx Duiven–QAR (disponible en la garita y la oficina).</div>`
      },
      it: {
        title: "QAR Duiven | Percorsi di guida e a piedi",
        intro: "Leggi la mappa del sito e le regole di sicurezza. Devi accettare il regolamento per continuare.",
        accept: "Accetto il regolamento del sito",
        cont: "Continua",
        open_pdf: "Apri il PDF in una nuova scheda",
        html: `<img class="hr-img" src="/static/house_rules/01_site_map.png" alt="Site map" />
<h3>Mappa del sito – posizioni chiave</h3>
<ul>
  <li><b>Uffici/servizi autisti</b>: A – Guardhouse / lounge autisti; B – Briefing autisti &amp; ufficio documenti</li>
  <li><b>Aree di parcheggio</b>: C – Parcheggio riposo autisti; D – Parcheggio rimorchi; E – Parcheggio relay → camion con rimorchio; F – Parcheggio relay → camion solo; G – Parcheggio visitatori</li>
  <li><b>Pedoni</b>: H – Ingresso lounge; I – Scale verso la passerella per l’ufficio; J – Cancello pedonale per l’ufficio</li>
  <li><b>Barriere</b>: K – Barriera ingresso camion; L – Uscita cortile verso parcheggio autisti; M – Ingresso cortile dall’area autisti; N – Barriera parcheggio visitatori; O – Barriera uscita camion</li>
</ul>

<h3>GUIDA – entrare/uscire dal parcheggio riposo autisti</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/02_after_gatehouse.png" alt="After passing the gatehouse" />
  Dopo la guardhouse (A), prosegui dritto fino al parcheggio riposo (C).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/03_enter_driver_rest_parking.png" alt="Enter the driver rest parking area" />
  Entra nel parcheggio riposo passando dalla barriera (L).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/04_go_through_barrier_gate.png" alt="Go through the barrier gate" />
  Passa la barriera (M) per uscire dal parcheggio riposo e accedere al cortile.
</li>
</ul>

<h3>A PIEDI</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/05_stairs_leading_from.png" alt="Stairs leading from" />
  <img class="hr-img" src="/static/house_rules/06_walk_along_fence.png" alt="Walk along the fence" />
  <img class="hr-img" src="/static/house_rules/07_pass_through_gate.png" alt="Pass through the gate" />
  <b>Dalla lounge all’ufficio</b>: usa le scale (I) dalla lounge (A) alla passerella, cammina lungo la recinzione seguendo i cartelli, poi attraversa il cancello (J) per arrivare all’ufficio briefing (B).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/08_walk_along_path.png" alt="Walk along the path" />
  <img class="hr-img" src="/static/house_rules/09_continue_straight.png" alt="Continue straight" />
  <b>Dal parcheggio relay all’ufficio</b>: segui il percorso con segnaletica verde lungo la recinzione, continua dritto e attraversa il cancello (J) per arrivare all’ufficio briefing (B).
</li>
</ul>

<h3>In caso di emergenza</h3>
<ul>
  <li>Mantieni la calma.</li>
  <li>Chiama prima il <b>112</b>.</li>
  <li>Chiama la sicurezza FedEx: <b>0316 799 476</b>.</li>
  <li>In caso di incendio o allarme: esci dall’uscita di emergenza più vicina e raggiungi il punto di raccolta.</li>
</ul>

<h3>Regole di sicurezza del sito</h3>
<ul>
  <li>Limite di velocità: <b>15 km/h</b>.</li>
  <li>Traffico a senso unico nel cortile.</li>
  <li>Cammina solo sui percorsi pedonali dedicati.</li>
  <li>Indossa gilet ad alta visibilità chiuso e scarpe di sicurezza (classe S3) nel cortile.</li>
  <li>Usa l’attrezzatura di sicurezza adeguata durante l’aggancio o la movimentazione dei rimorchi.</li>
  <li>Segui sempre le regole locali e le istruzioni del personale di sicurezza e degli shunters.</li>
</ul>
<div class="muted">Per le regole complete, vedi il leaflet “FedEx Duiven–QAR site yard rules” (disponibile alla guardhouse e in ufficio).</div>`
      },
      ro: {
        title: "QAR Duiven | Rute de condus și de mers pe jos",
        intro: "Citește harta site-ului și regulile de siguranță. Trebuie să accepți regulamentul pentru a continua.",
        accept: "Accept regulamentul (house rules)",
        cont: "Continuă",
        open_pdf: "Deschide PDF-ul într-un tab nou",
        html: `<img class="hr-img" src="/static/house_rules/01_site_map.png" alt="Site map" />
<h3>Harta site-ului – locații cheie</h3>
<ul>
  <li><b>Birouri/facilități șoferi</b>: A – Poartă / lounge șoferi; B – Briefing șoferi &amp; birou documente</li>
  <li><b>Parcări</b>: C – Parcare odihnă șoferi; D – Parcare remorci; E – Parcare relay → cap tractor + remorcă; F – Parcare relay → cap tractor singur; G – Parcare vizitatori</li>
  <li><b>Pietoni</b>: H – Intrare lounge; I – Scări către pasarela spre birou; J – Poartă pietonală către birou</li>
  <li><b>Bariere</b>: K – Barieră intrare camioane; L – Ieșire curte către parcare șoferi; M – Intrare curte din zona șoferi; N – Barieră parcare vizitatori; O – Barieră ieșire camioane</li>
</ul>

<h3>CONDUS – intrare/ieșire parcare odihnă șoferi</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/02_after_gatehouse.png" alt="After passing the gatehouse" />
  După punctul de control (A), mergeți drept până la parcarea de odihnă (C).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/03_enter_driver_rest_parking.png" alt="Enter the driver rest parking area" />
  Intrați în parcarea de odihnă prin bariera (L).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/04_go_through_barrier_gate.png" alt="Go through the barrier gate" />
  Treceți prin bariera (M) pentru a ieși din parcare și a accesa curtea.
</li>
</ul>

<h3>PE JOS</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/05_stairs_leading_from.png" alt="Stairs leading from" />
  <img class="hr-img" src="/static/house_rules/06_walk_along_fence.png" alt="Walk along the fence" />
  <img class="hr-img" src="/static/house_rules/07_pass_through_gate.png" alt="Pass through the gate" />
  <b>Din lounge la birou</b>: urcați scările (I) din lounge (A) pe pasarelă, mergeți pe lângă gard și urmați indicatoarele, apoi treceți prin poarta (J) către biroul de briefing (B).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/08_walk_along_path.png" alt="Walk along the path" />
  <img class="hr-img" src="/static/house_rules/09_continue_straight.png" alt="Continue straight" />
  <b>Din parcarea relay la birou</b>: urmați traseul cu marcaje verzi lângă gard, continuați drept, apoi treceți prin poarta (J) către biroul de briefing (B).
</li>
</ul>

<h3>În caz de urgență</h3>
<ul>
  <li>Păstrați calmul.</li>
  <li>Sunați mai întâi la <b>112</b>.</li>
  <li>Sunați securitatea FedEx: <b>0316 799 476</b>.</li>
  <li>În caz de incendiu sau alarmă: ieșiți pe cea mai apropiată ieșire de urgență și mergeți la punctul de adunare.</li>
</ul>

<h3>Reguli de siguranță pe site</h3>
<ul>
  <li>Limită de viteză: <b>15 km/h</b>.</li>
  <li>Trafic într-un singur sens în curte.</li>
  <li>Mergeți doar pe căile pietonale marcate.</li>
  <li>Purtați vestă reflectorizantă închisă și încălțăminte de protecție (clasa S3) în curte.</li>
  <li>Folosiți echipamentul de siguranță adecvat la cuplare sau manipularea remorcilor.</li>
  <li>Respectați regulile locale și instrucțiunile personalului de securitate și shunterilor.</li>
</ul>
<div class="muted">Pentru reguli complete, consultați pliantul „FedEx Duiven–QAR site yard rules” (disponibil la poartă și în birou).</div>`
      },
      ru: {
        title: "QAR Duiven | Маршруты движения и пешие маршруты",
        intro: "Пожалуйста, ознакомьтесь со схемой площадки и правилами безопасности. Чтобы продолжить, нужно принять правила площадки.",
        accept: "Я принимаю правила площадки",
        cont: "Продолжить",
        open_pdf: "Открыть PDF в новой вкладке",
        html: `<img class="hr-img" src="/static/house_rules/01_site_map.png" alt="Site map" />
<h3>Схема площадки – ключевые места</h3>
<ul>
  <li><b>Офисы/зоны для водителей</b>: A – КПП / лаунж для водителей; B – Офис брифинга и документов</li>
  <li><b>Парковки</b>: C – Парковка отдыха водителей; D – Парковка прицепов; E – Relay-парковка → тягач с прицепом; F – Relay-парковка → тягач без прицепа; G – Парковка для посетителей</li>
  <li><b>Пешеходы</b>: H – Вход в лаунж; I – Лестница на пешеходный проход к офису; J – Пешеходные ворота к офису</li>
  <li><b>Шлагбаумы</b>: K – Въезд грузовиков; L – Выезд с двора на парковку водителей; M – Въезд во двор из зоны водителей; N – Шлагбаум парковки посетителей; O – Выезд грузовиков</li>
</ul>

<h3>ДВИЖЕНИЕ – въезд/выезд на парковку отдыха водителей</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/02_after_gatehouse.png" alt="After passing the gatehouse" />
  Проехав КПП (A), двигайтесь прямо к парковке отдыха (C).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/03_enter_driver_rest_parking.png" alt="Enter the driver rest parking area" />
  Въезд на парковку отдыха через шлагбаум (L).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/04_go_through_barrier_gate.png" alt="Go through the barrier gate" />
  Для выезда и доступа во двор проезжайте через шлагбаум (M).
</li>
</ul>

<h3>ПЕШКОМ</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/05_stairs_leading_from.png" alt="Stairs leading from" />
  <img class="hr-img" src="/static/house_rules/06_walk_along_fence.png" alt="Walk along the fence" />
  <img class="hr-img" src="/static/house_rules/07_pass_through_gate.png" alt="Pass through the gate" />
  <b>Из лаунжа в офис</b>: поднимитесь по лестнице (I) из лаунжа (A) на проход, идите вдоль забора и следуйте указателям, затем пройдите через ворота (J) к офису брифинга (B).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/08_walk_along_path.png" alt="Walk along the path" />
  <img class="hr-img" src="/static/house_rules/09_continue_straight.png" alt="Continue straight" />
  <b>С relay-парковки в офис</b>: идите по дорожке с зелёной разметкой вдоль забора, продолжайте прямо и пройдите через ворота (J) к офису брифинга (B).
</li>
</ul>

<h3>В экстренной ситуации</h3>
<ul>
  <li>Сохраняйте спокойствие.</li>
  <li>Сначала звоните <b>112</b>.</li>
  <li>Безопасность FedEx: <b>0316 799 476</b>.</li>
  <li>При пожаре/сработавшей сигнализации: выйдите через ближайший аварийный выход и направляйтесь к месту сбора.</li>
</ul>

<h3>Правила безопасности на площадке</h3>
<ul>
  <li>Ограничение скорости: <b>15 км/ч</b>.</li>
  <li>Одностороннее движение во дворе.</li>
  <li>Ходите только по обозначенным пешеходным маршрутам.</li>
  <li>Носите закрытый светоотражающий жилет и защитную обувь (класс S3) во дворе.</li>
  <li>Используйте соответствующие средства защиты при стыковке/работе с прицепами.</li>
  <li>Всегда соблюдайте местные правила и указания охраны и шунтеров.</li>
</ul>
<div class="muted">Полные правила: буклет FedEx Duiven–QAR (доступен на КПП и в офисе).</div>`
      },
      lt: {
        title: "QAR Duiven | Važiavimo ir ėjimo maršrutai",
        intro: "Perskaitykite aikštelės žemėlapį ir saugos taisykles. Norėdami tęsti, turite sutikti su taisyklėmis.",
        accept: "Sutinku su aikštelės taisyklėmis",
        cont: "Tęsti",
        open_pdf: "Atidaryti PDF naujame skirtuke",
        html: `<img class="hr-img" src="/static/house_rules/01_site_map.png" alt="Site map" />
<h3>Aikštelės žemėlapis – pagrindinės vietos</h3>
<ul>
  <li><b>Biurai / vairuotojų patalpos</b>: A – Sargybinė / vairuotojų poilsio zona; B – Vairuotojų instruktažas &amp; dokumentų biuras</li>
  <li><b>Parkavimo zonos</b>: C – Vairuotojų poilsio parkavimas; D – Priekabų parkavimas; E – Relay parkavimas → vilkikas su priekaba; F – Relay parkavimas → vilkikas be priekabos; G – Lankytojų parkavimas</li>
  <li><b>Pėstieji</b>: H – Įėjimas į vairuotojų zoną; I – Laiptai į taką link biuro; J – Pėsčiųjų varteliai į biurą</li>
  <li><b>Užtvarai</b>: K – Sunkvežimių įvažiavimo užtvaras; L – Išvažiavimas iš kiemo į vairuotojų parkavimą; M – Įvažiavimas į kiemą iš vairuotojų zonos; N – Lankytojų parkavimo užtvaras; O – Sunkvežimių išvažiavimo užtvaras</li>
</ul>

<h3>VAŽIAVIMAS – įvažiavimas/išvažiavimas iš vairuotojų poilsio parkavimo</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/02_after_gatehouse.png" alt="After passing the gatehouse" />
  Praėjus sargybinei (A), važiuokite tiesiai į vairuotojų poilsio parkavimą (C).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/03_enter_driver_rest_parking.png" alt="Enter the driver rest parking area" />
  Įvažiuokite per užtvarą (L).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/04_go_through_barrier_gate.png" alt="Go through the barrier gate" />
  Norėdami išvažiuoti ir patekti į kiemą, pravažiuokite per užtvarą (M).
</li>
</ul>

<h3>EJIMAS PĖSČIOMIS</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/05_stairs_leading_from.png" alt="Stairs leading from" />
  <img class="hr-img" src="/static/house_rules/06_walk_along_fence.png" alt="Walk along the fence" />
  <img class="hr-img" src="/static/house_rules/07_pass_through_gate.png" alt="Pass through the gate" />
  <b>Iš poilsio zonos į biurą</b>: lipkite laiptais (I) iš zonos (A) į taką, eikite palei tvorą ir sekite ženklus, tada praeikite pro vartelius (J) iki briefing biuro (B).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/08_walk_along_path.png" alt="Walk along the path" />
  <img class="hr-img" src="/static/house_rules/09_continue_straight.png" alt="Continue straight" />
  <b>Iš relay parkavimo į biurą</b>: eikite taku su žaliomis žymomis palei tvorą, eikite tiesiai ir praeikite pro vartelius (J) iki briefing biuro (B).
</li>
</ul>

<h3>Avarijos atveju</h3>
<ul>
  <li>Išlikite ramūs.</li>
  <li>Pirmiausia skambinkite <b>112</b>.</li>
  <li>Skambinkite FedEx apsaugai: <b>0316 799 476</b>.</li>
  <li>Gaisro ar signalo atveju: išeikite per artimiausią avarinį išėjimą ir eikite į susirinkimo vietą.</li>
</ul>

<h3>Saugos taisyklės aikštelėje</h3>
<ul>
  <li>Greičio limitas: <b>15 km/h</b>.</li>
  <li>Kieme – vienos krypties eismas.</li>
  <li>Eikite tik pažymėtais pėsčiųjų takais.</li>
  <li>Kieme dėvėkite užsegtą ryškiaspalvę liemenę ir apsauginius batus (S3).</li>
  <li>Prijungiant ar tvarkant priekabas naudokite tinkamas apsaugos priemones.</li>
  <li>Visada laikykitės vietinių taisyklių ir apsaugos bei shunterių nurodymų.</li>
</ul>
<div class="muted">Pilnos taisyklės: bukletas „FedEx Duiven–QAR site yard rules“ (gaunamas sargybinėje ir biure).</div>`
      },
      kk: {
        title: "QAR Duiven | Көлікпен және жаяу жүру бағыттары",
        intro: "Алаң картасын және қауіпсіздік ережелерін оқып шығыңыз. Жалғастыру үшін алаң ережелерін қабылдауыңыз керек.",
        accept: "Мен алаң ережелерін қабылдаймын",
        cont: "Жалғастыру",
        open_pdf: "PDF-ті жаңа қойындыда ашу",
        html: `<img class="hr-img" src="/static/house_rules/01_site_map.png" alt="Site map" />
<h3>Алаң картасы – негізгі орындар</h3>
<ul>
  <li><b>Кеңсе/жүргізуші аймақтары</b>: A – Күзет бекеті / жүргізуші лаунжы; B – Брифинг және құжаттар кеңсесі</li>
  <li><b>Тұрақ аймақтары</b>: C – Жүргізуші демалыс тұрағы; D – Тіркеме тұрағы; E – Relay тұрағы → тіркемесі бар тартқыш; F – Relay тұрағы → жеке тартқыш; G – Келушілер тұрағы</li>
  <li><b>Жаяу жүргінші</b>: H – Лаунжға кіреберіс; I – Кеңсеге баратын жаяу жолға баспалдақ; J – Кеңсеге жаяу қақпа</li>
  <li><b>Шлагбаумдар</b>: K – Жүк көлігі кіреберіс шлагбаумы; L – Аулада шығу → жүргізуші тұрағы; M – Жүргізуші аймағынан аулаға кіру; N – Келушілер тұрағы шлагбаумы; O – Жүк көлігі шығу шлагбаумы</li>
</ul>

<h3>КӨЛІКПЕН – демалыс тұрағына кіру/шығу</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/02_after_gatehouse.png" alt="After passing the gatehouse" />
  Күзет бекетінен (A) өткен соң, демалыс тұрағына (C) дейін түзу жүріңіз.
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/03_enter_driver_rest_parking.png" alt="Enter the driver rest parking area" />
  Шлагбаум (L) арқылы демалыс тұрағына кіріңіз.
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/04_go_through_barrier_gate.png" alt="Go through the barrier gate" />
  Аулаға шығу үшін шлагбаумнан (M) өтіңіз.
</li>
</ul>

<h3>ЖАЯУ</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/05_stairs_leading_from.png" alt="Stairs leading from" />
  <img class="hr-img" src="/static/house_rules/06_walk_along_fence.png" alt="Walk along the fence" />
  <img class="hr-img" src="/static/house_rules/07_pass_through_gate.png" alt="Pass through the gate" />
  <b>Лаунждан кеңсеге</b>: лаунждан (A) баспалдақпен (I) жаяу жолға шығып, қоршау бойымен жүріп, белгілерді қадағалаңыз, содан кейін қақпадан (J) өтіп брифинг кеңсесіне (B) барыңыз.
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/08_walk_along_path.png" alt="Walk along the path" />
  <img class="hr-img" src="/static/house_rules/09_continue_straight.png" alt="Continue straight" />
  <b>Relay тұрағынан кеңсеге</b>: қоршау жанындағы жасыл белгіленген жолмен жүріңіз, түзу жалғастырыңыз, содан кейін қақпадан (J) өтіп брифинг кеңсесіне (B) барыңыз.
</li>
</ul>

<h3>Төтенше жағдайда</h3>
<ul>
  <li>Сабыр сақтаңыз.</li>
  <li>Алдымен <b>112</b> нөміріне қоңырау шалыңыз.</li>
  <li>FedEx қауіпсіздігі: <b>0316 799 476</b>.</li>
  <li>Өрт немесе өрт дабылы кезінде: ең жақын авариялық шығу арқылы шығып, жиналу нүктесіне барыңыз.</li>
</ul>

<h3>Алаңдағы қауіпсіздік ережелері</h3>
<ul>
  <li>Жылдамдық шегі: <b>15 км/сағ</b>.</li>
  <li>Аулада бір бағытты қозғалыс.</li>
  <li>Тек белгіленген жаяу жолдармен жүріңіз.</li>
  <li>Аулада жабық жарыққайтарғыш кеудеше және қауіпсіздік аяқ киімі (S3) киіңіз.</li>
  <li>Тіркемені докқа қою/өңдеу кезінде тиісті қорғаныс құралдарын қолданыңыз.</li>
  <li>Жергілікті ережелерді және күзет пен shunter нұсқауларын әрдайым орындаңыз.</li>
</ul>
<div class="muted">Толық ережелер: FedEx Duiven–QAR алаң ережелері буклеті (күзет бекетінде және кеңседе бар).</div>`
      },
      hi: {
        title: "QAR Duiven | ड्राइविंग और पैदल मार्ग",
        intro: "कृपया साइट मैप और सुरक्षा नियम पढ़ें। आगे बढ़ने के लिए आपको साइट के नियम स्वीकार करने होंगे।",
        accept: "मैं साइट के नियम स्वीकार करता/करती हूँ",
        cont: "जारी रखें",
        open_pdf: "PDF नई टैब में खोलें",
        html: `<img class="hr-img" src="/static/house_rules/01_site_map.png" alt="Site map" />
<h3>साइट मैप – प्रमुख स्थान</h3>
<ul>
  <li><b>ऑफिस/ड्राइवर सुविधाएँ</b>: A – गेटहाउस / ड्राइवर लाउंज; B – ड्राइवर ब्रीफिंग और दस्तावेज़ कार्यालय</li>
  <li><b>पार्किंग क्षेत्र</b>: C – ड्राइवर रेस्ट पार्किंग; D – ट्रेलर पार्किंग; E – रिले पार्किंग → ट्रक + ट्रेलर; F – रिले पार्किंग → केवल ट्रक; G – विज़िटर पार्किंग</li>
  <li><b>पैदल यात्री</b>: H – ड्राइवर लाउंज प्रवेश; I – ऑफिस वॉकवे के लिए सीढ़ियाँ; J – ऑफिस के लिए पैदल गेट</li>
  <li><b>बैARRIER गेट</b>: K – ट्रक एंट्रेंस बैARRIER; L – यार्ड से ड्राइवर पार्किंग की ओर निकास; M – ड्राइवर क्षेत्र से यार्ड में प्रवेश; N – विज़िटर पार्किंग बैARRIER; O – ट्रक एग्ज़िट बैARRIER</li>
</ul>

<h3>ड्राइविंग – ड्राइवर रेस्ट पार्किंग में प्रवेश/निकास</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/02_after_gatehouse.png" alt="After passing the gatehouse" />
  गेटहाउस (A) के बाद सीधे ड्राइवर रेस्ट पार्किंग (C) तक जाएँ।
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/03_enter_driver_rest_parking.png" alt="Enter the driver rest parking area" />
  बैARRIER गेट (L) से होकर पार्किंग में प्रवेश करें।
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/04_go_through_barrier_gate.png" alt="Go through the barrier gate" />
  यार्ड तक पहुँचने के लिए बैARRIER गेट (M) से होकर बाहर निकलें।
</li>
</ul>

<h3>पैदल</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/05_stairs_leading_from.png" alt="Stairs leading from" />
  <img class="hr-img" src="/static/house_rules/06_walk_along_fence.png" alt="Walk along the fence" />
  <img class="hr-img" src="/static/house_rules/07_pass_through_gate.png" alt="Pass through the gate" />
  <b>ड्राइवर लाउंज से ऑफिस</b>: लाउंज (A) से सीढ़ियाँ (I) लेकर वॉकवे पर जाएँ, बाड़ के साथ चलते हुए संकेतों का पालन करें, फिर गेट (J) से होकर ब्रीफिंग ऑफिस (B) पहुँचें।
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/08_walk_along_path.png" alt="Walk along the path" />
  <img class="hr-img" src="/static/house_rules/09_continue_straight.png" alt="Continue straight" />
  <b>रिले पार्किंग से ऑफिस</b>: बाड़ के पास हरे निशान वाले रास्ते पर चलें, सीधे आगे बढ़ें, फिर गेट (J) से होकर ब्रीफिंग ऑफिस (B) पहुँचें।
</li>
</ul>

<h3>आपात स्थिति में</h3>
<ul>
  <li>शांत रहें।</li>
  <li>सबसे पहले <b>112</b> पर कॉल करें।</li>
  <li>FedEx सुरक्षा: <b>0316 799 476</b>।</li>
  <li>आग/फायर अलार्म में: नज़दीकी आपात निकास से बाहर जाएँ और असेंबली पॉइंट पर जाएँ।</li>
</ul>

<h3>साइट सुरक्षा नियम</h3>
<ul>
  <li>गति सीमा: <b>15 किमी/घं</b>।</li>
  <li>यार्ड में एक-तरफ़ा ट्रैफ़िक।</li>
  <li>केवल निर्धारित पैदल मार्गों पर चलें।</li>
  <li>यार्ड में बंद हाई-विज़िबिलिटी सेफ़्टी वेस्ट और सेफ़्टी शूज़ (क्लास S3) पहनें।</li>
  <li>ट्रेलर डॉकिंग/हैंडलिंग के समय उपयुक्त सुरक्षा उपकरण का उपयोग करें।</li>
  <li>स्थानीय नियमों तथा सुरक्षा कर्मियों और शंटरों के निर्देशों का पालन करें।</li>
</ul>
<div class="muted">पूरे नियम: FedEx Duiven–QAR साइट यार्ड नियम पुस्तिका (गेटहाउस और ऑफिस में उपलब्ध)।</div>`
      },
      pl: {
        title: "QAR Duiven | Trasy dojazdu i dojścia",
        intro: "Zapoznaj się z mapą terenu i zasadami bezpieczeństwa. Aby kontynuować, musisz zaakceptować regulamin.",
        accept: "Akceptuję regulamin terenu",
        cont: "Dalej",
        open_pdf: "Otwórz PDF w nowej karcie",
        html: `<img class="hr-img" src="/static/house_rules/01_site_map.png" alt="Site map" />
<h3>Mapa terenu – kluczowe miejsca</h3>
<ul>
  <li><b>Biura/strefy kierowców</b>: A – Portiernia / lounge kierowców; B – Biuro briefingu i dokumentów</li>
  <li><b>Parkingi</b>: C – Parking odpoczynku kierowców; D – Parking naczep; E – Parking relay → ciągnik z naczepą; F – Parking relay → sam ciągnik; G – Parking dla gości</li>
  <li><b>Piesi</b>: H – Wejście do lounge; I – Schody na kładkę do biura; J – Brama piesza do biura</li>
  <li><b>Szlabany</b>: K – Szlaban wjazdowy dla ciężarówek; L – Wyjazd z placu na parking kierowców; M – Wjazd na plac ze strefy kierowców; N – Szlaban parkingu gości; O – Szlaban wyjazdowy dla ciężarówek</li>
</ul>

<h3>JAZDA – wjazd/wyjazd z parkingu odpoczynku</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/02_after_gatehouse.png" alt="After passing the gatehouse" />
  Po minięciu portierni (A) jedź prosto do parkingu odpoczynku (C).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/03_enter_driver_rest_parking.png" alt="Enter the driver rest parking area" />
  Wjedź na parking przez szlaban (L).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/04_go_through_barrier_gate.png" alt="Go through the barrier gate" />
  Aby wyjechać i dostać się na plac, przejedź przez szlaban (M).
</li>
</ul>

<h3>PIESZO</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/05_stairs_leading_from.png" alt="Stairs leading from" />
  <img class="hr-img" src="/static/house_rules/06_walk_along_fence.png" alt="Walk along the fence" />
  <img class="hr-img" src="/static/house_rules/07_pass_through_gate.png" alt="Pass through the gate" />
  <b>Z lounge do biura</b>: wejdź schodami (I) z lounge (A) na kładkę, idź wzdłuż ogrodzenia i kieruj się znakami, następnie przejdź przez bramę (J) do biura briefingu (B).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/08_walk_along_path.png" alt="Walk along the path" />
  <img class="hr-img" src="/static/house_rules/09_continue_straight.png" alt="Continue straight" />
  <b>Z parkingu relay do biura</b>: idź ścieżką z zielonymi oznaczeniami przy ogrodzeniu, idź prosto, następnie przejdź przez bramę (J) do biura briefingu (B).
</li>
</ul>

<h3>W sytuacji awaryjnej</h3>
<ul>
  <li>Zachowaj spokój.</li>
  <li>Najpierw zadzwoń pod <b>112</b>.</li>
  <li>Zadzwoń do ochrony FedEx: <b>0316 799 476</b>.</li>
  <li>W razie pożaru/alarmu: wyjdź najbliższym wyjściem ewakuacyjnym i idź do punktu zbiórki.</li>
</ul>

<h3>Zasady bezpieczeństwa na terenie</h3>
<ul>
  <li>Limit prędkości: <b>15 km/h</b>.</li>
  <li>Ruch jednokierunkowy na placu.</li>
  <li>Poruszaj się tylko wyznaczonymi ciągami pieszymi.</li>
  <li>Na placu noś zapiętą kamizelkę odblaskową i buty ochronne (klasa S3).</li>
  <li>Używaj odpowiednich środków ochrony przy dokowaniu/obsłudze naczep.</li>
  <li>Zawsze stosuj się do lokalnych zasad oraz poleceń ochrony i shunterów.</li>
</ul>
<div class="muted">Pełne zasady: broszura „FedEx Duiven–QAR site yard rules” (dostępna w portierni i biurze).</div>`
      },
      hu: {
        title: "QAR Duiven | Behajtási és gyalogos útvonalak",
        intro: "Kérlek olvasd el a telephely térképét és a biztonsági szabályokat. A folytatáshoz el kell fogadnod a házirendet.",
        accept: "Elfogadom a házirendet",
        cont: "Tovább",
        open_pdf: "PDF megnyitása új lapon",
        html: `<img class="hr-img" src="/static/house_rules/01_site_map.png" alt="Site map" />
<h3>Telephely térkép – kulcs helyszínek</h3>
<ul>
  <li><b>Irodák/sofőr létesítmények</b>: A – Kapuőrház / sofőr pihenő; B – Sofőr briefing &amp; papírmunka iroda</li>
  <li><b>Parkolók</b>: C – Sofőr pihenő parkoló; D – Pótkocsi parkoló; E – Relay parkoló → vontató + pótkocsi; F – Relay parkoló → solo vontató; G – Látogatói parkoló</li>
  <li><b>Gyalogosok</b>: H – Bejárat a sofőr pihenőbe; I – Lépcső az irodához vezető sétányra; J – Gyalogos kapu az irodához</li>
  <li><b>Sorompók</b>: K – Teherautó bejárati sorompó; L – Udvar kijárat a sofőr parkolóhoz; M – Udvar bejárat a sofőr zónából; N – Látogatói parkoló sorompó; O – Teherautó kijárati sorompó</li>
</ul>

<h3>VEZETÉS – sofőr pihenő parkoló be/ki</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/02_after_gatehouse.png" alt="After passing the gatehouse" />
  A kapuőrház (A) után hajts egyenesen a sofőr pihenő parkolóba (C).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/03_enter_driver_rest_parking.png" alt="Enter the driver rest parking area" />
  A sorompón (L) keresztül hajts be a pihenő parkolóba.
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/04_go_through_barrier_gate.png" alt="Go through the barrier gate" />
  Az udvar eléréséhez hajts át a sorompón (M) és hagyd el a pihenő parkolót.
</li>
</ul>

<h3>GYALOG</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/05_stairs_leading_from.png" alt="Stairs leading from" />
  <img class="hr-img" src="/static/house_rules/06_walk_along_fence.png" alt="Walk along the fence" />
  <img class="hr-img" src="/static/house_rules/07_pass_through_gate.png" alt="Pass through the gate" />
  <b>A pihenőből az irodába</b>: menj fel a lépcsőn (I) a pihenőből (A) a sétányra, haladj a kerítés mellett és kövesd a táblákat, majd menj át a kapun (J) a briefing irodához (B).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/08_walk_along_path.png" alt="Walk along the path" />
  <img class="hr-img" src="/static/house_rules/09_continue_straight.png" alt="Continue straight" />
  <b>A relay parkolóból az irodába</b>: kövesd a zöld jelölésű utat a kerítés mellett, menj egyenesen, majd menj át a kapun (J) a briefing irodához (B).
</li>
</ul>

<h3>Vészhelyzet esetén</h3>
<ul>
  <li>Maradj nyugodt.</li>
  <li>Először hívd a <b>112</b>-t.</li>
  <li>Hívd a FedEx security-t: <b>0316 799 476</b>.</li>
  <li>Tűz vagy tűzriadó esetén: a legközelebbi vészkijáraton hagyd el az épületet és menj a gyülekezési pontra.</li>
</ul>

<h3>Telephelyi biztonsági szabályok</h3>
<ul>
  <li>Sebességhatár: <b>15 km/h</b>.</li>
  <li>Az udvaron egyirányú forgalom van.</li>
  <li>Csak kijelölt gyalogos útvonalon közlekedj.</li>
  <li>Az udvaron zárt, láthatósági mellényt és munkavédelmi cipőt (S3) viselj.</li>
  <li>Pótkocsi dokkolásakor/kezelésekor megfelelő védőeszközt használj.</li>
  <li>Mindig kövesd a helyi szabályokat, valamint a security és a shunterek utasításait.</li>
</ul>
<div class="muted">Teljes szabályok: FedEx Duiven–QAR telephelyi szabályok szórólap (a kapuőrháznál és az irodában elérhető).</div>`
      },
      uz: {
        title: "QAR Duiven | Haydash va piyoda yo‘nalishlar",
        intro: "Iltimos, hudud xaritasi va xavfsizlik qoidalarini o‘qing. Davom etish uchun hudud qoidalarini qabul qilishingiz kerak.",
        accept: "Men hudud qoidalarini qabul qilaman",
        cont: "Davom etish",
        open_pdf: "PDF-ni yangi yorliqda ochish",
        html: `<img class="hr-img" src="/static/house_rules/01_site_map.png" alt="Site map" />
<h3>Hudud xaritasi – muhim joylar</h3>
<ul>
  <li><b>Ofislar/haydovchi hududi</b>: A – Qo‘riqlash punkti / haydovchilar dam olish joyi; B – Brifing va hujjatlar ofisi</li>
  <li><b>To‘xtash joylari</b>: C – Haydovchilar dam olish parkingi; D – Tirkama parkingi; E – Relay parking → yuk mashinasi+tirkama; F – Relay parking → solo yuk mashinasi; G – Mehmonlar parkingi</li>
  <li><b>Piyodalar</b>: H – Dam olish joyiga kirish; I – Ofis yo‘lagiga olib boruvchi zinapoya; J – Ofisga piyoda darvoza</li>
  <li><b>To‘siq darvozalar</b>: K – Yuk mashinasi kirish to‘sig‘i; L – Hududdan haydovchi parkingiga chiqish; M – Haydovchi hududidan hududga kirish; N – Mehmonlar parkingi to‘sig‘i; O – Yuk mashinasi chiqish to‘sig‘i</li>
</ul>

<h3>HAYDASH – dam olish parkingiga kirish/chiqish</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/02_after_gatehouse.png" alt="After passing the gatehouse" />
  Qo‘riqlash punktidan (A) o‘tgach, to‘g‘ri yurib dam olish parkingiga (C) boring.
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/03_enter_driver_rest_parking.png" alt="Enter the driver rest parking area" />
  To‘siq darvoza (L) orqali parkingga kiring.
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/04_go_through_barrier_gate.png" alt="Go through the barrier gate" />
  Hududga chiqish uchun to‘siq darvoza (M) orqali o‘ting.
</li>
</ul>

<h3>PIYODA</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/05_stairs_leading_from.png" alt="Stairs leading from" />
  <img class="hr-img" src="/static/house_rules/06_walk_along_fence.png" alt="Walk along the fence" />
  <img class="hr-img" src="/static/house_rules/07_pass_through_gate.png" alt="Pass through the gate" />
  <b>Dam olish joyidan ofisga</b>: dam olish joyidan (A) zinapoya (I) orqali yo‘lakka chiqing, panjara bo‘ylab yuring va belgilarga amal qiling, so‘ng darvozadan (J) o‘tib brifing ofisiga (B) boring.
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/08_walk_along_path.png" alt="Walk along the path" />
  <img class="hr-img" src="/static/house_rules/09_continue_straight.png" alt="Continue straight" />
  <b>Relay parkingdan ofisga</b>: panjara yonidagi yashil belgilangan yo‘ldan yuring, to‘g‘ri davom eting, so‘ng darvozadan (J) o‘tib brifing ofisiga (B) boring.
</li>
</ul>

<h3>Favqulodda holatda</h3>
<ul>
  <li>Xotirjam bo‘ling.</li>
  <li>Avval <b>112</b> ga qo‘ng‘iroq qiling.</li>
  <li>FedEx xavfsizligi: <b>0316 799 476</b>.</li>
  <li>Yong‘in/yong‘in signali bo‘lsa: eng yaqin favqulodda chiqishdan chiqing va yig‘ilish nuqtasiga boring.</li>
</ul>

<h3>Hudud xavfsizlik qoidalari</h3>
<ul>
  <li>Tezlik cheklovi: <b>15 km/soat</b>.</li>
  <li>Hududda bir yo‘nalishli harakat.</li>
  <li>Faqat belgilangan piyoda yo‘laklaridan yuring.</li>
  <li>Hududda yopiq hi-vis jilet va xavfsizlik poyabzali (S3) kiying.</li>
  <li>Tirkamalarni doklash/ishlashda mos xavfsizlik jihozlaridan foydalaning.</li>
  <li>Har doim mahalliy qoidalar va xavfsizlik xodimlari hamda shunter ko‘rsatmalariga amal qiling.</li>
</ul>
<div class="muted">To‘liq qoidalar: FedEx Duiven–QAR hudud qoidalari bukleti (qo‘riqlash punkti va ofisda mavjud).</div>`
      },
      tg: {
        title: "QAR Duiven | Роҳҳои рондан ва пиёдагардӣ",
        intro: "Лутфан нақшаи маҳал ва қоидаҳои бехатариро хонед. Барои идома қоидаҳои маҳалро қабул кардан лозим аст.",
        accept: "Ман қоидаҳои маҳалро қабул мекунам",
        cont: "Идома",
        open_pdf: "PDF-ро дар ҷадвали нав кушоед",
        html: `<img class="hr-img" src="/static/house_rules/01_site_map.png" alt="Site map" />
<h3>Нақшаи маҳал – ҷойҳои муҳим</h3>
<ul>
  <li><b>Идора/шароити ронандагон</b>: A – Посбонхона / лонҷи ронанда; B – Офиси брифинг ва ҳуҷҷатҳо</li>
  <li><b>Майдони таваққуф</b>: C – Таваққуфи истироҳати ронанда; D – Таваққуфи прицепҳо; E – Relay таваққуф → мошин бо прицеп; F – Relay таваққуф → мошини танҳо; G – Таваққуфи меҳмонон</li>
  <li><b>Пиёдагард</b>: H – Даромад ба лонҷ; I – Зинапоя ба роҳрави идора; J – Дарвозаи пиёдагард ба идора</li>
  <li><b>Монеаҳо</b>: K – Монеаи даромади мошинҳо; L – Баромад аз ҳавлӣ ба таваққуфи ронанда; M – Даромад ба ҳавлӣ аз минтақаи ронанда; N – Монеаи таваққуфи меҳмонон; O – Монеаи баромади мошинҳо</li>
</ul>

<h3>РОНДАН – даромад/баромад аз таваққуфи истироҳат</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/02_after_gatehouse.png" alt="After passing the gatehouse" />
  Пас аз посбонхона (A) рост ба таваққуфи истироҳат (C) равед.
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/03_enter_driver_rest_parking.png" alt="Enter the driver rest parking area" />
  Аз монеа (L) гузашта ба таваққуф дароед.
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/04_go_through_barrier_gate.png" alt="Go through the barrier gate" />
  Барои баромадан ва ба ҳавлӣ расидан аз монеа (M) гузаред.
</li>
</ul>

<h3>ПИЁДА</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/05_stairs_leading_from.png" alt="Stairs leading from" />
  <img class="hr-img" src="/static/house_rules/06_walk_along_fence.png" alt="Walk along the fence" />
  <img class="hr-img" src="/static/house_rules/07_pass_through_gate.png" alt="Pass through the gate" />
  <b>Аз лонҷ ба идора</b>: аз лонҷ (A) бо зинапоя (I) ба роҳрав бароед, дар канори девор (панҷара) қадам занед ва нишонаҳоро пайравӣ кунед, сипас аз дарвоза (J) гузашта ба офиси брифинг (B) расед.
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/08_walk_along_path.png" alt="Walk along the path" />
  <img class="hr-img" src="/static/house_rules/09_continue_straight.png" alt="Continue straight" />
  <b>Аз relay таваққуф ба идора</b>: роҳро бо аломатҳои сабз дар канори панҷара пайравӣ кунед, рост идома диҳед ва аз дарвоза (J) гузашта ба офиси брифинг (B) расед.
</li>
</ul>

<h3>Дар ҳолати фавқулода</h3>
<ul>
  <li>Ором бошед.</li>
  <li>Аввал ба <b>112</b> занг занед.</li>
  <li>Амнияти FedEx: <b>0316 799 476</b>.</li>
  <li>Ҳангоми сӯхтор ё ҳушдор: аз наздиктарин баромади фавқулода бароед ва ба нуқтаи ҷамъшавӣ равед.</li>
</ul>

<h3>Қоидаҳои бехатарӣ дар маҳал</h3>
<ul>
  <li>Ҳадди суръат: <b>15 км/соат</b>.</li>
  <li>Ҳаракати яктарафа дар ҳавлӣ.</li>
  <li>Танҳо дар роҳҳои махсуси пиёдагард ҳаракат кунед.</li>
  <li>Дар ҳавлӣ жилети намоён (hi-vis) ва пойафзоли бехатарӣ (S3) пӯшед.</li>
  <li>Ҳангоми док кардан ё кор бо прицепҳо таҷҳизоти мувофиқи бехатариро истифода баред.</li>
  <li>Ҳамеша қоидаҳои маҳаллӣ ва дастурҳои амният ва shunter-ҳоро риоя кунед.</li>
</ul>
<div class="muted">Қоидаҳои пурра: буклети FedEx Duiven–QAR (дар посбонхона ва идора дастрас аст).</div>`
      },
      ky: {
        title: "QAR Duiven | Айдоо жана жөө жүрүү маршруттары",
        intro: "Сураныч, аянттын картасын жана коопсуздук эрежелерин окуңуз. Улантуу үчүн аянт эрежелерин кабыл алуу керек.",
        accept: "Мен аянттын эрежелерин кабыл алам",
        cont: "Улантуу",
        open_pdf: "PDFти жаңы өтмөктө ачуу",
        html: `<img class="hr-img" src="/static/house_rules/01_site_map.png" alt="Site map" />
<h3>Аянт картасы – негизги жерлер</h3>
<ul>
  <li><b>Офис/айдоочулар үчүн жайлар</b>: A – Күзөт пункту / айдоочу лаунжу; B – Брифинг жана документтер офиси</li>
  <li><b>Токтотмо жайлар</b>: C – Айдоочу эс алуу паркинги; D – Прицеп паркинги; E – Relay паркинги → прицептүү тягач; F – Relay паркинги → жалгыз тягач; G – Коноктор паркинги</li>
  <li><b>Жөө жүрүүчү</b>: H – Лаунжга кирүү; I – Офиске баруучу жолго тепкич; J – Офиске жөө дарбаза</li>
  <li><b>Шлагбаумдар</b>: K – Жүк авто кирүү шлагбаумы; L – Аянттан айдоочу паркингине чыгуу; M – Айдоочу зонасынан аянтка кирүү; N – Коноктор паркинги шлагбаумы; O – Жүк авто чыгуу шлагбаумы</li>
</ul>

<h3>АЙДОО – эс алуу паркингине кирүү/чыгуу</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/02_after_gatehouse.png" alt="After passing the gatehouse" />
  Күзөт пунктунан (A) өткөндөн кийин түз эле эс алуу паркингине (C) айдаңыз.
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/03_enter_driver_rest_parking.png" alt="Enter the driver rest parking area" />
  Шлагбаум (L) аркылуу паркингге кириңиз.
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/04_go_through_barrier_gate.png" alt="Go through the barrier gate" />
  Аянтка чыгуу үчүн шлагбаумдан (M) өтүңүз.
</li>
</ul>

<h3>ЖӨӨ</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/05_stairs_leading_from.png" alt="Stairs leading from" />
  <img class="hr-img" src="/static/house_rules/06_walk_along_fence.png" alt="Walk along the fence" />
  <img class="hr-img" src="/static/house_rules/07_pass_through_gate.png" alt="Pass through the gate" />
  <b>Лаунждан офиске</b>: лаунждан (A) тепкич (I) менен жөө жолго чыгып, тосмонун жанынан жүрүп белгилерди ээрчиңиз, андан кийин дарбазадан (J) өтүп брифинг офисине (B) барыңыз.
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/08_walk_along_path.png" alt="Walk along the path" />
  <img class="hr-img" src="/static/house_rules/09_continue_straight.png" alt="Continue straight" />
  <b>Relay паркингинен офиске</b>: тосмонун жанындагы жашыл белгилүү жол менен жүрүңүз, түз улантыңыз, анан дарбазадан (J) өтүп брифинг офисине (B) барыңыз.
</li>
</ul>

<h3>Өзгөчө кырдаалда</h3>
<ul>
  <li>Тынч болуңуз.</li>
  <li>Адегенде <b>112</b> номерине чалыңыз.</li>
  <li>FedEx коопсуздугу: <b>0316 799 476</b>.</li>
  <li>Өрт/өрт сигналы болсо: эң жакын авариялык чыгуу аркылуу чыгып, жыйналуу пунктуна барыңыз.</li>
</ul>

<h3>Аянттагы коопсуздук эрежелери</h3>
<ul>
  <li>Ылдамдык чеги: <b>15 км/саат</b>.</li>
  <li>Аянтта бир багыттуу кыймыл.</li>
  <li>Белгиленген жөө жолдор менен гана жүрүңүз.</li>
  <li>Аянтта жабык hi-vis жилет жана коопсуздук бут кийими (S3) кийиңиз.</li>
  <li>Прицептерди докко коюуда/кармоодо ылайыктуу коргонуу каражаттарын колдонуңуз.</li>
  <li>Жергиликтүү эрежелерди жана коопсуздук кызматкерлери менен shunter көрсөтмөлөрүн дайыма аткарыңыз.</li>
</ul>
<div class="muted">Толук эрежелер: FedEx Duiven–QAR буклети (күзөт пунктунда жана офисте бар).</div>`
      },
      be: {
        title: "QAR Duiven | Маршруты руху і пешыя маршруты",
        intro: "Калі ласка, азнаёмцеся з картай пляцоўкі і правіламі бяспекі. Каб працягнуць, трэба прыняць правілы пляцоўкі.",
        accept: "Я прымаю правілы пляцоўкі",
        cont: "Працягнуць",
        open_pdf: "Адкрыць PDF у новай укладцы",
        html: `<img class="hr-img" src="/static/house_rules/01_site_map.png" alt="Site map" />
<h3>Карта пляцоўкі – ключавыя месцы</h3>
<ul>
  <li><b>Офіс/зоны кіроўцаў</b>: A – Прапускны пункт / лаунж кіроўцаў; B – Офіс брыфінгу і дакументаў</li>
  <li><b>Паркоўкі</b>: C – Паркоўка адпачынку кіроўцаў; D – Паркоўка прычэпаў; E – Relay-паркоўка → цягач з прычэпам; F – Relay-паркоўка → цягач без прычэпа; G – Паркоўка для наведвальнікаў</li>
  <li><b>Пешаходы</b>: H – Уваход у лаунж; I – Лесвіца на пешаходны праход да офіса; J – Пешая брама да офіса</li>
  <li><b>Шлагбаумы</b>: K – Уваход грузавікоў; L – Выезд з двара на паркоўку кіроўцаў; M – Уезд у двор з зоны кіроўцаў; N – Шлагбаум паркоўкі наведвальнікаў; O – Выезд грузавікоў</li>
</ul>

<h3>РУХ – уезд/выезд на паркоўку адпачынку</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/02_after_gatehouse.png" alt="After passing the gatehouse" />
  Пасля КПП (A) рухайцеся прама да паркоўкі адпачынку (C).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/03_enter_driver_rest_parking.png" alt="Enter the driver rest parking area" />
  Уезд на паркоўку праз шлагбаум (L).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/04_go_through_barrier_gate.png" alt="Go through the barrier gate" />
  Каб выехаць і трапіць у двор, праязджайце праз шлагбаум (M).
</li>
</ul>

<h3>ПЕШКІ</h3>
<ul>
  <li>
  <img class="hr-img" src="/static/house_rules/05_stairs_leading_from.png" alt="Stairs leading from" />
  <img class="hr-img" src="/static/house_rules/06_walk_along_fence.png" alt="Walk along the fence" />
  <img class="hr-img" src="/static/house_rules/07_pass_through_gate.png" alt="Pass through the gate" />
  <b>З лаунжа ў офіс</b>: падыміцеся па лесвіцы (I) з лаунжа (A) на праход, ідзіце ўздоўж плота і па ўказальніках, затым прайдзіце праз браму (J) да офіса брыфінгу (B).
</li>
  <li>
  <img class="hr-img" src="/static/house_rules/08_walk_along_path.png" alt="Walk along the path" />
  <img class="hr-img" src="/static/house_rules/09_continue_straight.png" alt="Continue straight" />
  <b>З relay-паркоўкі ў офіс</b>: ідзіце па дарожцы з зялёнай разметкай уздоўж плота, працягвайце прама і прайдзіце праз браму (J) да офіса брыфінгу (B).
</li>
</ul>

<h3>У выпадку надзвычайнай сітуацыі</h3>
<ul>
  <li>Захоўвайце спакой.</li>
  <li>Спачатку тэлефануйце <b>112</b>.</li>
  <li>Бяспека FedEx: <b>0316 799 476</b>.</li>
  <li>Пры пажары/сігнале: выйдзіце праз бліжэйшы аварыйны выхад і ідзіце да пункта збору.</li>
</ul>

<h3>Правілы бяспекі на пляцоўцы</h3>
<ul>
  <li>Абмежаванне хуткасці: <b>15 км/г</b>.</li>
  <li>Аднабаковы рух у двары.</li>
  <li>Хадзіце толькі па пазначаных пешаходных маршрутах.</li>
  <li>Носіце зашпіленую светоадбівальную камізэльку і абутак бяспекі (клас S3) у двары.</li>
  <li>Выкарыстоўвайце адпаведныя сродкі аховы пры докаванні/працы з прычэпамі.</li>
  <li>Заўсёды выконвайце мясцовыя правілы і інструкцыі аховы і shunter’аў.</li>
</ul>
<div class="muted">Поўныя правілы: буклет FedEx Duiven–QAR (даступны на КПП і ў офісе).</div>`
      }
    };


    function _houseRulesPack() {
      return HOUSE_RULES[CURRENT_LANG] || HOUSE_RULES.en;
    }

    function showHouseRulesModal(plate) {
      const pack = _houseRulesPack();
      const backdrop = document.getElementById("hrBackdrop");
      const titleEl = document.getElementById("hrTitle");
      const introEl = document.getElementById("hrIntro");
      const textEl = document.getElementById("hrText");
const acceptEl = document.getElementById("hrAccept");
      const acceptLabelEl = document.getElementById("hrAcceptLabel");
      const contBtn = document.getElementById("hrContinue");

      titleEl.textContent = pack.title || "House rules";
      introEl.textContent = pack.intro || "";
      acceptLabelEl.textContent = pack.accept || "I accept";
      contBtn.textContent = pack.cont || "Continue";
textEl.innerHTML = pack.html || "";

      acceptEl.checked = false;
      contBtn.disabled = true;

      const onToggle = () => {
        contBtn.disabled = !acceptEl.checked;
      };

      return new Promise((resolve) => {
        const onContinue = async () => {
          if (!acceptEl.checked) return;

          contBtn.disabled = true;
          const contText = pack.cont || "Continue";
          contBtn.textContent = contText + "…";

          try {
            const res = await fetch(`${API_BASE}/api/house_rules_accept`, {
              method: "POST",
              headers: { "Content-Type": "application/json" },
              body: JSON.stringify({ plate })
            });
            const data = await readJsonOrText(res);

            if (!res.ok) {
              introEl.innerHTML = `<b>${t("err_error")}:</b> ${data.detail || res.statusText}`;
              contBtn.textContent = contText;
              contBtn.disabled = false;
              return;
            }

            backdrop.style.display = "none";
            cleanup();
            contBtn.textContent = contText;
            resolve(true);
          } catch (e) {
            introEl.innerHTML = `<b>${t("err_network")}:</b> ${e.message}`;
            contBtn.textContent = pack.cont || "Continue";
            contBtn.disabled = false;
          }
        };

        const cleanup = () => {
          acceptEl.removeEventListener("change", onToggle);
          contBtn.removeEventListener("click", onContinue);
        };

        acceptEl.addEventListener("change", onToggle);
        contBtn.addEventListener("click", onContinue);

        // block closing without acceptance (no outside click handler)
        backdrop.style.display = "flex";
      });
    }


    function t(key) {
      const pack = UI[CURRENT_LANG] || UI.en;
      return (pack && pack[key]) || (UI.en && UI.en[key]) || key;
    }

    function setNotifyMsg(html, kind) {
      const el = document.getElementById("notifyMsg");
      if (!el) return;
      el.style.display = html ? "block" : "none";
      el.innerHTML = html || "";
      el.style.color = (kind === "err") ? "#a00000" : "";
    }

    function normalizePlate(v) {
      return (v || "").toUpperCase().trim().replaceAll(" ", "").replaceAll("-", "");
    }

    function getInitialLang() {
      try {
        const l = new URLSearchParams(window.location.search).get("lang") || "";
        const ln = tryLang(l);
        if (ln) return ln;
      } catch (e) {}
      try {
        const ls = localStorage.getItem("lang") || "";
        const ln2 = tryLang(ls);
        if (ln2) return ln2;
      } catch (e) {}
      try {
        const list = (navigator.languages && navigator.languages.length) ? navigator.languages : [];
        for (const cand of list) {
          const ln3 = tryLang(cand);
          if (ln3) return ln3;
        }
      } catch (e) {}
      try {
        const nav = (navigator.language || navigator.userLanguage || "") || "";
        const ln4 = tryLang(nav);
        if (ln4) return ln4;
      } catch (e) {}
      return "en";
    }

    function setCurrentLang(lang, opts) {
      const ln = normLang(lang);
      const changed = (ln !== CURRENT_LANG);
      CURRENT_LANG = ln;

      try { localStorage.setItem("lang", ln); } catch (e) {}

      try {
        const u = new URL(window.location.href);
        u.searchParams.set("lang", ln);

        // Keep current plate in the URL as well (if user already typed it)
        const pEl = document.getElementById("plate");
        const pNow = pEl ? normalizePlate(pEl.value) : "";
        if (pNow) u.searchParams.set("plate", pNow);

        if (opts && opts.reload && changed) {
          window.location.replace(u.toString()); // full reload
          return;
        }
        history.replaceState(null, "", u.toString());
      } catch (e) {}

      try { document.documentElement.lang = ln; } catch (e) {}
      applyLangUI();
      updateLangButtons();
    }

    function updateLangButtons() {
      const bar = document.getElementById("langbar");
      if (!bar) return;
      const btns = bar.querySelectorAll("button[data-lang]");
      btns.forEach((b) => {
        const l = normLang(b.getAttribute("data-lang") || "");
        if (l === CURRENT_LANG) b.classList.add("active");
        else b.classList.remove("active");
      });
    }

    function applyLangUI() {
      const h2 = document.getElementById("titleH2");
      if (h2) h2.textContent = t("title");

      const plate = document.getElementById("plate");
      if (plate) plate.setAttribute("placeholder", t("plate_ph"));

      const btn = document.getElementById("btn");
      if (btn) btn.textContent = t("btn_check");

      const bn = document.getElementById("btnNotify");
      if (bn && bn.style.display !== "none") {
        if (bn.disabled && bn.textContent === UI.en.btn_enabled) bn.textContent = t("btn_enabled");
        else bn.textContent = t("btn_notify");
      }
    }

    function getInitialPlate() {
      try {
        const p = new URLSearchParams(window.location.search).get("plate") || "";
        const pn = normalizePlate(p);
        if (pn) return pn;
      } catch (e) {}
      try {
        return normalizePlate(localStorage.getItem("last_plate") || "");
      } catch (e) {
        return "";
      }
    }

    function setCurrentPlate(p) {
      const pn = normalizePlate(p);
      if (!pn) return;
      try { localStorage.setItem("last_plate", pn); } catch (e) {}
      try {
        const u = new URL(window.location.href);
        u.searchParams.set("plate", pn);
        u.searchParams.set("lang", CURRENT_LANG);
        history.replaceState(null, "", u.toString());
      } catch (e) {}
    }

    async function readJsonOrText(res) {
      const ct = (res.headers.get("content-type") || "").toLowerCase();
      if (ct.includes("application/json")) {
        return await res.json();
      }
      const txt = await res.text();
      return { detail: txt };
    }

    let _map = null;
    let _routeLine = null;

    function destroyMap() {
      try {
        if (_map) _map.remove();
      } catch (e) {}
      _map = null;
      _routeLine = null;
    }

    function setMapNote(msg, isErr) {
      const el = document.getElementById("mapNote");
      if (!el) return;
      el.textContent = msg || "";
      el.style.color = isErr ? "#a00000" : "";
    }

    async function renderRouteMap(plate, loc) {
      const mapDiv = document.getElementById("map");
      if (!mapDiv || typeof L === "undefined") return;

      destroyMap();

      _map = L.map("map", { zoomControl: true, scrollWheelZoom: true });
      L.tileLayer("https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png", {
        maxZoom: 19,
        attribution: "&copy; OpenStreetMap contributors",
      }).addTo(_map);

      setMapNote(t("loading_route"), false);

      try {
        const url = `${API_BASE}/api/route?plate=${encodeURIComponent(plate)}&lat=${encodeURIComponent(loc.lat)}&lon=${encodeURIComponent(loc.lon)}&ts=${encodeURIComponent(loc.ts)}&lang=${encodeURIComponent(CURRENT_LANG)}`;
        const res = await fetch(url);
        const data = await readJsonOrText(res);

        if (!res.ok) {
          setMapNote(data.detail || res.statusText, true);
          _map.setView([loc.lat, loc.lon], 10);
          setTimeout(() => { if (_map) _map.invalidateSize(); }, 80);
          return;
        }

        const pts = data.route || [];
        if (pts.length >= 2) {
          _routeLine = L.polyline(pts, { color: "#4D148C", weight: 4, opacity: 0.9 }).addTo(_map);
          _map.fitBounds(_routeLine.getBounds(), { padding: [12, 12] });
        } else {
          _map.setView([loc.lat, loc.lon], 10);
        }

        if (data.origin && data.origin.lat != null && data.origin.lon != null) {
          L.marker([data.origin.lat, data.origin.lon]).addTo(_map).bindPopup(t("origin"));
        }
        if (data.dest && data.dest.lat != null && data.dest.lon != null) {
          L.marker([data.dest.lat, data.dest.lon]).addTo(_map).bindPopup(t("destination_pin"));
        }

        setMapNote(data.note || "", false);
        setTimeout(() => { if (_map) _map.invalidateSize(); }, 80);
      } catch (e) {
        setMapNote(t("route_error") + ": " + e, true);
        try { _map.setView([loc.lat, loc.lon], 10); } catch (e2) {}
        setTimeout(() => { if (_map) _map.invalidateSize(); }, 80);
      }
    }

    function show(html, klass) {
      const out = document.getElementById("out");
      out.className = "card " + (klass || "");
      out.style.display = "block";
      out.innerHTML = html;
    }

    function getLocation() {
      return new Promise((resolve, reject) => {
        if (!navigator.geolocation) {
          reject(new Error("Geolocation not supported on this device."));
          return;
        }
        navigator.geolocation.getCurrentPosition(
          (pos) => {
            resolve({
              lat: pos.coords.latitude,
              lon: pos.coords.longitude,
              ts: Math.floor(Date.now() / 1000)
            });
          },
          (err) => reject(new Error(err.message || "Location denied.")),
          { enableHighAccuracy: true, timeout: 15000, maximumAge: 0 }
        );
      });
    }

    async function checkStatus() {
      stopDeveloperView();

      const plate = normalizePlate(document.getElementById("plate").value);
      if (!plate) return;

      if (plate === DEV_PLATE) {
        showDeveloperView();
        return;
      }

      setCurrentPlate(plate);

      destroyMap();

      show(`<div class="muted">${t("getting_location")}</div>`);

      let loc;
      try {
        loc = await getLocation();
      } catch (e) {
        destroyMap();
        show(`<b>${t("err_location")}:</b> ${e.message}<div class="muted">${t("help_location")}</div>`, "err");
        return;
      }

      show(`<div class="muted">${t("loading_status")}</div>`);

      try {
        const url = `${API_BASE}/api/status?plate=${encodeURIComponent(plate)}&lat=${encodeURIComponent(loc.lat)}&lon=${encodeURIComponent(loc.lon)}&ts=${encodeURIComponent(loc.ts)}&lang=${encodeURIComponent(CURRENT_LANG)}`;
        const res = await fetch(url);
        const data = await readJsonOrText(res);

        if (!res.ok) {
          destroyMap();
          show(`<b>${t("err_error")}:</b> ${data.detail || res.statusText}`, "err");
          document.getElementById("btnNotify").style.display = "none";
          setNotifyMsg("", "");
          return;
        }

        if (data && data.house_rules_required) {
          await showHouseRulesModal(plate);
        }

        const last = data.last_refresh || "-";

        if (!data.found) {
          destroyMap();
          show(`
            <div class="status-big">${t("no_movement")}</div>
            <div class="muted">${t("last_refresh")}: ${last}</div>
          `, "warn");
          document.getElementById("btnNotify").style.display = "none";
          setNotifyMsg("", "");
          return;
        }

        const destText = data.destination_text || "-";
        const destLink = data.destination_nav_url ? `<a href="${data.destination_nav_url}" target="_blank" rel="noopener">${destText}</a>` : destText;

        const trailerVal = (data.trailer || "").trim();
        const locVal = (data.location || "").trim();

        let extraTrailerPlace = "";
        if (locVal) {
          let placeText = locVal;

          // If location begins with "P" → show "Parking <number>"
          if (/^[Pp]/.test(locVal)) {
            let rest = locVal.slice(1).trim();
            const digits = (rest.match(/\d+/g) || []).join("");
            if (digits) {
              const num = digits.replace(/^0+/, "") || "0";
              placeText = `${t("parking")} ${num}`;
            } else if (rest) {
              placeText = `${t("parking")} ${rest}`;
            } else {
              placeText = t("parking");
            }
          }
          // If location begins with a number → show "Dock <location>"
          else if (/^\d/.test(locVal)) {
            placeText = `${t("dock")} ${locVal}`;
          }

          extraTrailerPlace = `
          <div style="margin-top:6px;"><b>${t("trailer")}:</b> ${trailerVal || "-"}</div>
          <div><b>${t("place")}:</b> ${placeText}</div>`;
        }

        show(`
          <div class="status-big">"${data.status_text}"</div>
          <hr style="border:none;border-top:1px solid #ddd;margin:12px 0;">
          <div><b>${t("destination")}:</b> ${destLink}</div>
          <div><b>${t("departure_time")}:</b> ${data.scheduled_departure || "-"}</div>
          <div><b>${t("report_office")}:</b> ${data.report_in_office_at || "-"}</div>
          ${extraTrailerPlace}
          <div class="muted" style="margin-top:8px;">${t("last_refresh")}: ${last}</div>

          <div style="margin-top:12px;"><b>${t("route_map")}:</b></div>
          <div id="map"></div>
          <div id="mapNote" class="muted" style="margin-top:6px;"></div>
        `, "ok");

        setTimeout(() => renderRouteMap(plate, loc), 0);

        if (data.push_enabled && data.vapid_public_key) {
          const bn = document.getElementById("btnNotify");
          bn.style.display = "block";
          bn.onclick = () => enableNotifications(plate, loc, data.vapid_public_key);
          bn.disabled = false;
          bn.textContent = t("btn_notify");
          bn.style.opacity = "";
          setNotifyMsg("", "");
        } else {
          document.getElementById("btnNotify").style.display = "none";
          setNotifyMsg("", "");
        }

      } catch (e) {
        destroyMap();
        show(`<b>${t("err_network")}:</b> ${e}`, "err");
        document.getElementById("btnNotify").style.display = "none";
        setNotifyMsg("", "");
      }
    }

    function urlBase64ToUint8Array(base64String) {
      const padding = '='.repeat((4 - base64String.length % 4) % 4);
      const base64 = (base64String + padding).replace(/-/g, '+').replace(/_/g, '/');
      const rawData = window.atob(base64);
      const outputArray = new Uint8Array(rawData.length);
      for (let i = 0; i < rawData.length; ++i) outputArray[i] = rawData.charCodeAt(i);
      return outputArray;
    }

    async function enableNotifications(plate, loc, vapidPublicKey) {
      try {
        const bn0 = document.getElementById("btnNotify");
        if (bn0) {
          bn0.disabled = true;
          bn0.textContent = t("btn_enabling");
          bn0.style.opacity = "0.75";
        }
        setNotifyMsg(`<div class="muted">${t("btn_enabling")}</div>`, "");

        if (!('serviceWorker' in navigator) || !('PushManager' in window)) {
          setNotifyMsg(`<b>${t("notify_not_supported")}</b><div class="muted">${t("notify_not_supported_help")}</div>`, "err");
          const bn1 = document.getElementById("btnNotify");
          if (bn1) { bn1.disabled = false; bn1.textContent = t("btn_notify"); bn1.style.opacity = ""; }
          return;
        }

        const perm = await Notification.requestPermission();
        if (perm !== 'granted') {
          setNotifyMsg(`<b>${t("notify_denied")}</b><div class="muted">${t("notify_denied_help")}</div>`, "err");
          const bn2 = document.getElementById("btnNotify");
          if (bn2) { bn2.disabled = false; bn2.textContent = t("btn_notify"); bn2.style.opacity = ""; }
          return;
        }

        const reg = await navigator.serviceWorker.register('/sw.js');

        // Wait until the service worker is activated (otherwise PushManager.subscribe can fail with "no active Service Worker")
        const sw = reg.installing || reg.waiting || reg.active;
        if (!sw) throw new Error("Service Worker registration failed.");
        await new Promise((resolve, reject) => {
          if (sw.state === 'activated') return resolve();
          const tmo = setTimeout(() => reject(new Error("Service Worker activation timeout.")), 8000);
          sw.addEventListener('statechange', () => {
            if (sw.state === 'activated') {
              clearTimeout(tmo);
              resolve();
            }
          });
        });

        const existing = await reg.pushManager.getSubscription();
        const sub = existing || await reg.pushManager.subscribe({
          userVisibleOnly: true,
          applicationServerKey: urlBase64ToUint8Array(vapidPublicKey),
        });

        const resp = await fetch(`${API_BASE}/api/subscribe?plate=${encodeURIComponent(plate)}&lat=${encodeURIComponent(loc.lat)}&lon=${encodeURIComponent(loc.lon)}&ts=${encodeURIComponent(loc.ts)}&lang=${encodeURIComponent(CURRENT_LANG)}`, {
          method: 'POST',
          headers: { 'Content-Type': 'application/json' },
          body: JSON.stringify(sub),
        });

        const data = await readJsonOrText(resp);
        if (!resp.ok) {
          setNotifyMsg(`<b>${t("notify_failed")}:</b> ${data.detail || resp.statusText}`, "err");
          const bn3 = document.getElementById("btnNotify");
          if (bn3) { bn3.disabled = false; bn3.textContent = t("btn_notify"); bn3.style.opacity = ""; }
          return;
        }

        const bn = document.getElementById("btnNotify");
        if (bn) {
          bn.disabled = true;
          bn.textContent = t("btn_enabled");
          bn.style.opacity = "0.75";
          bn.onclick = null;
        }
        setNotifyMsg(`<b>${t("notify_enabled_msg")}</b><div class="muted">${t("notify_enabled_help")}</div>`, "");
      } catch (e) {
        setNotifyMsg(`<b>${t("subscribe_error")}:</b> ${e}`, "err");
        const bn4 = document.getElementById("btnNotify");
        if (bn4) { bn4.disabled = false; bn4.textContent = t("btn_notify"); bn4.style.opacity = ""; }
      }
    }

    
    // -----------------------------
    // Developer window (DEV_PLATE)
    // -----------------------------
    let _devTimer = null;

    function stopDeveloperView() {
      try {
        if (_devTimer) clearInterval(_devTimer);
      } catch (e) {}
      _devTimer = null;
    }

    function fmtIso(iso) {
      try {
        if (!iso) return "-";
        const d = new Date(iso);
        if (!isNaN(d.getTime())) return d.toLocaleString();
      } catch (e) {}
      return String(iso || "-");
    }

    async function fetchDevSummary() {
      const url = `${API_BASE}/api/dev/summary?key=${encodeURIComponent(DEV_PLATE)}`;
      const res = await fetch(url);
      const data = await readJsonOrText(res);
      if (!res.ok) throw new Error(data.detail || res.statusText);
      return data;
    }

    async function setAdminMonitorEnabled(enabled) {
      const url = `${API_BASE}/api/dev/admin_notify?key=${encodeURIComponent(DEV_PLATE)}`;
      const res = await fetch(url, {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ enabled: !!enabled }),
      });
      const data = await readJsonOrText(res);
      if (!res.ok) throw new Error(data.detail || res.statusText);
      return data;
    }

    async function ensureAdminSubscription(vapidPublicKey) {
      if (!('serviceWorker' in navigator) || !('PushManager' in window)) {
        throw new Error(t("notify_not_supported_help"));
      }

      const perm = await Notification.requestPermission();
      if (perm !== 'granted') {
        throw new Error(t("notify_denied_help"));
      }

      const reg = await navigator.serviceWorker.register('/sw.js');

      const sw = reg.installing || reg.waiting || reg.active;
      if (!sw) throw new Error("Service Worker registration failed.");
      await new Promise((resolve, reject) => {
        if (sw.state === 'activated') return resolve();
        const tmo = setTimeout(() => reject(new Error("Service Worker activation timeout.")), 8000);
        sw.addEventListener('statechange', () => {
          if (sw.state === 'activated') {
            clearTimeout(tmo);
            resolve();
          }
        });
      });

      const existing = await reg.pushManager.getSubscription();
      const sub = existing || await reg.pushManager.subscribe({
        userVisibleOnly: true,
        applicationServerKey: urlBase64ToUint8Array(vapidPublicKey),
      });

      const resp = await fetch(`${API_BASE}/api/dev/subscribe_admin?key=${encodeURIComponent(DEV_PLATE)}&lang=${encodeURIComponent(CURRENT_LANG)}`, {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify(sub),
      });

      const data = await readJsonOrText(resp);
      if (!resp.ok) throw new Error(data.detail || resp.statusText);
      return data;
    }

    function escapeHtml(v) {
      return String(v == null ? "" : v)
        .replaceAll("&", "&amp;")
        .replaceAll("<", "&lt;")
        .replaceAll(">", "&gt;")
        .replaceAll('"', "&quot;")
        .replaceAll("'", "&#39;");
    }

    async function sendAdminPlateMessage(plate, message) {
      const resp = await fetch(`${API_BASE}/api/dev/send_message?key=${encodeURIComponent(DEV_PLATE)}`, {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ plate, message }),
      });

      const data = await readJsonOrText(resp);
      if (!resp.ok) throw new Error(data.detail || resp.statusText);
      return data;
    }

    function devRowHtml(it) {
      const bell = it.bell ? "🔔" : "";
      const st = escapeHtml(it.status_text || "-");
      const dest = escapeHtml(it.destination_text || "-");
      const dep = escapeHtml(it.scheduled_departure || "-");
      const lc = fmtIso(it.last_check || "");
      const cnt = (it.count_12h != null) ? String(it.count_12h) : "-";
      const plate = escapeHtml(it.plate || "-");

      const missing = it.movement_found ? "" : ' <span class="muted">(not found)</span>';
      const canSend = !!it.bell;
      const msgBtnLabel = canSend ? "Message" : "No push user";
      const msgBtnDisabled = canSend ? "" : " disabled";
      const msgBtnOpacity = canSend ? "1" : "0.55";
      const msgBtnCursor = canSend ? "pointer" : "not-allowed";

      return `<tr>
        <td style="padding:6px 8px; border-bottom:1px solid rgba(0,0,0,0.10); white-space:nowrap;"><b>${plate}</b>${missing}</td>
        <td style="padding:6px 8px; border-bottom:1px solid rgba(0,0,0,0.10); white-space:nowrap;">${lc}</td>
        <td style="padding:6px 8px; border-bottom:1px solid rgba(0,0,0,0.10); text-align:right;">${cnt}</td>
        <td style="padding:6px 8px; border-bottom:1px solid rgba(0,0,0,0.10); white-space:nowrap;">${dep}</td>
        <td style="padding:6px 8px; border-bottom:1px solid rgba(0,0,0,0.10);">${dest}</td>
        <td style="padding:6px 8px; border-bottom:1px solid rgba(0,0,0,0.10);">${st}</td>
        <td style="padding:6px 8px; border-bottom:1px solid rgba(0,0,0,0.10); text-align:center;">${bell}</td>
        <td style="padding:6px 8px; border-bottom:1px solid rgba(0,0,0,0.10); text-align:right; white-space:nowrap;">
          <button
            class="btn btn-secondary"
            data-dev-send="1"
            data-plate="${plate}"
            style="min-width:110px; opacity:${msgBtnOpacity}; cursor:${msgBtnCursor};"
            ${msgBtnDisabled}
          >${msgBtnLabel}</button>
        </td>
      </tr>`;
    }

    function renderDeveloperView(data) {
      const items = (data && data.items) ? data.items : [];
      const adminEnabled = !!(data && data.admin_notify_enabled);
      const adminSubscribed = !!(data && data.admin_subscribed);
      const pushEnabled = !!(data && data.push_enabled);
      const vapidKey = (data && data.vapid_public_key) ? String(data.vapid_public_key) : "";

      const btnLabel = adminEnabled ? "Admin notifications: ON" : "Admin notifications: OFF";
      const btnClass = adminEnabled ? "btn btn-primary" : "btn btn-secondary";

      let adminHint = "";
      if (!pushEnabled) {
        adminHint = `<div class="muted" style="margin-top:6px;"><b>Push disabled on server</b> (missing VAPID keys).</div>`;
      } else if (!adminSubscribed) {
        adminHint = `<div class="muted" style="margin-top:6px;">Admin is not subscribed yet (click the button to subscribe).</div>`;
      } else {
        adminHint = `<div class="muted" style="margin-top:6px;">Admin subscribed: ✅</div>`;
      }

      const rows = items.length ? items.map(devRowHtml).join("") : `<tr><td colspan="8" class="muted" style="padding:10px;">No checks in the last 12 hours.</td></tr>`;

      show(`
        <div style="display:flex; align-items:center; justify-content:space-between; gap:10px; flex-wrap:wrap;">
          <div>
            <div class="status-big">Developer monitor</div>
            <div class="muted">Plates checked in the last 12 hours</div>
          </div>
          <div class="row" style="flex:0 0 auto;">
            <button id="devRefresh" class="btn btn-secondary" style="min-width:120px;">Refresh</button>
          </div>
        </div>

        <div style="margin-top:10px;">
          <button id="btnAdminMonitor" class="${btnClass}" style="width:100%;">${btnLabel}</button>
          ${adminHint}
          <div id="devMsg" class="muted" style="margin-top:6px; display:none;"></div>
        </div>

        <div style="margin-top:12px; overflow:auto;">
          <table style="width:100%; border-collapse:collapse; font-size:14px;">
            <thead>
              <tr>
                <th style="text-align:left; padding:6px 8px; border-bottom:1px solid rgba(0,0,0,0.18);">Plate</th>
                <th style="text-align:left; padding:6px 8px; border-bottom:1px solid rgba(0,0,0,0.18);">Last check</th>
                <th style="text-align:right; padding:6px 8px; border-bottom:1px solid rgba(0,0,0,0.18);">#</th>
                <th style="text-align:left; padding:6px 8px; border-bottom:1px solid rgba(0,0,0,0.18);">Departure</th>
                <th style="text-align:left; padding:6px 8px; border-bottom:1px solid rgba(0,0,0,0.18);">Destination</th>
                <th style="text-align:left; padding:6px 8px; border-bottom:1px solid rgba(0,0,0,0.18);">Status</th>
                <th style="text-align:center; padding:6px 8px; border-bottom:1px solid rgba(0,0,0,0.18);">🔔</th>
                <th style="text-align:right; padding:6px 8px; border-bottom:1px solid rgba(0,0,0,0.18);">Message</th>
              </tr>
            </thead>
            <tbody>${rows}</tbody>
          </table>
        </div>
      `, "ok");

      // Hook buttons
      const rbtn = document.getElementById("devRefresh");
      if (rbtn) rbtn.onclick = () => refreshDeveloperView();

      const abtn = document.getElementById("btnAdminMonitor");
      if (abtn) {
        abtn.onclick = async () => {
          const devMsg = document.getElementById("devMsg");
          const setDevMsg = (txt, isErr) => {
            if (!devMsg) return;
            devMsg.style.display = txt ? "block" : "none";
            devMsg.textContent = txt || "";
            devMsg.style.color = isErr ? "#a00000" : "";
          };

          try {
            const turnOn = !adminEnabled;

            if (turnOn) {
              if (!pushEnabled || !vapidKey) {
                throw new Error("Push is disabled on server (missing VAPID keys).");
              }
              // Ensure admin browser has a subscription
              if (!adminSubscribed) {
                setDevMsg("Subscribing admin push…", false);
                await ensureAdminSubscription(vapidKey);
              }
              setDevMsg("Turning admin monitor ON…", false);
              await setAdminMonitorEnabled(true);
            } else {
              setDevMsg("Turning admin monitor OFF…", false);
              await setAdminMonitorEnabled(false);
            }

            await refreshDeveloperView();
            setDevMsg("", false);
          } catch (e) {
            setDevMsg(String(e && e.message ? e.message : e), true);
          }
        };
      }

      const devMsg = document.getElementById("devMsg");
      const setDevMsg = (txt, isErr) => {
        if (!devMsg) return;
        devMsg.style.display = txt ? "block" : "none";
        devMsg.textContent = txt || "";
        devMsg.style.color = isErr ? "#a00000" : "";
      };

      document.querySelectorAll("button[data-dev-send='1']").forEach((btn) => {
        btn.onclick = async () => {
          const plate = String(btn.getAttribute("data-plate") || "").trim();
          if (!plate) return;

          const msg = window.prompt(`Send message to ${plate}:`, "");
          if (msg == null) return;

          const cleanMsg = String(msg || "").trim();
          if (!cleanMsg) {
            setDevMsg("Message is empty.", true);
            return;
          }

          const oldLabel = btn.textContent || "Message";
          btn.disabled = true;
          btn.textContent = "Sending…";

          try {
            await sendAdminPlateMessage(plate, cleanMsg);
            setDevMsg(`Message sent to ${plate}.`, false);
            await refreshDeveloperView();
          } catch (e) {
            setDevMsg(String(e && e.message ? e.message : e), true);
            btn.disabled = false;
            btn.textContent = oldLabel;
          }
        };
      });
    }

    async function refreshDeveloperView() {
      try {
        const data = await fetchDevSummary();
        renderDeveloperView(data);
      } catch (e) {
        show(`<b>Developer monitor error:</b> ${String(e && e.message ? e.message : e)}`, "err");
      }
    }

    function showDeveloperView() {
      destroyMap();
      try { document.getElementById("btnNotify").style.display = "none"; } catch (e) {}
      setNotifyMsg("", "");
      show(`<div class="muted">Loading developer monitor…</div>`, "ok");

      refreshDeveloperView();
      stopDeveloperView();
      _devTimer = setInterval(refreshDeveloperView, 10000);
    }


    document.getElementById("btn").addEventListener("click", checkStatus);
    document.getElementById("plate").addEventListener("keydown", (e) => {
      if (e.key === "Enter") checkStatus();
    });

    // Language buttons
    (function initLang() {
      setCurrentLang(getInitialLang(), { reload: false });

      const bar = document.getElementById("langbar");
      if (bar) {
        bar.addEventListener("click", (ev) => {
          const btn = ev.target && ev.target.closest ? ev.target.closest("button[data-lang]") : null;
          if (!btn) return;
          const l = btn.getAttribute("data-lang") || "en";
          setCurrentLang(l, { reload: true });
        });
      }
    })();

    // Restore plate from URL or last usage and auto-run once
    (function initPlate() {
      const p = getInitialPlate();
      if (p) {
        document.getElementById("plate").value = p;
        setTimeout(() => { checkStatus(); }, 50);
      } else {
        applyLangUI();
        updateLangButtons();
      }
    })();
  </script>
</body>
</html>"""



SERVICE_WORKER_JS = r"""
self.addEventListener('install', function(event) {
  // Activate immediately on first load
  self.skipWaiting();
});

self.addEventListener('activate', function(event) {
  // Take control without requiring a reload
  event.waitUntil(clients.claim());
});

self.addEventListener('push', function(event) {
  let data = {};
  try { data = event.data.json(); } catch (e) { data = { title: 'Update', body: event.data && event.data.text() }; }
  const title = data.title || 'Status update';
  const options = { body: data.body || '', icon: '/static/icon-192.png', badge: '/static/icon-192.png', data: { url: (data.url || '/') } };
  event.waitUntil(self.registration.showNotification(title, options));
});

self.addEventListener('notificationclick', function(event) {
  event.notification.close();
  const url = (event.notification && event.notification.data && event.notification.data.url) || '/';
  event.waitUntil(clients.openWindow(url));
});
"""
